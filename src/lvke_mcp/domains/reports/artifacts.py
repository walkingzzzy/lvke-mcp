"""报告域交付工件 —— MCP 自有实现（零外部依赖）。

为既有交付工件逻辑的无签审子集：被引 7 符号
（``create_draft_export`` / ``create_deliverable_artifact`` / ``_artifact_root`` /
``record_internal_release`` / ``list_artifacts`` / ``get_artifact`` /
``read_artifact_download``）及其闭包，仅改 import 路径、存储根与降级语义，
不重写业务逻辑：

- doc_service → ``lvke_mcp.domains.reports.doc_service``（MCP 自有）
- professional_review（专业复核）无 MCP 等价域 → 无签审子集：
  ``_inspect_review`` 恒返回 none 状态，正式工件门禁不再校验专业复核批准记录
- report_artifacts._path → 读工作区根 ``{name}.json``；governed 快照裁剪为
  MCP 域内实际存在的 ``evidence_pack``（读 ``lvke_data_analysis.EVIDENCE_STORE``）
- source_files_api.source_basis_snapshot → 读 ``lvke_data_acquisition.SOURCE_STORE``
  构造同 schema 快照（存储不可用 fail-closed 进入 basis 指纹）
- finance 审计/门禁 → ``lvke_mcp.domains.finance.run_store`` /
  ``lvke_mcp.domains.finance.gate``；MCP 边界无持久化 finance_binding，
  绑定退化为最新 run（与 MCP gate 语义一致），门禁显式传 expected_run_id
- docx_fonts.normalize_docx_fonts → ``lvke_mcp.domains.reports.docx_fonts``
- 统一审查绑定（``_require_unified_release_review``）剔除：内部发布记录不
  绑定统一审查 review（release 权限门禁由审查域自身承载）
- 删 tenant 形参（MCP 无租户边界；tenant_scope_hash 固定为 "local"，与
  hermes 默认租户同值，state 文件兼容）
"""

from __future__ import annotations

import copy
import hashlib
import io
import json
import mimetypes
import os
import re
import shutil
import threading
import uuid
from collections.abc import Mapping, Sequence
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any

from filelock import FileLock

from lvke_mcp.domains.reports import doc_service


SCHEMA_VERSION = "deliverable_artifacts.v1"
MANIFEST_SCHEMA_VERSION = "deliverable_manifest.v1"
BASIS_SCHEMA_VERSION = "deliverable_basis.v1"
DEFAULT_TEMPLATE_VERSION = "feasibility-report.v1"
DRAFT_MARKER = "专家参考稿/内部复核·非报批终稿"

_SAFE_WORKSPACE_ID = re.compile(r"^[A-Za-z0-9_.-]{1,64}$")
_SAFE_ARTIFACT_ID = re.compile(r"^deliverable_[0-9a-f]{32}$")
_SHA256_EVIDENCE_HASH = re.compile(r"^sha256:[0-9a-f]{64}$")
_SAFE_REVISION_ID = re.compile(r"^[A-Za-z0-9_.-]{1,160}$")
_SAFE_TEMPLATE_VERSION = re.compile(r"^[A-Za-z0-9_.:-]{1,120}$")
_SAFE_OPERATION_ID = re.compile(r"^[A-Za-z0-9_.:-]{1,200}$")
_LOCK = threading.RLock()
# MCP 域内 governed 工件只有 evidence_pack（数据链 EVIDENCE_STORE）；
# fact_pack / research_decisions / appendix_manifest 为 hermes 独有概念，
# MCP 无对应持久化，不参与工件依据指纹。
_GOVERNED_SNAPSHOTS = ("evidence_pack",)
_SUPPORT_SUFFIXES = {
    ".csv",
    ".docx",
    ".jpeg",
    ".jpg",
    ".pdf",
    ".png",
    ".xls",
    ".xlsx",
}
_VERIFIED_APPENDIX_STATES = {"approved", "ready", "reviewed", "verified"}
_INTERNAL_ACTOR = {
    "actor_id": "system:deliverable-artifacts",
    "authenticated": True,
    "display_name": "Deliverable artifact integrity service",
    "auth_method": "internal",
}
# MCP 无租户边界：scope hash 固定为 hermes 默认租户 "local"，
# 使 state.json 的 tenant_scope_hash 字段与 hermes 默认租户产物同值。
_LOCAL_TENANT_SCOPE = "sha256:" + hashlib.sha256(b"local").hexdigest()


class DeliverableArtifactError(RuntimeError):
    """Machine-readable service error."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _canonical_hash(value: Any) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return "sha256:" + hashlib.sha256(payload).hexdigest()


def _bytes_hash(value: bytes) -> str:
    return "sha256:" + hashlib.sha256(value).hexdigest()


def _file_hash(path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    try:
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
                size += len(chunk)
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_FILE_UNREADABLE",
            "工件文件不可读取",
            details={"filename": path.name, "error": type(exc).__name__},
        ) from exc
    return "sha256:" + digest.hexdigest(), size


def _without_volatile_timestamps(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            str(key): _without_volatile_timestamps(item)
            for key, item in value.items()
            if str(key) != "computed_at"
        }
    if isinstance(value, list):
        return [_without_volatile_timestamps(item) for item in value]
    return value


def _validate_workspace_id(workspace_id: str) -> str:
    value = str(workspace_id or "").strip()
    if not _SAFE_WORKSPACE_ID.fullmatch(value):
        raise DeliverableArtifactError("INVALID_WORKSPACE_ID", "工作区 id 不合法")
    return value


def _validate_artifact_id(artifact_id: str) -> str:
    value = str(artifact_id or "").strip()
    if not _SAFE_ARTIFACT_ID.fullmatch(value):
        raise DeliverableArtifactError("INVALID_ARTIFACT_ID", "工件 id 不合法")
    return value


def _validate_template_version(template_version: str) -> str:
    value = str(template_version or DEFAULT_TEMPLATE_VERSION).strip()
    if not _SAFE_TEMPLATE_VERSION.fullmatch(value):
        raise DeliverableArtifactError(
            "INVALID_TEMPLATE_VERSION", "报告模板版本不合法",
        )
    return value


def _workspace_root(workspace_id: str) -> Path:
    return doc_service._workspace_root(_validate_workspace_id(workspace_id))  # noqa: SLF001


def _require_workspace(workspace_id: str) -> Path:
    root = _workspace_root(workspace_id)
    if not (root / "workspace_meta.json").is_file():
        raise DeliverableArtifactError("WORKSPACE_NOT_FOUND", "工作区不存在")
    return root


def _service_root(workspace_id: str) -> Path:
    return _workspace_root(workspace_id) / "deliverable_artifacts"


def _artifacts_root(workspace_id: str) -> Path:
    return _service_root(workspace_id) / "artifacts"


def _artifact_root(
    workspace_id: str,
    artifact_id: str,
) -> Path:
    return _artifacts_root(workspace_id) / _validate_artifact_id(artifact_id)


def _state_path(workspace_id: str) -> Path:
    return _service_root(workspace_id) / "state.json"


def _empty_state(workspace_id: str) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "workspace_id": workspace_id,
        "tenant_scope_hash": _LOCAL_TENANT_SCOPE,
        "current": {"draft": "", "formal": ""},
        "created_at": "",
        "updated_at": "",
        "artifacts": {},
        "history": [],
    }


def _strict_read_json(path: Path, *, code: str, label: str) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise DeliverableArtifactError(code, f"{label}不存在") from exc
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise DeliverableArtifactError(
            code,
            f"{label}损坏或不可读",
            details={"error": type(exc).__name__},
        ) from exc


def _read_state(workspace_id: str) -> dict[str, Any]:
    path = _state_path(workspace_id)
    if not path.exists():
        return _empty_state(workspace_id)
    value = _strict_read_json(
        path, code="ARTIFACT_STATE_CORRUPT", label="交付工件状态",
    )
    if not isinstance(value, dict):
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "交付工件状态必须是对象",
        )
    if value.get("schema_version") != SCHEMA_VERSION:
        raise DeliverableArtifactError(
            "ARTIFACT_SCHEMA_UNSUPPORTED",
            "交付工件状态 schema 版本不受支持",
            details={"schema_version": value.get("schema_version")},
        )
    if str(value.get("workspace_id") or "") != workspace_id:
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "交付工件状态与工作区不匹配",
        )
    stored_scope = str(value.get("tenant_scope_hash") or "")
    if stored_scope and stored_scope != _LOCAL_TENANT_SCOPE:
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "交付工件状态与租户边界不匹配",
        )
    artifacts = value.get("artifacts")
    history = value.get("history")
    current = value.get("current")
    if not isinstance(artifacts, dict) or any(
        not isinstance(item, dict) for item in artifacts.values()
    ):
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "交付工件索引格式错误",
        )
    if not isinstance(history, list) or any(not isinstance(item, dict) for item in history):
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "交付工件历史格式错误",
        )
    if not isinstance(current, dict):
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_CORRUPT", "当前交付工件指针格式错误",
        )
    return value


def _write_json_atomic(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f"{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temp.open("w", encoding="utf-8", newline="\n") as handle:
            json.dump(value, handle, ensure_ascii=False, indent=2, default=str)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp, path)
    except (OSError, TypeError, ValueError) as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_STATE_WRITE_FAILED",
            "交付工件状态保存失败",
            details={"path": str(path), "error": type(exc).__name__},
        ) from exc
    finally:
        temp.unlink(missing_ok=True)


def _write_bytes(path: Path, value: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("xb") as handle:
            handle.write(value)
            handle.flush()
            os.fsync(handle.fileno())
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_GENERATION_FAILED",
            "交付工件文件写入失败",
            details={"filename": path.name, "error": type(exc).__name__},
        ) from exc


def _write_json_file(path: Path, value: Any) -> None:
    payload = json.dumps(
        value, ensure_ascii=False, indent=2, default=str,
    ).encode("utf-8") + b"\n"
    _write_bytes(path, payload)


# ── 工作区根键值工件（原 report_artifacts.load/save/bind_finance_run 的无 tenant 版）──
# MCP 无租户边界：路径恒为 ``workspace_root/{name}.json``（默认租户语义），
# 读写在主线程锁下进行（写失败静默降级，与原 save 的 warning 语义一致）。


def load(workspace_id: str, name: str, default: Any = None) -> Any:
    """读工作区根 ``{name}.json`` 键值工件；文件缺失返回 ``default``。"""
    path = _workspace_root(workspace_id) / f"{name}.json"
    if not path.is_file():
        return default
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, ValueError):
        return default


def save(workspace_id: str, name: str, obj: Any) -> None:
    """写工作区根 ``{name}.json`` 键值工件（写失败静默降级，不阻断调用链）。"""
    try:
        _write_json_atomic(_workspace_root(workspace_id) / f"{name}.json", obj)
    except DeliverableArtifactError:
        pass


def bind_finance_run(
    workspace_id: str,
    run_id: str,
    *,
    section: str = "",
    fin: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """把正文/附表绑定到指定 finance_run_id（P1），写入 ``finance_binding.json``。"""
    if not run_id:
        return {}
    payload: dict[str, Any] = {
        "workspace_id": workspace_id,
        "finance_run_id": str(run_id),
        "section": section or "",
        "bound_at": _now(),
    }
    if isinstance(fin, Mapping):
        # Keep the legacy run fields while allowing a formal acquisition
        # artifact to bind the report to the exact immutable run/spec/fact
        # revision that produced it.  ``None`` is retained intentionally: it
        # makes an incomplete binding observable to the publish gate instead
        # of silently falling back to another revision.
        for key in (
            "input_hash",
            "spec_hash",
            "spec_id",
            "fact_revision",
            "spec_snapshot_hash",
            "evidence_binding_version",
            "evidence_binding_hash",
            "model_version",
            "template_version",
            "assurance_level",
            "review_status",
            "artifact_id",
            "artifact_job_id",
            "artifact_status",
            "report_data_hash",
            "binding_kind",
            "acquisition_tables_package_id",
            "acquisition_tables_basis_hash",
        ):
            if key in fin:
                payload[key] = fin.get(key)
    save(workspace_id, "finance_binding", payload)
    return payload


@contextmanager
def _state_guard(workspace_id: str):
    root = _service_root(workspace_id)
    root.mkdir(parents=True, exist_ok=True)
    with _LOCK, FileLock(str(root / "state.lock"), timeout=30):
        yield


def _validated_actor(actor: Mapping[str, Any] | None) -> dict[str, Any]:
    """Validate and sanitize an authenticated actor without persisting tokens.

    MCP 版本地实现（hermes 委托 professional_review.validate_authenticated_actor；
    MCP 无专业复核域，按同语义校验）。缺省用内部服务身份。
    """
    if actor is None:
        actor = _INTERNAL_ACTOR
    if not isinstance(actor, Mapping):
        raise DeliverableArtifactError(
            "AUTHENTICATION_REQUIRED", "工件操作必须提供已认证身份",
        )
    actor_id = str(actor.get("actor_id") or "").strip()
    if actor.get("authenticated") is not True or not actor_id:
        raise DeliverableArtifactError(
            "AUTHENTICATION_REQUIRED", "工件操作必须提供已认证身份",
        )
    if len(actor_id) > 160 or any(ord(char) < 32 for char in actor_id):
        raise DeliverableArtifactError("INVALID_ACTOR", "操作人标识不合法")
    return {
        "actor_id": actor_id,
        "authenticated": True,
        "display_name": str(actor.get("display_name") or "").strip()[:160],
        "auth_method": str(actor.get("auth_method") or "").strip()[:80],
    }


def _internal_release_integrity_material(release: Mapping[str, Any]) -> dict[str, Any]:
    """Return the stable internal-release envelope, excluding later signoff data."""

    return {
        str(key): copy.deepcopy(value)
        for key, value in release.items()
        if str(key) not in {"professional_signoff", "release_integrity_hash"}
    }


def _document_snapshot(workspace_id: str) -> tuple[dict[str, Any], str, dict[str, Any]]:
    root = _workspace_root(workspace_id)
    meta_path = root / "workspace_meta.json"
    if not meta_path.is_file():
        raise DeliverableArtifactError("WORKSPACE_NOT_FOUND", "工作区不存在")
    meta = _strict_read_json(
        meta_path, code="WORKSPACE_STATE_CORRUPT", label="工作区元数据",
    )
    if not isinstance(meta, dict):
        raise DeliverableArtifactError(
            "WORKSPACE_STATE_CORRUPT", "工作区元数据必须是对象",
        )
    revision_id = str(meta.get("current_revision_id") or "")
    if not _SAFE_REVISION_ID.fullmatch(revision_id):
        raise DeliverableArtifactError(
            "DOCUMENT_REVISION_INVALID", "当前文档修订 id 缺失或不合法",
        )
    report_path = root / "revisions" / revision_id / "report.md"
    try:
        content = report_path.read_text(encoding="utf-8")
    except FileNotFoundError as exc:
        raise DeliverableArtifactError(
            "DOCUMENT_REVISION_NOT_FOUND", "当前文档修订正文不存在",
        ) from exc
    except (OSError, UnicodeError) as exc:
        raise DeliverableArtifactError(
            "DOCUMENT_REVISION_UNREADABLE", "当前文档修订正文不可读",
        ) from exc
    metadata_material = {
        "title": str(meta.get("title") or ""),
        "report_type": str(meta.get("report_type") or ""),
        "doc_kind": str(meta.get("doc_kind") or ""),
        "requirement": copy.deepcopy(meta.get("requirement") or {}),
        "cover": copy.deepcopy(meta.get("cover") or {}),
        "current_revision_id": revision_id,
        "workspace_version": meta.get("workspace_version"),
    }
    snapshot = {
        "revision_id": revision_id,
        "content_hash": _bytes_hash(content.encode("utf-8")),
        "content_bytes": len(content.encode("utf-8")),
        "metadata_hash": _canonical_hash(metadata_material),
    }
    return snapshot, content, meta


def _evidence_pack_snapshot(workspace_id: str) -> tuple[dict[str, Any], Any]:
    """MCP 等价于 hermes ``evidence_pack.json`` 工件：读数据链 EVIDENCE_STORE。

    快照 hash 用最新记录的服务端固化 ``content_hash``（内容变化即变化）；
    存储不可用或空按缺失处理（present=False，进入 basis 指纹）。
    """
    try:
        from lvke_mcp.servers.lvke_data_analysis.service import EVIDENCE_STORE

        records = EVIDENCE_STORE.list(workspace_id) or []
    except Exception as exc:  # noqa: BLE001 - 数据链不可用按缺失处理
        return {
            "present": False,
            "hash": _canonical_hash({
                "artifact": "evidence_pack",
                "present": False,
                "error": type(exc).__name__,
            }),
        }, None
    records = [r for r in records if isinstance(r, dict)]
    if not records:
        return {
            "present": False,
            "hash": _canonical_hash({"artifact": "evidence_pack", "present": False}),
        }, None
    latest = sorted(
        records,
        key=lambda row: str(row.get("created_at") or ""),
        reverse=True,
    )[0]
    record = dict(latest)
    hash_value = str(record.get("content_hash") or "")
    if not hash_value:
        hash_value = _canonical_hash(record)
    payload = record.get("payload")
    return (
        {"present": True, "hash": hash_value},
        payload if isinstance(payload, dict) else {},
    )


def _json_snapshot(
    workspace_id: str,
    name: str,
) -> tuple[dict[str, Any], Any]:
    if name == "evidence_pack":
        return _evidence_pack_snapshot(workspace_id)
    path = _workspace_root(workspace_id) / f"{name}.json"
    if not path.exists():
        return {
            "present": False,
            "hash": _canonical_hash({"artifact": name, "present": False}),
        }, None
    try:
        raw = path.read_bytes()
        value = json.loads(raw.decode("utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        try:
            raw = path.read_bytes()
        except OSError:
            raw = b""
        return {
            "present": True,
            "hash": _bytes_hash(raw),
            "error": f"{name}_corrupt:{type(exc).__name__}",
        }, None
    return {"present": True, "hash": _canonical_hash(value)}, value


def _source_basis_snapshot(workspace_id: str) -> dict[str, Any]:
    """MCP 等价于 hermes ``source_files_api.source_basis_snapshot``。

    读 ``lvke_data_acquisition.SOURCE_STORE`` 的 source_snapshot 记录，
    构造同 schema（source_basis.v1 / files / jobs / hash）的稳定快照；
    存储不可用时 fail-closed（error 进入 basis 指纹）。
    """
    unavailable = ""
    try:
        from lvke_mcp.servers.lvke_data_acquisition.service import SOURCE_STORE

        records = SOURCE_STORE.list(workspace_id) or []
    except Exception as exc:  # noqa: BLE001 - 资料链不可用 fail-closed
        records = []
        unavailable = type(exc).__name__
    files = []
    for row in records:
        if not isinstance(row, dict):
            continue
        files.append({
            "object_id": str(row.get("object_id") or ""),
            "content_hash": str(row.get("content_hash") or ""),
            "basis_hash": str(row.get("basis_hash") or ""),
        })
    files.sort(key=lambda item: str(item.get("object_id") or ""))
    material = {
        "schema_version": "source_basis.v1",
        "workspace_id": workspace_id,
        "files": files,
        "jobs": [],
    }
    if unavailable:
        material["error"] = unavailable
    encoded = json.dumps(
        material, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return {
        **material,
        "file_count": len(files),
        "job_count": 0,
        "hash": "sha256:" + hashlib.sha256(encoded).hexdigest(),
    }


def _fresh_readiness(workspace_id: str) -> dict[str, Any]:
    try:
        from lvke_mcp.domains.reports import readiness as report_readiness

        value = report_readiness.build_readiness(workspace_id, persist=False)
    except Exception as exc:  # noqa: BLE001 - captured as a fail-closed basis error
        return {
            "workspace_id": workspace_id,
            "publishable": False,
            "blocking_issues": ["readiness_compute_failed"],
            "blockers": [{
                "code": "readiness_compute_failed",
                "message": f"发布就绪度计算失败: {type(exc).__name__}",
            }],
            "error": type(exc).__name__,
        }
    if not isinstance(value, dict):
        return {
            "workspace_id": workspace_id,
            "publishable": False,
            "blocking_issues": ["readiness_invalid_result"],
            "blockers": [{
                "code": "readiness_invalid_result",
                "message": "发布就绪度计算结果格式错误",
            }],
            "error": "invalid_result",
        }
    return value


def _inspect_review(workspace_id: str) -> dict[str, Any]:
    """无签审子集：MCP 无专业复核域，恒返回 none 状态。

    basis 的 professional_review 摘要保持 schema 形状（空记录），
    hermes 的批准记录校验在 ``_assert_formal_basis`` 中已裁剪。
    """
    return {
        "status": "none",
        "basis_matches": False,
        "review_id": "",
    }


def _load_finance_run(workspace_id: str, run_id: str) -> dict[str, Any] | None:
    if not run_id or run_id.startswith("acqrun_"):
        return None
    try:
        from lvke_mcp.domains.finance import run_store

        value = run_store.load_run(workspace_id, run_id)
    except Exception as exc:  # noqa: BLE001 - error is bound into the basis
        return {"_load_error": type(exc).__name__, "run_id": run_id}
    return value if isinstance(value, dict) and value else None


def _strict_finance_gate(
    workspace_id: str,
    *,
    expected_run_id: str = "",
) -> dict[str, Any]:
    """Run the authoritative finance publish gate even without finance evidence."""

    try:
        from lvke_mcp.domains.finance import gate as finance_gate

        value = finance_gate.assert_publish_finance_binding(
            workspace_id,
            strict=True,
            expected_run_id=expected_run_id,
        )
    except Exception as exc:  # noqa: BLE001 - normalized as a blocking result
        return {
            "ok": False,
            "blockers": [{
                "code": "finance_publish_gate_failed",
                "message": f"财务发布门禁执行失败: {type(exc).__name__}",
            }],
            "error": type(exc).__name__,
        }
    if not isinstance(value, dict):
        return {
            "ok": False,
            "blockers": [{
                "code": "finance_publish_gate_invalid",
                "message": "财务发布门禁返回格式错误",
            }],
            "error": "invalid_result",
        }
    return value


def _capture_basis(
    workspace_id: str,
    *,
    template_version: str,
) -> tuple[dict[str, Any], str, dict[str, Any]]:
    document, content, meta = _document_snapshot(workspace_id)
    readiness = _without_volatile_timestamps(
        _fresh_readiness(workspace_id)
    )
    review = _inspect_review(workspace_id)
    sources = _source_basis_snapshot(workspace_id)

    artifacts: dict[str, dict[str, Any]] = {}
    artifact_values: dict[str, Any] = {}
    for name in _GOVERNED_SNAPSHOTS:
        summary, value = _json_snapshot(
            workspace_id,
            name,
        )
        artifacts[name] = summary
        artifact_values[name] = value
    appendix_files = _appendix_files_snapshot(
        workspace_id, artifact_values.get("appendix_manifest"),
    )

    binding_summary, binding_value = _json_snapshot(
        workspace_id,
        "finance_binding",
    )
    binding = binding_value if isinstance(binding_value, dict) else {}
    run_id = str(binding.get("finance_run_id") or "")
    if not run_id:
        # MCP 边界无持久化 finance_binding：绑定退化为最新 run（MCP gate 语义）。
        try:
            from lvke_mcp.domains.finance import run_store

            run_id = str((run_store.latest_run(workspace_id) or {}).get("run_id") or "")
        except Exception:  # noqa: BLE001 - 无 run 时按未绑定处理
            run_id = ""
    run = _load_finance_run(workspace_id, run_id)
    finance = {
        "binding": binding_summary,
        "binding_snapshot": copy.deepcopy(binding),
        "run_id": run_id,
        "run_kind": (
            "asset_acquisition" if run_id.startswith("acqrun_")
            else "feasibility_finance" if run_id else "none"
        ),
        "run_present": isinstance(run, dict) and not run.get("_load_error"),
        "run_hash": _canonical_hash(
            run if isinstance(run, dict) else {"run_id": run_id, "present": False}
        ),
        "run_snapshot": copy.deepcopy(run) if isinstance(run, dict) else None,
        "publish_gate": _without_volatile_timestamps(
            _strict_finance_gate(workspace_id, expected_run_id=run_id)
        ),
    }
    review_summary = {
        "review_id": review.get("review_id"),
        "status": review.get("status"),
        "recorded_decision": review.get("recorded_decision"),
        "basis_matches": review.get("basis_matches") is True,
        "basis_fingerprint": review.get("basis_fingerprint"),
        "current_basis_fingerprint": review.get("current_basis_fingerprint"),
        "error": review.get("error"),
    }
    meta_doc_kind = str(meta.get("doc_kind") or "")
    material = {
        "schema_version": BASIS_SCHEMA_VERSION,
        "workspace_id": workspace_id,
        "tenant_scope_hash": _LOCAL_TENANT_SCOPE,
        "workspace_version": meta.get("workspace_version"),
        "report_type": str(meta.get("report_type") or ""),
        "doc_kind": meta_doc_kind or doc_service.DEFAULT_DOC_KIND,
        "template_version": template_version,
        "document": document,
        "sources": sources,
        "readiness": {
            "hash": _canonical_hash(readiness),
            "snapshot": readiness,
        },
        "professional_review": review_summary,
        "finance": finance,
        "artifacts": artifacts,
        "appendix_files": appendix_files,
    }
    basis = {**material, "fingerprint": _canonical_hash(material)}
    context = {
        "meta": meta,
        "artifact_values": artifact_values,
    }
    return basis, content, context


def _basis_problem(code: str, message: str, **details: Any) -> dict[str, Any]:
    return {"code": code, "message": message, "details": details}


def _readiness_blockers(readiness: dict[str, Any]) -> list[dict[str, Any]]:
    blockers: list[dict[str, Any]] = []
    for raw in readiness.get("blockers") or []:
        if isinstance(raw, dict):
            blockers.append({
                "code": str(raw.get("code") or "readiness_blocker"),
                "message": str(raw.get("message") or "发布就绪度存在阻断项"),
                "details": copy.deepcopy(raw.get("details") or {}),
            })
        else:
            blockers.append({
                "code": "readiness_blocker",
                "message": str(raw),
                "details": {},
            })
    known = {item["code"] for item in blockers}
    for code in readiness.get("blocking_issues") or []:
        value = str(code or "readiness_blocker")
        if value not in known:
            blockers.append({"code": value, "message": value, "details": {}})
            known.add(value)
    if readiness.get("error") and "readiness_error" not in known:
        blockers.append({
            "code": "readiness_error",
            "message": "发布就绪度计算存在错误",
            "details": {"error": readiness.get("error")},
        })
    return blockers


def _assert_formal_basis(basis: dict[str, Any], context: dict[str, Any]) -> None:
    """无签审子集：hermes 版的专业复核批准/依据指纹比对段已裁剪（MCP 无该域）。"""
    if basis.get("doc_kind") != "feasibility" or basis.get("report_type") == "asset_acquisition":
        raise DeliverableArtifactError(
            "FORMAL_ARTIFACT_TYPE_UNSUPPORTED",
            "通用正式工件仅支持非资产收购可行性研究报告",
            details={
                "doc_kind": basis.get("doc_kind"),
                "report_type": basis.get("report_type"),
            },
        )

    readiness = (basis.get("readiness") or {}).get("snapshot") or {}
    blockers = _readiness_blockers(readiness)
    if readiness.get("publishable") is not True or blockers:
        raise DeliverableArtifactError(
            "FORMAL_READINESS_BLOCKED",
            "当前发布就绪度存在阻断项，不能生成正式工件",
            details={"blockers": blockers, "readiness": copy.deepcopy(readiness)},
        )

    problems: list[dict[str, Any]] = []

    current_sources = basis.get("sources") or {}
    if current_sources.get("error"):
        problems.append(_basis_problem(
            "SOURCE_BASIS_UNAVAILABLE",
            "原始资料快照不可用",
            error=current_sources.get("error"),
        ))

    for name in _GOVERNED_SNAPSHOTS:
        current = (basis.get("artifacts") or {}).get(name) or {}
        if current.get("present") is not True or current.get("error"):
            problems.append(_basis_problem(
                "GOVERNED_SNAPSHOT_INCOMPLETE",
                f"{name} 快照缺失或损坏",
                artifact=name,
                error=current.get("error"),
            ))
    for appendix_file in basis.get("appendix_files") or []:
        if appendix_file.get("ok") is not True:
            problems.append(_basis_problem(
                "GOVERNED_APPENDIX_FILE_INCONSISTENT",
                "附表清单声明的已就绪文件不可用或哈希不一致",
                appendix_id=appendix_file.get("appendix_id"),
                source=appendix_file.get("source"),
                error=appendix_file.get("error"),
            ))

    finance = basis.get("finance") or {}
    run = finance.get("run_snapshot")
    binding = finance.get("binding_snapshot") or {}
    run_id = str(finance.get("run_id") or "")
    if str(binding.get("binding_kind") or "") == "asset_acquisition":
        problems.append(_basis_problem(
            "FINANCE_BINDING_KIND_UNSUPPORTED",
            "通用可研正式工件不能使用资产收购财务绑定",
            binding_kind=binding.get("binding_kind"),
        ))
    if finance.get("run_kind") != "feasibility_finance" or not run_id:
        problems.append(_basis_problem(
            "FINANCE_BINDING_REQUIRED",
            "正式工件必须绑定非资产收购财务 run",
            run_id=run_id,
            run_kind=finance.get("run_kind"),
        ))
    if not isinstance(run, dict) or run.get("_load_error"):
        problems.append(_basis_problem(
            "FINANCE_RUN_UNAVAILABLE",
            "finance_binding 指向的财务 run 不存在或不可读",
            run_id=run_id,
            error=(run or {}).get("_load_error") if isinstance(run, dict) else "not_found",
        ))
    elif str(run.get("review_status") or "") != "approved":
        problems.append(_basis_problem(
            "FINANCE_RUN_NOT_APPROVED",
            "finance_binding 指向的财务 run 尚未批准",
            run_id=run_id,
            review_status=run.get("review_status"),
        ))
    elif str(run.get("workspace_id") or basis.get("workspace_id") or "") != basis.get("workspace_id"):
        problems.append(_basis_problem(
            "FINANCE_RUN_WORKSPACE_MISMATCH",
            "财务 run 与工作区不匹配",
            run_id=run_id,
        ))
    else:
        for field in (
            "input_hash",
            "spec_hash",
            "table_bundle_hash",
            "manifest_hash",
            "template_version",
        ):
            bound = binding.get(field)
            current = run.get(field)
            if bound not in (None, "") and bound != current:
                problems.append(_basis_problem(
                    "FINANCE_BINDING_HASH_MISMATCH",
                    "财务绑定字段与批准 run 不一致",
                    field=field,
                    expected=bound,
                    actual=current,
                ))

    finance_gate = finance.get("publish_gate") or {}
    if finance_gate.get("ok") is not True or finance_gate.get("blockers"):
        problems.append(_basis_problem(
            "FINANCE_PUBLISH_GATE_BLOCKED",
            "财务正式发布门禁未通过",
            blockers=copy.deepcopy(finance_gate.get("blockers") or []),
            error=finance_gate.get("error"),
        ))
    if str(finance_gate.get("bound_run_id") or "") != run_id:
        problems.append(_basis_problem(
            "FINANCE_GATE_BOUND_RUN_MISMATCH",
            "财务门禁返回的绑定 run 与工件依据不一致",
            expected=run_id,
            actual=finance_gate.get("bound_run_id"),
        ))
    if str(finance_gate.get("approved_run_id") or "") != run_id:
        problems.append(_basis_problem(
            "FINANCE_GATE_APPROVED_RUN_MISMATCH",
            "财务门禁返回的批准 run 与工件依据不一致",
            expected=run_id,
            actual=finance_gate.get("approved_run_id"),
        ))

    if problems:
        raise DeliverableArtifactError(
            "FORMAL_BASIS_INCONSISTENT",
            "正式工件的文档、财务或治理快照不一致",
            details={"problems": problems},
        )


def _marker_markdown(
    content: str,
    readiness: dict[str, Any],
    *,
    additional_blockers: Sequence[dict[str, Any]] = (),
) -> tuple[str, dict[str, Any]]:
    blockers = [
        *_readiness_blockers(readiness),
        *(copy.deepcopy(item) for item in additional_blockers),
    ]
    deduped_blockers: list[dict[str, Any]] = []
    seen_blockers: set[tuple[str, str]] = set()
    for item in blockers:
        key = (str(item.get("code") or ""), str(item.get("message") or ""))
        if key not in seen_blockers:
            seen_blockers.add(key)
            deduped_blockers.append(item)
    blockers = deduped_blockers
    warnings: list[dict[str, Any]] = []
    for raw in readiness.get("warnings") or []:
        if isinstance(raw, dict):
            warnings.append({
                "code": str(raw.get("code") or "warning"),
                "message": str(raw.get("message") or "需人工复核"),
                "details": copy.deepcopy(raw.get("details") or {}),
            })
        else:
            warnings.append({"code": "warning", "message": str(raw), "details": {}})
    lines = [
        f"# {DRAFT_MARKER}",
        "",
        "> 本文件为**专家参考稿/内部复核材料**，供专业人员修订使用，**不是报批终稿**，不构成法律或审批效力。系统与 AI **不承担**专业结论责任；采用前须人工审核。",
        "",
        "## 阻断项与警告摘要",
        "",
    ]
    if blockers:
        lines.extend(
            f"- 阻断项 [{item['code']}]：{item['message']}" for item in blockers
        )
    else:
        lines.append("- 当前自动检查未发现阻断项；本文件仍仅供内部复核。")
    lines.extend(
        f"- 警告 [{item['code']}]：{item['message']}" for item in warnings
    )
    lines.extend(["", "---", "", content])
    summary = {
        "blockers": blockers,
        "warnings": warnings,
        "blocker_count": len(blockers),
        "warning_count": len(warnings),
    }
    return "\n".join(lines), summary


def _draft_basis_blockers(
    basis: dict[str, Any],
    context: dict[str, Any],
) -> list[dict[str, Any]]:
    """Summarize non-readiness gates that still prevent a formal artifact."""

    blockers: list[dict[str, Any]] = []
    for name in _GOVERNED_SNAPSHOTS:
        snapshot = (basis.get("artifacts") or {}).get(name) or {}
        if snapshot.get("present") is not True or snapshot.get("error"):
            blockers.append({
                "code": "governed_snapshot_incomplete",
                "message": f"{name} 快照缺失或损坏",
                "details": {"artifact": name, "error": snapshot.get("error")},
            })
    finance = basis.get("finance") or {}
    run = finance.get("run_snapshot")
    if not finance.get("run_id"):
        blockers.append({
            "code": "finance_binding_required",
            "message": "尚未绑定财务 run",
            "details": {},
        })
    elif not isinstance(run, dict) or run.get("_load_error"):
        blockers.append({
            "code": "finance_run_unavailable",
            "message": "绑定的财务 run 不存在或不可读",
            "details": {"run_id": finance.get("run_id")},
        })
    elif run.get("review_status") != "approved":
        blockers.append({
            "code": "finance_run_not_approved",
            "message": "绑定的财务 run 尚未批准",
            "details": {"run_id": finance.get("run_id")},
        })
    finance_gate = finance.get("publish_gate") or {}
    for raw in finance_gate.get("blockers") or []:
        if isinstance(raw, dict):
            blockers.append({
                "code": str(raw.get("code") or "finance_publish_gate_blocked"),
                "message": str(raw.get("message") or "财务正式发布门禁未通过"),
                "details": copy.deepcopy(raw.get("details") or {}),
            })
    for appendix_file in basis.get("appendix_files") or []:
        if appendix_file.get("ok") is not True:
            blockers.append({
                "code": "appendix_file_inconsistent",
                "message": "已就绪附表文件不存在或哈希不一致",
                "details": copy.deepcopy(appendix_file),
            })
    return blockers


def _set_docx_metadata(
    data: bytes,
    *,
    title: str,
    subject: str,
    keywords: Sequence[str],
    comments: str,
) -> bytes:
    try:
        from docx import Document

        document = Document(io.BytesIO(data))
        properties = document.core_properties
        properties.title = str(title or "可行性研究报告")[:255]
        properties.subject = str(subject or "")[:255]
        properties.keywords = ", ".join(str(item) for item in keywords)[:255]
        properties.comments = str(comments or "")[:255]
        output = io.BytesIO()
        document.save(output)
        return output.getvalue()
    except Exception as exc:  # noqa: BLE001 - metadata is part of the contract
        raise DeliverableArtifactError(
            "DOCX_METADATA_WRITE_FAILED",
            "DOCX 元数据写入失败",
            details={"error": type(exc).__name__},
        ) from exc


def _normalize_declared_hash(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if text.startswith("sha256:"):
        digest = text.removeprefix("sha256:")
    else:
        digest = text
    return f"sha256:{digest}" if re.fullmatch(r"[0-9a-f]{64}", digest) else ""


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def _path_has_symlink(path: Path, root: Path) -> bool:
    """Return whether a path or one of its descendants below root is a symlink."""

    current = path.absolute()
    stop = root.absolute()
    while True:
        if current.is_symlink():
            return True
        if current == stop:
            return False
        if current.parent == current:
            return True
        current = current.parent


def _verify_finance_workbook(path: Path, run: dict[str, Any]) -> tuple[bool, str]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if "Meta" not in workbook.sheetnames:
                return False, "meta_sheet_missing"
            rows = workbook["Meta"].iter_rows(values_only=True)
            metadata = {
                str(row[0] or "").strip(): row[1]
                for row in rows
                if row and len(row) >= 2 and str(row[0] or "").strip() not in {"", "key"}
            }
        finally:
            workbook.close()
    except Exception as exc:  # noqa: BLE001 - optional support file is skipped
        return False, f"workbook_unreadable:{type(exc).__name__}"
    if str(metadata.get("run_id") or "") != str(run.get("run_id") or ""):
        return False, "run_id_mismatch"
    formal_value = metadata.get("formal_delivery_ready")
    if str(formal_value).strip().lower() not in {"1", "true", "yes"}:
        return False, "formal_delivery_not_ready"
    expected_manifest = str(run.get("manifest_hash") or "")
    actual_manifest = str(metadata.get("manifest_hash") or "")
    if expected_manifest and actual_manifest != expected_manifest:
        return False, "manifest_hash_mismatch"
    return True, ""


def _safe_support_source(workspace_root: Path, raw_path: Any) -> Path | None:
    text = str(raw_path or "").strip()
    if not text or "\x00" in text:
        return None
    source = Path(text)
    if not source.is_absolute():
        source = workspace_root / source
    try:
        resolved = source.resolve(strict=True)
        root = workspace_root.resolve(strict=True)
    except OSError:
        return None
    if not _is_relative_to(resolved, root):
        return None
    if not resolved.is_file() or _path_has_symlink(source, workspace_root):
        return None
    if resolved.suffix.lower() not in _SUPPORT_SUFFIXES:
        return None
    return resolved


def _safe_filename(value: str) -> str:
    name = re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff]+", "_", value).strip(" ._")
    return (name or "appendix")[:160]


def _appendix_path_rows(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, dict):
        return []
    rows: list[dict[str, Any]] = []
    for collection in ("tables", "figures", "attachments"):
        for raw in value.get(collection) or []:
            if not isinstance(raw, dict):
                continue
            status = str(
                raw.get("review_state") or raw.get("status") or ""
            ).lower()
            if status not in _VERIFIED_APPENDIX_STATES:
                continue
            source_path = next((
                raw.get(key)
                for key in (
                    "path", "file_path", "storage_path", "artifact_path", "source_path",
                )
                if raw.get(key)
            ), None)
            if source_path:
                rows.append({
                    "source_path": source_path,
                    "declared_hash": next((
                        raw.get(key)
                        for key in ("sha256", "content_hash", "hash")
                        if raw.get(key)
                    ), ""),
                    "collection": collection,
                    "id": str(raw.get("id") or raw.get("template_id") or ""),
                })
    return rows


def _appendix_files_snapshot(workspace_id: str, value: Any) -> list[dict[str, Any]]:
    """Hash every governed, file-backed appendix without trusting its claim."""

    workspace_root = _workspace_root(workspace_id).resolve()
    snapshots: list[dict[str, Any]] = []
    for row in _appendix_path_rows(value):
        raw_source = row.get("source_path")
        source = _safe_support_source(workspace_root, raw_source)
        raw_declared = str(row.get("declared_hash") or "").strip()
        declared = _normalize_declared_hash(raw_declared)
        item: dict[str, Any] = {
            "appendix_id": row.get("id"),
            "collection": row.get("collection"),
            "source": str(raw_source or ""),
            "declared_hash": declared or raw_declared,
            "ok": False,
        }
        if source is None:
            item["error"] = "file_unavailable"
        elif raw_declared and not declared:
            item["error"] = "declared_hash_invalid"
        else:
            try:
                actual_hash, size = _file_hash(source)
            except DeliverableArtifactError:
                item["error"] = "file_unreadable"
            else:
                item.update({
                    "source": source.relative_to(workspace_root).as_posix(),
                    "sha256": actual_hash,
                    "size_bytes": size,
                    "ok": not declared or declared == actual_hash,
                })
                if declared and declared != actual_hash:
                    item["error"] = "declared_hash_mismatch"
        snapshots.append(item)
    return snapshots


def _copy_support_file(source: Path, target: Path) -> dict[str, Any]:
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        with source.open("rb") as reader, target.open("xb") as writer:
            shutil.copyfileobj(reader, writer, length=1024 * 1024)
            writer.flush()
            os.fsync(writer.fileno())
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_SUPPORT_COPY_FAILED",
            "交付工件附件复制失败",
            details={"source": source.name, "error": type(exc).__name__},
        ) from exc
    digest, size = _file_hash(target)
    return {"sha256": digest, "size_bytes": size}


def _collect_support_files(
    workspace_id: str,
    temp_root: Path,
    basis: dict[str, Any],
    context: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    copied: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    workspace_root = _workspace_root(workspace_id).resolve()
    seen: set[Path] = set()
    finance = basis.get("finance") or {}
    run = finance.get("run_snapshot")
    run_id = str(finance.get("run_id") or "")
    if isinstance(run, dict) and run_id:
        finance_root = workspace_root / "finance_artifacts" / run_id
        try:
            candidates = sorted(finance_root.rglob("*.xlsx")) if finance_root.is_dir() else []
        except OSError:
            candidates = []
        copied_finance_count = 0
        for source in candidates:
            source = _safe_support_source(workspace_root, source)
            if source is None or source in seen:
                continue
            verified, reason = _verify_finance_workbook(source, run)
            if not verified:
                warnings.append({
                    "code": "FINANCE_XLSX_NOT_VERIFIED",
                    "filename": source.name,
                    "reason": reason,
                })
                continue
            source_hash_before, _source_size_before = _file_hash(source)
            copied_finance_count += 1
            relative = PurePosixPath("finance") / (
                f"{copied_finance_count:02d}_{_safe_filename(source.name)}"
            )
            target = temp_root.joinpath(*relative.parts)
            metadata = _copy_support_file(source, target)
            target_verified, target_reason = _verify_finance_workbook(target, run)
            source_hash_after, _source_size_after = _file_hash(source)
            if (
                not target_verified
                or metadata.get("sha256") != source_hash_before
                or metadata.get("sha256") != source_hash_after
            ):
                target.unlink(missing_ok=True)
                warnings.append({
                    "code": "FINANCE_XLSX_COPY_NOT_VERIFIED",
                    "filename": source.name,
                    "reason": target_reason or "source_changed_during_copy",
                })
                continue
            seen.add(source)
            copied.append({
                "name": relative.as_posix(),
                "role": "verified_finance_xlsx",
                "media_type": mimetypes.guess_type(source.name)[0]
                or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                **metadata,
            })

    appendix = (context.get("artifact_values") or {}).get("appendix_manifest")
    appendix_basis = basis.get("appendix_files") or []
    for index, row in enumerate(_appendix_path_rows(appendix), start=1):
        source = _safe_support_source(workspace_root, row.get("source_path"))
        if source is None:
            warnings.append({
                "code": "APPENDIX_FILE_UNAVAILABLE",
                "appendix_id": row.get("id"),
            })
            continue
        if source in seen:
            continue
        raw_declared_hash = str(row.get("declared_hash") or "").strip()
        declared_hash = _normalize_declared_hash(raw_declared_hash)
        if raw_declared_hash and not declared_hash:
            warnings.append({
                "code": "APPENDIX_HASH_INVALID",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        relative_source = source.relative_to(workspace_root).as_posix()
        expected_snapshot = next((
            item for item in appendix_basis
            if item.get("source") == relative_source
            and item.get("appendix_id") == row.get("id")
        ), None)
        if not isinstance(expected_snapshot, dict) or expected_snapshot.get("ok") is not True:
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        relative = PurePosixPath("appendices") / (
            f"{index:02d}_{_safe_filename(source.name)}"
        )
        target = temp_root.joinpath(*relative.parts)
        metadata = _copy_support_file(source, target)
        if metadata.get("sha256") != expected_snapshot.get("sha256"):
            target.unlink(missing_ok=True)
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        if declared_hash and metadata.get("sha256") != declared_hash:
            target.unlink(missing_ok=True)
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        seen.add(source)
        copied.append({
            "name": relative.as_posix(),
            "role": "verified_appendix",
            "appendix_id": row.get("id"),
            "media_type": mimetypes.guess_type(source.name)[0]
            or "application/octet-stream",
            **metadata,
        })
    return copied, warnings


def _file_entry(path: Path, root: Path, *, role: str) -> dict[str, Any]:
    digest, size = _file_hash(path)
    relative = path.relative_to(root).as_posix()
    return {
        "name": relative,
        "role": role,
        "sha256": digest,
        "size_bytes": size,
        "media_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
    }


def _build_artifact_directory(
    workspace_id: str,
    artifact_id: str,
    *,
    kind: str,
    docx_bytes: bytes,
    basis: dict[str, Any],
    blocker_summary: dict[str, Any],
    actor: dict[str, Any],
    context: dict[str, Any],
    docx_font_audit: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    artifacts_root = _artifacts_root(workspace_id)
    artifacts_root.mkdir(parents=True, exist_ok=True)
    final_root = _artifact_root(workspace_id, artifact_id)
    temp_root = artifacts_root / f".{artifact_id}.{uuid.uuid4().hex}.tmp"
    renamed = False
    if final_root.exists():
        raise DeliverableArtifactError(
            "ARTIFACT_ALREADY_EXISTS", "交付工件目录已存在",
        )
    try:
        temp_root.mkdir(parents=False, exist_ok=False)
        report_path = temp_root / "report.docx"
        _write_bytes(report_path, docx_bytes)
        basis_path = temp_root / "basis_snapshot.json"
        _write_json_file(basis_path, basis)

        payload_files = [
            _file_entry(report_path, temp_root, role="report_docx"),
            _file_entry(basis_path, temp_root, role="basis_snapshot"),
        ]
        support_warnings: list[dict[str, Any]] = []
        if kind == "formal":
            support_files, support_warnings = _collect_support_files(
                workspace_id, temp_root, basis, context,
            )
            inconsistent_appendices = [
                item for item in support_warnings
                if item.get("code") in {
                    "APPENDIX_FILE_UNAVAILABLE",
                    "APPENDIX_HASH_INVALID",
                    "APPENDIX_HASH_MISMATCH",
                    "FINANCE_XLSX_COPY_NOT_VERIFIED",
                }
            ]
            if inconsistent_appendices:
                raise DeliverableArtifactError(
                    "FORMAL_BASIS_INCONSISTENT",
                    "附表清单声明的已就绪文件不存在或哈希不一致",
                    details={"problems": inconsistent_appendices},
                )
            payload_files.extend(support_files)

        finance_basis = basis.get("finance") or {}
        finance_run = finance_basis.get("run_snapshot") or {}
        manifest = {
            "schema_version": MANIFEST_SCHEMA_VERSION,
            "artifact_id": artifact_id,
            "workspace_id": workspace_id,
            "tenant_scope_hash": _LOCAL_TENANT_SCOPE,
            "kind": kind,
            "template_version": basis.get("template_version"),
            "basis_fingerprint": basis.get("fingerprint"),
            "document": copy.deepcopy(basis.get("document") or {}),
            "workspace_version": basis.get("workspace_version"),
            "finance": {
                "run_id": finance_basis.get("run_id"),
                "run_kind": finance_basis.get("run_kind"),
                "run_hash": finance_basis.get("run_hash"),
                "binding_hash": (finance_basis.get("binding") or {}).get("hash"),
                "input_hash": finance_run.get("input_hash"),
                "spec_hash": finance_run.get("spec_hash"),
                "table_bundle_hash": finance_run.get("table_bundle_hash"),
                "manifest_hash": finance_run.get("manifest_hash"),
                "model_version": finance_run.get("model_version"),
                "template_version": finance_run.get("template_version"),
                "review_status": finance_run.get("review_status"),
                "publish_gate_hash": _canonical_hash(
                    finance_basis.get("publish_gate") or {}
                ),
            },
            "professional_review": copy.deepcopy(
                basis.get("professional_review") or {}
            ),
            "governed_snapshots": copy.deepcopy(basis.get("artifacts") or {}),
            "appendix_file_snapshots": copy.deepcopy(
                basis.get("appendix_files") or []
            ),
            "blocker_summary": copy.deepcopy(blocker_summary),
            "support_file_warnings": copy.deepcopy(support_warnings),
            "created_by": copy.deepcopy(actor),
            "created_at": _now(),
            "payload_files": copy.deepcopy(payload_files),
            "docx_font_audit": copy.deepcopy(docx_font_audit or {}),
        }
        manifest_path = temp_root / "manifest.json"
        _write_json_file(manifest_path, manifest)
        manifest_entry = _file_entry(manifest_path, temp_root, role="manifest")
        index = {
            "schema_version": MANIFEST_SCHEMA_VERSION,
            "artifact_id": artifact_id,
            "workspace_id": workspace_id,
            "tenant_scope_hash": _LOCAL_TENANT_SCOPE,
            "kind": kind,
            "basis_fingerprint": basis.get("fingerprint"),
            "files": [*copy.deepcopy(payload_files), copy.deepcopy(manifest_entry)],
        }
        index_path = temp_root / "index.json"
        _write_json_file(index_path, index)
        index_entry = _file_entry(index_path, temp_root, role="index")
        files = [*payload_files, manifest_entry, index_entry]

        failures = _verify_files(temp_root, files)
        if failures:
            raise DeliverableArtifactError(
                "ARTIFACT_INTEGRITY_FAILED",
                "交付工件生成后的完整性校验失败",
                details={"failures": failures},
            )
        os.replace(temp_root, final_root)
        renamed = True
        failures = _verify_files(final_root, files)
        if failures:
            raise DeliverableArtifactError(
                "ARTIFACT_INTEGRITY_FAILED",
                "交付工件原子落盘后的完整性校验失败",
                details={"failures": failures},
            )
        return files, support_warnings
    except DeliverableArtifactError:
        if temp_root.exists():
            shutil.rmtree(temp_root, ignore_errors=True)
        if renamed and final_root.exists():
            shutil.rmtree(final_root, ignore_errors=True)
        raise
    except Exception as exc:  # noqa: BLE001 - normalize generation failures
        if temp_root.exists():
            shutil.rmtree(temp_root, ignore_errors=True)
        if renamed and final_root.exists():
            shutil.rmtree(final_root, ignore_errors=True)
        raise DeliverableArtifactError(
            "ARTIFACT_GENERATION_FAILED",
            "交付工件生成失败",
            details={"error": type(exc).__name__},
        ) from exc


def _safe_relative_name(filename: str) -> str:
    value = str(filename or "").strip()
    if (
        not value
        or "\x00" in value
        or "\\" in value
        or any(ord(char) < 32 or ord(char) == 127 for char in value)
    ):
        raise DeliverableArtifactError(
            "INVALID_FILENAME_PATH", "下载文件名不合法",
        )
    path = PurePosixPath(value)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise DeliverableArtifactError(
            "INVALID_FILENAME_PATH", "下载文件名不合法",
        )
    return path.as_posix()


def _verify_files(root: Path, files: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    failures: list[dict[str, Any]] = []
    try:
        resolved_root = root.resolve(strict=True)
    except OSError:
        return [{"code": "ARTIFACT_DIRECTORY_MISSING", "path": str(root)}]
    seen: set[str] = set()
    for entry in files:
        try:
            name = _safe_relative_name(str(entry.get("name") or ""))
        except DeliverableArtifactError:
            failures.append({"code": "INVALID_FILENAME_PATH", "name": entry.get("name")})
            continue
        if name in seen:
            failures.append({"code": "DUPLICATE_FILE_ENTRY", "name": name})
            continue
        seen.add(name)
        path = root.joinpath(*PurePosixPath(name).parts)
        try:
            resolved = path.resolve(strict=True)
        except OSError:
            failures.append({"code": "ARTIFACT_FILE_MISSING", "name": name})
            continue
        if (
            not _is_relative_to(resolved, resolved_root)
            or _path_has_symlink(path, root)
            or not resolved.is_file()
        ):
            failures.append({"code": "ARTIFACT_FILE_PATH_UNSAFE", "name": name})
            continue
        try:
            actual_hash, actual_size = _file_hash(resolved)
        except DeliverableArtifactError as exc:
            failures.append({"code": exc.code, "name": name})
            continue
        if actual_hash != entry.get("sha256"):
            failures.append({
                "code": "ARTIFACT_FILE_HASH_MISMATCH",
                "name": name,
                "expected": entry.get("sha256"),
                "actual": actual_hash,
            })
        if actual_size != entry.get("size_bytes"):
            failures.append({
                "code": "ARTIFACT_FILE_SIZE_MISMATCH",
                "name": name,
                "expected": entry.get("size_bytes"),
                "actual": actual_size,
            })
    try:
        actual_names = {
            path.relative_to(root).as_posix()
            for path in root.rglob("*")
            if path.is_file()
        }
    except OSError:
        actual_names = set()
    unexpected = sorted(actual_names - seen)
    for name in unexpected:
        failures.append({"code": "UNINDEXED_ARTIFACT_FILE", "name": name})
    by_role = {str(entry.get("role") or ""): entry for entry in files}
    manifest_entry = by_role.get("manifest")
    index_entry = by_role.get("index")
    if not isinstance(manifest_entry, dict) or not isinstance(index_entry, dict):
        failures.append({"code": "ARTIFACT_INDEX_REQUIRED"})
        return failures
    try:
        manifest_path = root.joinpath(
            *PurePosixPath(str(manifest_entry["name"])).parts
        )
        index_path = root.joinpath(*PurePosixPath(str(index_entry["name"])).parts)
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        index = json.loads(index_path.read_text(encoding="utf-8"))
    except (KeyError, OSError, UnicodeError, json.JSONDecodeError) as exc:
        failures.append({
            "code": "ARTIFACT_INDEX_UNREADABLE",
            "error": type(exc).__name__,
        })
        return failures
    expected_payload = [
        copy.deepcopy(entry)
        for entry in files
        if entry.get("role") not in {"manifest", "index"}
    ]
    expected_index = [
        copy.deepcopy(entry) for entry in files if entry.get("role") != "index"
    ]
    if not isinstance(manifest, dict) or _canonical_hash(
        manifest.get("payload_files") or []
    ) != _canonical_hash(expected_payload):
        failures.append({"code": "ARTIFACT_MANIFEST_FILE_INDEX_MISMATCH"})
    if not isinstance(index, dict) or _canonical_hash(
        index.get("files") or []
    ) != _canonical_hash(expected_index):
        failures.append({"code": "ARTIFACT_INDEX_FILE_LIST_MISMATCH"})
    if isinstance(manifest, dict) and isinstance(index, dict):
        for field in (
            "artifact_id", "workspace_id", "kind", "basis_fingerprint",
        ):
            if manifest.get(field) != index.get(field):
                failures.append({
                    "code": "ARTIFACT_MANIFEST_INDEX_METADATA_MISMATCH",
                    "field": field,
                })
    return failures


def _basis_change_reasons(
    stored: dict[str, Any],
    current: dict[str, Any],
) -> list[dict[str, Any]]:
    reasons: list[dict[str, Any]] = []

    def changed(code: str, message: str, expected: Any, actual: Any) -> None:
        if expected != actual:
            reasons.append({
                "code": code,
                "message": message,
                "expected": expected,
                "actual": actual,
            })

    stored_doc = stored.get("document") or {}
    current_doc = current.get("document") or {}
    changed(
        "DOCUMENT_REVISION_CHANGED", "当前文档修订已变化",
        stored_doc.get("revision_id"), current_doc.get("revision_id"),
    )
    changed(
        "DOCUMENT_CONTENT_CHANGED", "当前文档正文已变化",
        stored_doc.get("content_hash"), current_doc.get("content_hash"),
    )
    changed(
        "WORKSPACE_METADATA_CHANGED", "工作区报告元数据已变化",
        stored_doc.get("metadata_hash"), current_doc.get("metadata_hash"),
    )
    changed(
        "WORKSPACE_VERSION_CHANGED", "工作区版本已变化",
        stored.get("workspace_version"), current.get("workspace_version"),
    )
    changed(
        "SOURCE_FILES_CHANGED", "原始资料文件、解析或复核状态已变化",
        (stored.get("sources") or {}).get("hash"),
        (current.get("sources") or {}).get("hash"),
    )
    changed(
        "TEMPLATE_VERSION_CHANGED", "报告模板版本已变化",
        stored.get("template_version"), current.get("template_version"),
    )
    changed(
        "PUBLISH_READINESS_CHANGED", "发布就绪度依据已变化",
        (stored.get("readiness") or {}).get("hash"),
        (current.get("readiness") or {}).get("hash"),
    )

    stored_finance = stored.get("finance") or {}
    current_finance = current.get("finance") or {}
    changed(
        "FINANCE_BINDING_CHANGED", "财务绑定已变化",
        (stored_finance.get("binding") or {}).get("hash"),
        (current_finance.get("binding") or {}).get("hash"),
    )
    changed(
        "FINANCE_RUN_REBOUND", "报告已绑定到不同财务 run",
        stored_finance.get("run_id"), current_finance.get("run_id"),
    )
    changed(
        "FINANCE_RUN_CHANGED", "绑定财务 run 内容或批准状态已变化",
        stored_finance.get("run_hash"), current_finance.get("run_hash"),
    )
    changed(
        "FINANCE_PUBLISH_GATE_CHANGED", "财务正式发布门禁结果已变化",
        _canonical_hash(stored_finance.get("publish_gate") or {}),
        _canonical_hash(current_finance.get("publish_gate") or {}),
    )

    for name in _GOVERNED_SNAPSHOTS:
        changed(
            f"{name.upper()}_CHANGED", f"{name} 已变化",
            ((stored.get("artifacts") or {}).get(name) or {}).get("hash"),
            ((current.get("artifacts") or {}).get(name) or {}).get("hash"),
        )
    changed(
        "APPENDIX_FILES_CHANGED", "附表文件内容或可用性已变化",
        _canonical_hash(stored.get("appendix_files") or []),
        _canonical_hash(current.get("appendix_files") or []),
    )

    stored_review = stored.get("professional_review") or {}
    current_review = current.get("professional_review") or {}
    changed(
        "PROFESSIONAL_REVIEW_CHANGED", "专业复核记录或状态已变化",
        _canonical_hash(stored_review), _canonical_hash(current_review),
    )
    if stored.get("fingerprint") != current.get("fingerprint") and not reasons:
        reasons.append({
            "code": "ARTIFACT_BASIS_CHANGED",
            "message": "交付工件输入依据指纹已变化",
            "expected": stored.get("fingerprint"),
            "actual": current.get("fingerprint"),
        })
    return reasons


def _append_event(
    state: dict[str, Any],
    artifact_id: str,
    event: str,
    *,
    actor: dict[str, Any],
    details: dict[str, Any] | None = None,
) -> None:
    state.setdefault("history", []).append({
        "event_id": f"artevent_{uuid.uuid4().hex}",
        "artifact_id": artifact_id,
        "event": event,
        "actor": copy.deepcopy(actor),
        "details": copy.deepcopy(details or {}),
        "created_at": _now(),
    })


def _persist_new_record(
    workspace_id: str,
    state: dict[str, Any],
    record: dict[str, Any],
) -> None:
    artifact_id = str(record["artifact_id"])
    now = _now()
    kind = str(record["kind"])
    previous_id = str((state.setdefault("current", {})).get(kind) or "")
    previous = (state.setdefault("artifacts", {})).get(previous_id)
    if previous_id and isinstance(previous, dict) and previous_id != artifact_id:
        previous = copy.deepcopy(previous)
        previous["current"] = False
        previous["superseded_by"] = artifact_id
        previous["superseded_at"] = now
        previous["updated_at"] = now
        state["artifacts"][previous_id] = previous
        _append_event(
            state,
            previous_id,
            "superseded",
            actor=copy.deepcopy(record["created_by"]),
            details={"superseded_by": artifact_id, "kind": kind},
        )
    state.setdefault("artifacts", {})[artifact_id] = record
    state.setdefault("current", {})[kind] = artifact_id
    state["created_at"] = str(state.get("created_at") or now)
    state["updated_at"] = now
    _append_event(
        state, artifact_id, "created",
        actor=copy.deepcopy(record["created_by"]),
        details={
            "kind": record["kind"],
            "basis_fingerprint": record["basis_fingerprint"],
        },
    )
    _write_json_atomic(_state_path(workspace_id), state)


def _create(
    workspace_id: str,
    *,
    kind: str,
    actor: Mapping[str, Any] | None,
    template_version: str,
    operation_id: str = "",
) -> dict[str, Any]:
    workspace_id = _validate_workspace_id(workspace_id)
    _require_workspace(workspace_id)
    template_version = _validate_template_version(template_version)
    validated_actor = _validated_actor(actor)
    operation_id = str(operation_id or "").strip()
    if operation_id and not _SAFE_OPERATION_ID.fullmatch(operation_id):
        raise DeliverableArtifactError(
            "INVALID_OPERATION_ID", "交付工件操作标识不合法",
        )
    artifact_id = f"deliverable_{uuid.uuid4().hex}"
    final_root = _artifact_root(workspace_id, artifact_id)
    with _state_guard(workspace_id):
        state = _read_state(workspace_id)
        if operation_id:
            existing = next((
                item for item in (state.get("artifacts") or {}).values()
                if str(item.get("operation_id") or "") == operation_id
                and str(item.get("kind") or "") == kind
            ), None)
            if isinstance(existing, dict):
                replay = _refresh_record_locked(
                    workspace_id,
                    state,
                    str(existing.get("artifact_id") or ""),
                )
                return {**copy.deepcopy(replay), "idempotent_replay": True}
        basis, content, context = _capture_basis(
            workspace_id,
            template_version=template_version,
        )
        readiness = (basis.get("readiness") or {}).get("snapshot") or {}
        if kind == "formal":
            _assert_formal_basis(basis, context)
            report_content = content
            blocker_summary = {
                "blockers": [],
                "warnings": copy.deepcopy(readiness.get("warnings") or []),
                "blocker_count": 0,
                "warning_count": len(readiness.get("warnings") or []),
            }
            subject = "受控可行性研究报告交付工件"
            keywords = ["可行性研究报告", "受控交付工件"]
            comments = "由服务端基于批准 run 和有效专业复核依据生成；不表示法律签署。"
        else:
            report_content, blocker_summary = _marker_markdown(
                content,
                readiness,
                additional_blockers=_draft_basis_blockers(basis, context),
            )
            subject = DRAFT_MARKER
            keywords = [DRAFT_MARKER, "非正式发布件"]
            comments = "专家参考稿/内部复核·非报批终稿；可含未决差异与警告；AI不担责；不得作为正式发布件。"

        raw_docx = doc_service.markdown_to_docx(report_content)
        title = str((context.get("meta") or {}).get("title") or "可行性研究报告")
        docx_bytes = _set_docx_metadata(
            raw_docx,
            title=f"{DRAFT_MARKER} - {title}" if kind == "draft" else title,
            subject=subject,
            keywords=keywords,
            comments=comments,
        )
        from lvke_mcp.domains.reports.docx_fonts import normalize_docx_fonts

        docx_bytes, docx_font_audit = normalize_docx_fonts(docx_bytes)

        second_basis, _second_content, second_context = _capture_basis(
            workspace_id,
            template_version=template_version,
        )
        if second_basis.get("fingerprint") != basis.get("fingerprint"):
            raise DeliverableArtifactError(
                "BASIS_CHANGED_DURING_EXPORT",
                "交付工件生成期间输入依据发生变化，请刷新后重试",
                details={
                    "before": basis.get("fingerprint"),
                    "after": second_basis.get("fingerprint"),
                },
            )
        if kind == "formal":
            _assert_formal_basis(second_basis, second_context)

        files, support_warnings = _build_artifact_directory(
            workspace_id,
            artifact_id,
            kind=kind,
            docx_bytes=docx_bytes,
            basis=second_basis,
            blocker_summary=blocker_summary,
            actor=validated_actor,
            context=second_context,
            docx_font_audit=docx_font_audit,
        )
        try:
            final_basis, _final_content, final_context = _capture_basis(
                workspace_id,
                template_version=template_version,
            )
            if final_basis.get("fingerprint") != second_basis.get("fingerprint"):
                raise DeliverableArtifactError(
                    "BASIS_CHANGED_DURING_EXPORT",
                    "交付工件装配期间输入依据发生变化，请刷新后重试",
                    details={
                        "before": second_basis.get("fingerprint"),
                        "after": final_basis.get("fingerprint"),
                    },
                )
            if kind == "formal":
                _assert_formal_basis(final_basis, final_context)
        except Exception:
            if final_root.exists():
                shutil.rmtree(final_root, ignore_errors=True)
            raise
        now = _now()
        manifest_hash = next(
            (item.get("sha256") for item in files if item.get("role") == "manifest"),
            "",
        )
        index_hash = next(
            (item.get("sha256") for item in files if item.get("role") == "index"),
            "",
        )
        record = {
            "schema_version": SCHEMA_VERSION,
            "artifact_id": artifact_id,
            "workspace_id": workspace_id,
            "tenant_scope_hash": _LOCAL_TENANT_SCOPE,
            "kind": kind,
            "operation_id": operation_id,
            "status": "succeeded",
            "ok": True,
            "current": True,
            "template_version": template_version,
            "basis_fingerprint": final_basis.get("fingerprint"),
            "basis": final_basis,
            "document_revision_id": (final_basis.get("document") or {}).get(
                "revision_id"
            ),
            "finance_run_id": (final_basis.get("finance") or {}).get("run_id"),
            "run_id": (final_basis.get("finance") or {}).get("run_id"),
            "spec_hash": (
                ((final_basis.get("finance") or {}).get("run_snapshot") or {}).get(
                    "spec_hash"
                )
            ),
            "professional_review_id": (
                final_basis.get("professional_review") or {}
            ).get("review_id"),
            "blocker_summary": blocker_summary,
            "support_file_warnings": support_warnings,
            "files": files,
            "manifest_hash": manifest_hash,
            "index_hash": index_hash,
            "integrity_status": "passed",
            "created_by": validated_actor,
            "created_at": now,
            "updated_at": now,
            "invalidation_reasons": [],
            "release_status": "not_released",
            "release_history": [],
        }
        try:
            _persist_new_record(workspace_id, state, record)
        except Exception:
            if final_root.exists():
                shutil.rmtree(final_root, ignore_errors=True)
            raise
    return copy.deepcopy(record)


def create_draft_export(
    workspace_id: str,
    *,
    actor: Mapping[str, Any] | None = None,
    template_version: str = DEFAULT_TEMPLATE_VERSION,
    operation_id: str = "",
) -> dict[str, Any]:
    """Create an immutable, visibly watermarked internal-review DOCX.

    Draft export intentionally permits readiness blockers and warnings.  They
    are rendered into the document and stored in its immutable manifest.
    """

    return _create(
        workspace_id,
        kind="draft",
        actor=actor,
        template_version=template_version,
        operation_id=operation_id,
    )


def create_deliverable_artifact(
    workspace_id: str,
    *,
    actor: Mapping[str, Any] | None = None,
    template_version: str = DEFAULT_TEMPLATE_VERSION,
    operation_id: str = "",
) -> dict[str, Any]:
    """Create a formal generic-feasibility artifact after all gates pass."""

    return _create(
        workspace_id,
        kind="formal",
        actor=actor,
        template_version=template_version,
        operation_id=operation_id,
    )


def _invalidate_locked(
    workspace_id: str,
    state: dict[str, Any],
    record: dict[str, Any],
    reasons: list[dict[str, Any]],
) -> dict[str, Any]:
    if record.get("status") == "invalidated":
        return record
    now = _now()
    artifact_id = str(record["artifact_id"])
    previous_status = str(record.get("status") or "")
    previous_release_status = str(record.get("release_status") or "")
    record["status"] = "invalidated"
    record["current"] = False
    record["integrity_status"] = (
        "failed" if any(
            "FILE_" in str(item.get("code") or "")
            or item.get("code") == "PROFESSIONAL_SIGNOFF_INTEGRITY_FAILED"
            for item in reasons
        )
        else record.get("integrity_status") or "passed"
    )
    record["invalidation_reasons"] = copy.deepcopy(reasons)
    record["invalidated_at"] = now
    record["updated_at"] = now
    record["previous_status"] = previous_status
    if previous_release_status == "released":
        record["historical_release_status"] = "released"
        record["release_status"] = "invalidated"
    state.setdefault("artifacts", {})[artifact_id] = record
    kind = str(record.get("kind") or "")
    if (state.setdefault("current", {})).get(kind) == artifact_id:
        state["current"][kind] = ""
    state["updated_at"] = now
    _append_event(
        state,
        artifact_id,
        "invalidated",
        actor=_INTERNAL_ACTOR,
        details={"previous_status": previous_status, "reasons": reasons},
    )
    _write_json_atomic(_state_path(workspace_id), state)
    return record


def _refresh_record_locked(
    workspace_id: str,
    state: dict[str, Any],
    artifact_id: str,
) -> dict[str, Any]:
    raw = (state.get("artifacts") or {}).get(artifact_id)
    if not isinstance(raw, dict):
        raise DeliverableArtifactError("ARTIFACT_NOT_FOUND", "交付工件不存在")
    record = copy.deepcopy(raw)
    root = _artifact_root(workspace_id, artifact_id)
    integrity_failures = _verify_files(root, record.get("files") or [])
    reasons: list[dict[str, Any]] = []
    for failure in integrity_failures:
        reasons.append({
            "code": str(failure.get("code") or "ARTIFACT_INTEGRITY_FAILED"),
            "message": "交付工件文件完整性校验失败",
            "details": copy.deepcopy(failure),
        })
    try:
        current_basis, _content, _context = _capture_basis(
            workspace_id,
            template_version=str(record.get("template_version") or DEFAULT_TEMPLATE_VERSION),
        )
        reasons.extend(_basis_change_reasons(record.get("basis") or {}, current_basis))
    except DeliverableArtifactError as exc:
        reasons.append({
            "code": "CURRENT_BASIS_UNAVAILABLE",
            "message": "当前交付依据不可用",
            "details": {"error": exc.code, **copy.deepcopy(exc.details)},
        })
    if reasons:
        record = _invalidate_locked(
            workspace_id,
            state,
            record,
            reasons,
        )
    return record


def get_artifact(
    workspace_id: str,
    artifact_id: str,
) -> dict[str, Any]:
    """Return metadata after fresh basis and file-integrity validation."""

    workspace_id = _validate_workspace_id(workspace_id)
    _require_workspace(workspace_id)
    artifact_id = _validate_artifact_id(artifact_id)
    with _state_guard(workspace_id):
        state = _read_state(workspace_id)
        record = _refresh_record_locked(
            workspace_id,
            state,
            artifact_id,
        )
    return copy.deepcopy(record)


def list_artifacts(
    workspace_id: str,
    *,
    limit: int = 50,
) -> list[dict[str, Any]]:
    """List newest artifacts, refreshing each returned artifact first."""

    workspace_id = _validate_workspace_id(workspace_id)
    _require_workspace(workspace_id)
    bounded_limit = max(1, min(int(limit), 200))
    with _state_guard(workspace_id):
        state = _read_state(workspace_id)
        ordered = sorted(
            (state.get("artifacts") or {}).values(),
            key=lambda item: str(item.get("created_at") or ""),
            reverse=True,
        )[:bounded_limit]
        result = [
            copy.deepcopy(
                _refresh_record_locked(
                    workspace_id,
                    state,
                    str(item["artifact_id"]),
                )
            )
            for item in ordered
        ]
    return result


def _resolve_artifact_download(
    workspace_id: str,
    artifact_id: str,
    filename: str,
) -> dict[str, Any]:
    """Resolve a safe, current, hash-verified artifact file.

    无签审子集：hermes 版对正式工件要求统一审查发布绑定
    （``_require_current_release_review``），MCP 域无该绑定，删除此门禁；
    正式工件在 ``succeeded``（候选）与 ``released``（已内部发布）状态
    均可下载，发布权限由审查域自身承载。
    """

    safe_name = _safe_relative_name(filename)
    record = get_artifact(workspace_id, artifact_id)
    if record.get("status") not in {"succeeded", "released"} or not record.get("current"):
        raise DeliverableArtifactError(
            "ARTIFACT_NOT_CURRENT",
            "交付工件已失效或不是当前可下载工件",
            details={
                "status": record.get("status"),
                "invalidation_reasons": copy.deepcopy(
                    record.get("invalidation_reasons") or []
                ),
            },
        )
    entry = next(
        (item for item in record.get("files") or [] if item.get("name") == safe_name),
        None,
    )
    if not isinstance(entry, dict):
        raise DeliverableArtifactError(
            "ARTIFACT_FILE_NOT_FOUND", "交付工件文件不存在",
            details={"filename": safe_name},
        )
    root = _artifact_root(workspace_id, artifact_id)
    path = root.joinpath(*PurePosixPath(safe_name).parts)
    try:
        resolved_root = root.resolve(strict=True)
        resolved = path.resolve(strict=True)
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_FILE_NOT_FOUND", "交付工件文件不存在",
            details={"filename": safe_name},
        ) from exc
    if (
        not _is_relative_to(resolved, resolved_root)
        or _path_has_symlink(path, root)
        or not resolved.is_file()
    ):
        raise DeliverableArtifactError(
            "INVALID_FILENAME_PATH", "交付工件文件路径不安全",
        )
    actual_hash, actual_size = _file_hash(resolved)
    if actual_hash != entry.get("sha256") or actual_size != entry.get("size_bytes"):
        # A race after get_artifact still fails closed.  The next read will
        # durably persist invalidation; do that now as well.
        get_artifact(workspace_id, artifact_id)
        raise DeliverableArtifactError(
            "ARTIFACT_INTEGRITY_FAILED", "交付工件文件完整性校验失败",
            details={"filename": safe_name},
        )
    return {
        "ok": True,
        "workspace_id": workspace_id,
        "artifact_id": artifact_id,
        "kind": record.get("kind"),
        "status": record.get("status"),
        "filename": safe_name,
        "path": resolved,
        "sha256": actual_hash,
        "size_bytes": actual_size,
        "media_type": entry.get("media_type") or "application/octet-stream",
        "finance_run_id": record.get("finance_run_id"),
        "basis_fingerprint": record.get("basis_fingerprint"),
    }


def read_artifact_download(
    workspace_id: str,
    artifact_id: str,
    filename: str,
) -> dict[str, Any]:
    """Read verified bytes so an HTTP response cannot race a later file open."""

    resolved = _resolve_artifact_download(
        workspace_id,
        artifact_id,
        filename,
    )
    try:
        content = resolved["path"].read_bytes()
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_FILE_UNREADABLE", "交付工件文件不可读取",
            details={"filename": resolved.get("filename")},
        ) from exc
    if (
        _bytes_hash(content) != resolved.get("sha256")
        or len(content) != resolved.get("size_bytes")
    ):
        get_artifact(workspace_id, artifact_id)
        raise DeliverableArtifactError(
            "ARTIFACT_INTEGRITY_FAILED", "下载前工件文件内容发生变化",
            details={"filename": resolved.get("filename")},
        )
    return {**resolved, "content": content}


def record_internal_release(
    workspace_id: str,
    artifact_id: str,
    *,
    actor: Mapping[str, Any] | None = None,
    note: str = "",
) -> dict[str, Any]:
    """Record an authenticated internal release of a current formal artifact.

    无签审子集：hermes 版要求显式绑定已固化的统一审查 review_id
    （``_require_unified_release_review``），MCP 版内部发布不绑定统一审查
    （release 权限门禁由审查域自身承载），review 绑定字段保持空值以兼容
    schema 形状。该记录不是法律签署，也不构成专业复核证明。
    """

    workspace_id = _validate_workspace_id(workspace_id)
    _require_workspace(workspace_id)
    artifact_id = _validate_artifact_id(artifact_id)
    validated_actor = _validated_actor(actor)
    with _state_guard(workspace_id):
        state = _read_state(workspace_id)
        record = _refresh_record_locked(
            workspace_id,
            state,
            artifact_id,
        )
        if record.get("kind") != "formal":
            raise DeliverableArtifactError(
                "DRAFT_RELEASE_FORBIDDEN", "内部初稿不能记录为正式发布",
            )
        if record.get("status") == "released":
            previous = copy.deepcopy(record.get("release") or {})
            if (
                str((previous.get("actor") or {}).get("actor_id") or "")
                == str(validated_actor.get("actor_id") or "")
                and str(previous.get("note") or "")
                == str(note or "").strip()[:2000]
            ):
                return {**copy.deepcopy(record), "idempotent_replay": True}
            raise DeliverableArtifactError(
                "ARTIFACT_ALREADY_RELEASED", "交付工件已经记录过内部发布",
            )
        if record.get("status") != "succeeded" or not record.get("current"):
            raise DeliverableArtifactError(
                "ARTIFACT_NOT_CURRENT", "仅当前有效的正式工件可以记录内部发布",
                details={"status": record.get("status")},
            )
        final_basis, _content, final_context = _capture_basis(
            workspace_id,
            template_version=str(
                record.get("template_version") or DEFAULT_TEMPLATE_VERSION
            ),
        )
        release_reasons = _basis_change_reasons(
            record.get("basis") or {}, final_basis,
        )
        if release_reasons:
            _invalidate_locked(
                workspace_id,
                state,
                record,
                release_reasons,
            )
            raise DeliverableArtifactError(
                "ARTIFACT_NOT_CURRENT",
                "交付依据在内部发布提交前发生变化",
                details={"invalidation_reasons": release_reasons},
            )
        _assert_formal_basis(final_basis, final_context)
        release = {
            "release_id": f"release_{uuid.uuid4().hex}",
            "kind": "authorized_internal_release",
            "actor": validated_actor,
            "note": str(note or "").strip()[:2000],
            "basis_fingerprint": record.get("basis_fingerprint"),
            "review_id": "",
            "review_release_id": "",
            "review_release_hash": "",
            "review_release_basis_hash": "",
            "review_event_chain_hash": "",
            "review_target_sha256": "",
            "created_at": _now(),
            "legal_signature": False,
            "professional_signature": False,
        }
        release["release_integrity_hash"] = _canonical_hash(
            _internal_release_integrity_material(release)
        )
        record.setdefault("release_history", []).append(release)
        record["status"] = "released"
        record["release_status"] = "released"
        record["released_at"] = release["created_at"]
        record["released_by"] = validated_actor["actor_id"]
        record["release"] = copy.deepcopy(release)
        record["updated_at"] = release["created_at"]
        state.setdefault("artifacts", {})[artifact_id] = record
        state["updated_at"] = release["created_at"]
        _append_event(
            state,
            artifact_id,
            "internal_release_recorded",
            actor=validated_actor,
            details={"release_id": release["release_id"]},
        )
        _write_json_atomic(_state_path(workspace_id), state)
    return copy.deepcopy(record)