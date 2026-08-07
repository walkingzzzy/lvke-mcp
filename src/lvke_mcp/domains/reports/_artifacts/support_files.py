"""附件收集与路径安全：符号链接拒绝、workbook 校验与附录清单。"""

from __future__ import annotations

import mimetypes
import os
import re
import shutil
from pathlib import Path, PurePosixPath
from typing import Any



from .base import (
    DeliverableArtifactError,
    _SUPPORT_SUFFIXES,
    _VERIFIED_APPENDIX_STATES,
    _file_hash,
)

from .storage import (
    _finance_artifact_root,
    _workspace_root,
)


def _normalize_declared_hash(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if text.startswith("sha256:"):
        digest = text.removeprefix("sha256:")
    else:
        digest = text
    return f"sha256:{digest}" if re.fullmatch(r"[0-9a-f]{64}", digest) else ""


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def _path_has_symlink(path: Path, root: Path) -> bool:
    """Return whether a path or one of its descendants below root is a symlink."""

    current = path.absolute()
    stop = root.absolute()
    while True:
        if current.is_symlink():
            return True
        if current == stop:
            return False
        if current.parent == current:
            return True
        current = current.parent


def _verify_finance_workbook(path: Path, run: dict[str, Any]) -> tuple[bool, str]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if "Meta" not in workbook.sheetnames:
                return False, "meta_sheet_missing"
            rows = workbook["Meta"].iter_rows(values_only=True)
            metadata = {
                str(row[0] or "").strip(): row[1]
                for row in rows
                if row and len(row) >= 2 and str(row[0] or "").strip() not in {"", "key"}
            }
        finally:
            workbook.close()
    except Exception as exc:  # noqa: BLE001 - optional support file is skipped
        return False, f"workbook_unreadable:{type(exc).__name__}"
    if str(metadata.get("run_id") or "") != str(run.get("run_id") or ""):
        return False, "run_id_mismatch"
    formal_value = metadata.get("validation_complete")
    if str(formal_value).strip().lower() not in {"1", "true", "yes"}:
        return False, "formal_delivery_not_ready"
    expected_manifest = str(run.get("manifest_hash") or "")
    actual_manifest = str(metadata.get("manifest_hash") or "")
    if expected_manifest and actual_manifest != expected_manifest:
        return False, "manifest_hash_mismatch"
    return True, ""


def _safe_support_source(workspace_root: Path, raw_path: Any) -> Path | None:
    text = str(raw_path or "").strip()
    if not text or "\x00" in text:
        return None
    source = Path(text)
    if not source.is_absolute():
        source = workspace_root / source
    try:
        resolved = source.resolve(strict=True)
        root = workspace_root.resolve(strict=True)
    except OSError:
        return None
    if not _is_relative_to(resolved, root):
        return None
    if not resolved.is_file() or _path_has_symlink(source, workspace_root):
        return None
    if resolved.suffix.lower() not in _SUPPORT_SUFFIXES:
        return None
    return resolved


def _safe_filename(value: str) -> str:
    name = re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff]+", "_", value).strip(" ._")
    return (name or "appendix")[:160]


def _appendix_path_rows(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, dict):
        return []
    rows: list[dict[str, Any]] = []
    for collection in ("tables", "figures", "attachments"):
        for raw in value.get(collection) or []:
            if not isinstance(raw, dict):
                continue
            status = str(
                raw.get("review_state") or raw.get("status") or ""
            ).lower()
            if status not in _VERIFIED_APPENDIX_STATES:
                continue
            source_path = next((
                raw.get(key)
                for key in (
                    "path", "file_path", "storage_path", "artifact_path", "source_path",
                )
                if raw.get(key)
            ), None)
            if source_path:
                rows.append({
                    "source_path": source_path,
                    "declared_hash": next((
                        raw.get(key)
                        for key in ("sha256", "content_hash", "hash")
                        if raw.get(key)
                    ), ""),
                    "collection": collection,
                    "id": str(raw.get("id") or raw.get("template_id") or ""),
                })
    return rows


def _appendix_files_snapshot(workspace_id: str, value: Any) -> list[dict[str, Any]]:
    """Hash every governed, file-backed appendix without trusting its claim."""

    workspace_root = _workspace_root(workspace_id).resolve()
    snapshots: list[dict[str, Any]] = []
    for row in _appendix_path_rows(value):
        raw_source = row.get("source_path")
        source = _safe_support_source(workspace_root, raw_source)
        raw_declared = str(row.get("declared_hash") or "").strip()
        declared = _normalize_declared_hash(raw_declared)
        item: dict[str, Any] = {
            "appendix_id": row.get("id"),
            "collection": row.get("collection"),
            "source": str(raw_source or ""),
            "declared_hash": declared or raw_declared,
            "ok": False,
        }
        if source is None:
            item["error"] = "file_unavailable"
        elif raw_declared and not declared:
            item["error"] = "declared_hash_invalid"
        else:
            try:
                actual_hash, size = _file_hash(source)
            except DeliverableArtifactError:
                item["error"] = "file_unreadable"
            else:
                item.update({
                    "source": source.relative_to(workspace_root).as_posix(),
                    "sha256": actual_hash,
                    "size_bytes": size,
                    "ok": not declared or declared == actual_hash,
                })
                if declared and declared != actual_hash:
                    item["error"] = "declared_hash_mismatch"
        snapshots.append(item)
    return snapshots


def _copy_support_file(source: Path, target: Path) -> dict[str, Any]:
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        with source.open("rb") as reader, target.open("xb") as writer:
            shutil.copyfileobj(reader, writer, length=1024 * 1024)
            writer.flush()
            os.fsync(writer.fileno())
    except OSError as exc:
        raise DeliverableArtifactError(
            "ARTIFACT_SUPPORT_COPY_FAILED",
            "交付工件附件复制失败",
            details={"source": source.name, "error": type(exc).__name__},
        ) from exc
    digest, size = _file_hash(target)
    return {"sha256": digest, "size_bytes": size}


def _collect_support_files(
    workspace_id: str,
    temp_root: Path,
    basis: dict[str, Any],
    context: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    copied: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    workspace_root = _workspace_root(workspace_id).resolve()
    seen: set[Path] = set()
    finance = basis.get("finance") or {}
    run = finance.get("run_snapshot")
    run_id = str(finance.get("run_id") or "")
    if isinstance(run, dict) and run_id:
        # 财务可读工件已随十三表迁到交付物根（见 table_pack.default_artifact_dir），
        # 因此包含性校验必须用该根，否则附件会被 _safe_support_source 静默丢弃。
        finance_root = _finance_artifact_root(workspace_id, run_id)
        finance_containment_root = finance_root.parent.parent.resolve()
        try:
            candidates = sorted(finance_root.rglob("*.xlsx")) if finance_root.is_dir() else []
        except OSError:
            candidates = []
        copied_finance_count = 0
        for source in candidates:
            source = _safe_support_source(finance_containment_root, source)
            if source is None or source in seen:
                continue
            verified, reason = _verify_finance_workbook(source, run)
            if not verified:
                warnings.append({
                    "code": "FINANCE_XLSX_NOT_VERIFIED",
                    "filename": source.name,
                    "reason": reason,
                })
                continue
            source_hash_before, _source_size_before = _file_hash(source)
            copied_finance_count += 1
            relative = PurePosixPath("finance") / (
                f"{copied_finance_count:02d}_{_safe_filename(source.name)}"
            )
            target = temp_root.joinpath(*relative.parts)
            metadata = _copy_support_file(source, target)
            target_verified, target_reason = _verify_finance_workbook(target, run)
            source_hash_after, _source_size_after = _file_hash(source)
            if (
                not target_verified
                or metadata.get("sha256") != source_hash_before
                or metadata.get("sha256") != source_hash_after
            ):
                target.unlink(missing_ok=True)
                warnings.append({
                    "code": "FINANCE_XLSX_COPY_NOT_VERIFIED",
                    "filename": source.name,
                    "reason": target_reason or "source_changed_during_copy",
                })
                continue
            seen.add(source)
            copied.append({
                "name": relative.as_posix(),
                "role": "verified_finance_xlsx",
                "media_type": mimetypes.guess_type(source.name)[0]
                or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                **metadata,
            })

    appendix = (context.get("artifact_values") or {}).get("appendix_manifest")
    appendix_basis = basis.get("appendix_files") or []
    for index, row in enumerate(_appendix_path_rows(appendix), start=1):
        source = _safe_support_source(workspace_root, row.get("source_path"))
        if source is None:
            warnings.append({
                "code": "APPENDIX_FILE_UNAVAILABLE",
                "appendix_id": row.get("id"),
            })
            continue
        if source in seen:
            continue
        raw_declared_hash = str(row.get("declared_hash") or "").strip()
        declared_hash = _normalize_declared_hash(raw_declared_hash)
        if raw_declared_hash and not declared_hash:
            warnings.append({
                "code": "APPENDIX_HASH_INVALID",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        relative_source = source.relative_to(workspace_root).as_posix()
        expected_snapshot = next((
            item for item in appendix_basis
            if item.get("source") == relative_source
            and item.get("appendix_id") == row.get("id")
        ), None)
        if not isinstance(expected_snapshot, dict) or expected_snapshot.get("ok") is not True:
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        relative = PurePosixPath("appendices") / (
            f"{index:02d}_{_safe_filename(source.name)}"
        )
        target = temp_root.joinpath(*relative.parts)
        metadata = _copy_support_file(source, target)
        if metadata.get("sha256") != expected_snapshot.get("sha256"):
            target.unlink(missing_ok=True)
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        if declared_hash and metadata.get("sha256") != declared_hash:
            target.unlink(missing_ok=True)
            warnings.append({
                "code": "APPENDIX_HASH_MISMATCH",
                "appendix_id": row.get("id"),
                "filename": source.name,
            })
            continue
        seen.add(source)
        copied.append({
            "name": relative.as_posix(),
            "role": "verified_appendix",
            "appendix_id": row.get("id"),
            "media_type": mimetypes.guess_type(source.name)[0]
            or "application/octet-stream",
            **metadata,
        })
    return copied, warnings
