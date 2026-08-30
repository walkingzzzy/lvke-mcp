"""CSV / XLSX 导出与导出前置校验。"""

from __future__ import annotations

import csv
import hashlib
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from lvke_mcp.runtime.storage import require_safe_id

from .columns import (
    PACKAGE_STORE,
    _export_root,
)

from .query import (
    _CSV_PREVIEW_BANNER,
    _XLSX_PREVIEW_BANNER,
    _blocked,
    _release_grade,
    _result,
)

from .render import (
    _package,
)

from .rows import (
    _rows,
    _table_contract,
)


_SUPPLEMENTAL_COLUMNS: dict[str, tuple[tuple[str, str], ...]] = {
    "monthly_income_statement": (
        ("month", "月序号"), ("period_start", "期间开始日期"), ("period_end", "期间结束日期"),
        ("revenue_wan", "营业收入（万元）"), ("hotel_revenue_wan", "酒店收入（万元）"),
        ("lease_revenue_wan", "租赁收入（万元）"), ("operating_cost_wan", "经营成本（万元）"),
        ("depreciation_wan", "折旧（万元）"), ("interest_wan", "利息（万元）"),
        ("income_tax_wan", "所得税（万元）"), ("net_profit_wan", "净利润（万元）"),
    ),
    "monthly_balance_sheet": (
        ("month", "月序号"), ("period_start", "期间开始日期"), ("period_end", "期间结束日期"),
        ("cash_wan", "货币资金（万元）"), ("fixed_asset_net_wan", "固定资产净值（万元）"),
        ("total_assets_wan", "资产合计（万元）"), ("debt_wan", "有息负债（万元）"),
        ("equity_wan", "所有者权益（万元）"),
        ("total_liabilities_equity_wan", "负债和权益合计（万元）"),
    ),
}


def _export_manifest(
    payload: dict[str, Any],
    *,
    package_id: str,
    export_format: str,
    files: dict[str, str],
) -> dict[str, Any]:
    return {
        "manifest_schema": "acquisition_monthly_export_manifest.v1",
        "export_format": export_format,
        "package_id": package_id,
        "run_id": payload.get("run_id"),
        "spec_id": payload.get("spec_id"),
        "spec_hash": payload.get("spec_hash"),
        "input_hash": payload.get("input_hash"),
        "model_version": payload.get("model_version"),
        "monthly_driver_manifest": payload.get("monthly_driver_manifest") or {},
        "operating_calendar": payload.get("operating_calendar") or {},
        "annual_reconciliation": payload.get("annual_reconciliation") or [],
        "evidence_policy": payload.get("evidence_policy"),
        "evidence_origin": payload.get("evidence_origin"),
        "project_fact_certified": bool(payload.get("project_fact_certified", False)),
        "formal_promotion": payload.get("formal_promotion"),
        "lineage": payload.get("lineage"),
        "files": files,
    }


def _ensure_exportable(payload: dict[str, Any]) -> dict[str, Any] | None:
    if (payload.get("integrity") or {}).get("status") != "passed":
        return _blocked("TABLE_PACKAGE_INCOMPLETE", "收购十三表列级完整性或勾稽校验未通过")
    return None


def _export_cell(field: str, value: Any) -> Any:
    if field == "dynamic_payback_status":
        return {"recovered": "已回收", "not_recovered": "未回收"}.get(value, value)
    return "" if value is None else value


def export_csv(
    workspace_id: str,
    package_id: str,
) -> dict[str, Any]:
    record, payload = _package(
        workspace_id,
        package_id,
    )
    if record is None:
        return _blocked("TABLE_PACKAGE_NOT_FOUND", "未找到收购十三表 package")
    blocked = _ensure_exportable(payload)
    if blocked is not None:
        return blocked
    definitions, columns_by_key, _required = _table_contract(
        str(payload.get("asset_type") or "hotel_lease")
    )
    directory = (
        _export_root(workspace_id)
        / "csv"
        / require_safe_id(package_id, "package_id")
    )
    directory.mkdir(parents=True, exist_ok=True)
    grade, grade_reasons = _release_grade(payload)
    preview = grade == "technical_preview"
    uris: list[str] = []
    hashes: dict[str, str] = {}
    export_tables = [*definitions, *[
        (key, key) for key in _SUPPLEMENTAL_COLUMNS if (payload.get("tables") or {}).get(key)
    ]]
    for key, _name in export_tables:
        target = directory / f"{key}.csv"
        columns = columns_by_key.get(key) or _SUPPLEMENTAL_COLUMNS[key]
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\r\n")
            if preview:
                writer.writerow([_CSV_PREVIEW_BANNER])
            writer.writerow([label for _field, label in columns])
            for row in _rows((payload.get("tables") or {}).get(key)):
                writer.writerow([_export_cell(field, row.get(field)) for field, _label in columns])
        hashes[key] = "sha256:" + hashlib.sha256(target.read_bytes()).hexdigest()
        uris.append(
            PACKAGE_STORE.uri(workspace_id, package_id)
            + f"/csv/{key}"
        )
    manifest = _export_manifest(
        payload,
        package_id=package_id,
        export_format="csv",
        files=hashes,
    )
    manifest_target = directory / "manifest.json"
    manifest_target.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    manifest_hash = "sha256:" + hashlib.sha256(manifest_target.read_bytes()).hexdigest()
    manifest_uri = PACKAGE_STORE.uri(workspace_id, package_id) + "/csv/manifest"
    uris.append(manifest_uri)
    # release_grade / technical_preview / formal_usable / release_limitations
    # 与预览 warning 由 _result 统一给出（render/get/list/export 共用同一口径），
    # 这里只补导出专有字段。
    result = _result(record)
    result.update({
        "csv_resource_uris": uris,
        "csv_hashes": hashes,
        "monthly_export_manifest": manifest,
        "monthly_export_manifest_hash": manifest_hash,
        "monthly_export_manifest_uri": manifest_uri,
        # 交付物落盘绝对目录：收购各表 CSV 均在此目录下。
        "deliverable_path": str(directory),
        "resource_uris": [*result["resource_uris"], *uris],
    })
    return result


def export_xlsx(
    workspace_id: str,
    package_id: str,
) -> dict[str, Any]:
    record, payload = _package(
        workspace_id,
        package_id,
    )
    if record is None:
        return _blocked("TABLE_PACKAGE_NOT_FOUND", "未找到收购十三表 package")
    blocked = _ensure_exportable(payload)
    if blocked is not None:
        return blocked
    definitions, columns_by_key, _required = _table_contract(
        str(payload.get("asset_type") or "hotel_lease")
    )
    directory = _export_root(workspace_id) / "xlsx"
    directory.mkdir(parents=True, exist_ok=True)
    grade, grade_reasons = _release_grade(payload)
    preview = grade == "technical_preview"
    # 与通用财务域一致：预览件的文件名也带 .technical 后缀，让文件在磁盘上
    # 就能与正式件区分，而不是只靠响应字段。
    suffix = ".technical.xlsx" if preview else ".xlsx"
    target = directory / f"{require_safe_id(package_id, 'package_id')}{suffix}"
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")

    export_tables = [*definitions, *[
        (key, "月度利润表" if key == "monthly_income_statement" else "月度资产负债表")
        for key in _SUPPLEMENTAL_COLUMNS if (payload.get("tables") or {}).get(key)
    ]]
    for index, (key, name) in enumerate(export_tables, 1):
        sheet = workbook.create_sheet(title=f"{index:02d}-{name}"[:31])
        columns = columns_by_key.get(key) or _SUPPLEMENTAL_COLUMNS[key]
        if preview:
            sheet.append([_XLSX_PREVIEW_BANNER])
        sheet.append([label for _field, label in columns])
        for row in _rows((payload.get("tables") or {}).get(key)):
            sheet.append([_export_cell(field, row.get(field)) for field, _label in columns])
        # 预览横幅占掉第 1 行，表头与数据整体下移一行；冻结行与数字格式
        # 必须跟着移，否则会把横幅当表头、把表头当数据格式化。
        header_row = 2 if preview else 1
        first_data_row = header_row + 1
        sheet.freeze_panes = f"A{first_data_row}"
        sheet.auto_filter.ref = (
            f"A{header_row}:{sheet.cell(row=sheet.max_row, column=len(columns)).coordinate}"
            if sheet.max_row >= header_row
            else sheet.dimensions
        )
        sheet.sheet_view.showGridLines = False
        for column_index, (field, _label) in enumerate(columns, 1):
            if field.endswith("_wan"):
                for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=first_data_row):
                    for item in cell:
                        item.number_format = '#,##0.00;[Red](#,##0.00);-'
            elif field in {"financing_ratio", "residual_rate", "occupancy", "target_irr"}:
                for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=first_data_row):
                    for item in cell:
                        item.number_format = '0.0%'
        for column_cells in sheet.columns:
            maximum = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(maximum + 2, 10), 32)

    for sheet in workbook.worksheets:
        header_row = 2 if preview else 1
        if preview:
            banner_cell = sheet.cell(row=1, column=1)
            banner_cell.font = Font(color="9C0006", bold=True)
            banner_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[1].height = 22
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
        sheet.row_dimensions[header_row].height = 24
        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for column_cells in sheet.columns:
            maximum = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(maximum + 2, 10), 36)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(target)
    digest = "sha256:" + hashlib.sha256(target.read_bytes()).hexdigest()
    manifest = _export_manifest(
        payload,
        package_id=package_id,
        export_format="xlsx",
        files={"workbook": digest},
    )
    manifest_target = target.with_suffix(target.suffix + ".manifest.json")
    manifest_target.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    manifest_hash = "sha256:" + hashlib.sha256(manifest_target.read_bytes()).hexdigest()
    uri = PACKAGE_STORE.uri(workspace_id, package_id) + "/xlsx"
    manifest_uri = PACKAGE_STORE.uri(workspace_id, package_id) + "/xlsx/manifest"
    # 等级与限制说明同样由 _result 统一给出；XLSX 侧额外把横幅写进 A1
    # 与文件名后缀（见上方 preview 分支），那是文件内标记，不是响应字段。
    result = _result(record)
    result.update({
        "xlsx_resource_uri": uri,
        "xlsx_hash": digest,
        "monthly_export_manifest": manifest,
        "monthly_export_manifest_hash": manifest_hash,
        "monthly_export_manifest_uri": manifest_uri,
        # 交付物落盘绝对路径。
        "deliverable_path": str(target),
        "resource_uris": [*result["resource_uris"], uri, manifest_uri],
        "warnings": [
            *list(result.get("warnings") or []),
            *([f"技术预览：{_XLSX_PREVIEW_BANNER}"] if preview else []),
        ],
    })
    return result
