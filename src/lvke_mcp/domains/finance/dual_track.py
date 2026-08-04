"""Engine vs reference dual-track compare (P2-3)."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional, Union


DEFAULT_AMOUNT_TOL = 0.01  # 万元（正式真实样本验收）
DEFAULT_RATIO_TOL = 0.01   # 百分点


def _is_ratio_key(key: str) -> bool:
    k = key.lower()
    return any(s in k for s in ("irr", "pct", "rate", "bep", "ratio", "dscr", "icr"))


def _tol_for(key: str, amount_tol: float, ratio_tol: float) -> float:
    k = key.lower()
    # DSCR/ICR are pure ratios (not percent); allow 0.05 absolute (S2 T-debt)
    if "dscr" in k or "icr" in k:
        return max(ratio_tol, 0.05)
    return ratio_tol if _is_ratio_key(key) else amount_tol


def extract_engine_indicators(fin: dict[str, Any]) -> dict[str, float]:
    ind = fin.get("indicators") or {}
    inv = fin.get("investment") or {}
    fund = fin.get("funding") or {}
    out: dict[str, float] = {}
    mapping = {
        "total_investment": inv.get("total"),
        "construction_investment": inv.get("construction"),
        "working_capital": inv.get("working_capital"),
        "interest_during_construction": inv.get("interest"),
        "equity_capital": fund.get("capital"),
        "loan": fund.get("loan"),
        "project_irr_pct": ind.get("project_irr_pct"),
        "capital_irr_pct": (fin.get("annual") or {}).get("capital_irr_pct"),
        "npv_wan": ind.get("npv_wan"),
        "static_payback_years": ind.get("static_payback_years"),
        "dynamic_payback_years": ind.get("dynamic_payback_years"),
        "bep_pct": ind.get("bep_pct"),
    }
    for k, v in mapping.items():
        if v is None:
            continue
        try:
            out[k] = float(v)
        except (TypeError, ValueError):
            continue
    return out


def extract_period_map(fin: dict[str, Any], tables: Optional[list[str]] = None) -> dict[str, float]:
    """Flatten annual tables to locator keys: table.element.Y{year}."""
    annual = fin.get("annual") or {}
    want = set(tables or [
        "income_statement", "total_cost", "project_cashflow",
        "capital_cashflow", "financial_plan", "debt_service",
    ])
    out: dict[str, float] = {}
    for table, rows in annual.items():
        if table not in want or not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            year = row.get("year") if row.get("year") is not None else row.get("period")
            if year is None:
                continue
            for k, v in row.items():
                if k in ("year", "period", "phase", "method", "in_grace"):
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                out[f"{table}.{k}.Y{year}"] = fv
    return out


def load_reference(reference: Union[dict[str, Any], str, Path]) -> dict[str, Any]:
    if isinstance(reference, dict):
        return reference
    p = Path(reference)
    if p.suffix.lower() in {".json"}:
        return json.loads(p.read_text(encoding="utf-8"))
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_reference_xlsx(p)
    raise ValueError(f"unsupported reference type: {reference!r}")


def _load_reference_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except Exception:  # noqa: BLE001
        openpyxl = None  # type: ignore[assignment]
    if openpyxl is not None:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            indicators: dict[str, float] = {}
            if "Indicators" in wb.sheetnames:
                ws = wb["Indicators"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        indicators[str(row[0])] = float(row[1])
                    except (TypeError, ValueError):
                        continue
                return {"indicators": indicators, "periods": {}}
        finally:
            wb.close()

    # Native vendor workbooks have appendix sheets rather than ``Indicators``.
    from lvke_mcp.domains.finance.vendor_import import build_reference_pack

    pack = build_reference_pack(path)
    return {
        "indicators": dict(pack.get("indicators") or {}),
        "periods": dict(pack.get("periods") or {}),
        "reference_pack": pack,
    }


def _deviation_grade(engine_value: float, ref_value: float, delta: float, tol: float) -> dict[str, Any]:
    """Grade overlap using the review thresholds: 15% / 30%."""
    if abs(delta) <= tol:
        return {"deviation_pct": 0.0 if ref_value == 0 else abs(delta / ref_value) * 100.0,
                "verdict": "converged", "red_flag": False}
    if abs(ref_value) < 1e-12:
        return {"deviation_pct": None, "verdict": "red_flag", "red_flag": True}
    deviation = abs(delta / ref_value) * 100.0
    if deviation <= 15.0:
        verdict = "converged"
    elif deviation <= 30.0:
        verdict = "explain"
    else:
        verdict = "red_flag"
    return {"deviation_pct": deviation, "verdict": verdict, "red_flag": verdict == "red_flag"}


def compare_engine_to_reference(
    fin: dict[str, Any],
    reference: Union[dict[str, Any], str, Path],
    *,
    amount_tol: float = DEFAULT_AMOUNT_TOL,
    ratio_tol: float = DEFAULT_RATIO_TOL,
    include_periods: bool = True,
) -> dict[str, Any]:
    """Compare engine result to reference indicators (+ optional period cells)."""
    ref = load_reference(reference)
    eng_ind = extract_engine_indicators(fin)
    ref_ind = {}
    raw_ind = ref.get("indicators") or {}
    for k, v in raw_ind.items():
        try:
            ref_ind[str(k)] = float(v)
        except (TypeError, ValueError):
            continue

    matched: list[dict[str, Any]] = []
    mismatched: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []

    keys = sorted(set(eng_ind) | set(ref_ind))
    for key in keys:
        locator = f"indicators.{key}"
        if key not in eng_ind:
            missing.append({"locator": locator, "side": "engine", "ref_value": ref_ind.get(key)})
            continue
        if key not in ref_ind:
            missing.append({"locator": locator, "side": "reference", "engine_value": eng_ind.get(key)})
            continue
        ev, rv = eng_ind[key], ref_ind[key]
        tol = _tol_for(key, amount_tol, ratio_tol)
        delta = ev - rv
        item = {
            "locator": locator,
            "engine_value": ev,
            "ref_value": rv,
            "delta": delta,
            "tol": tol,
        }
        item.update(_deviation_grade(ev, rv, delta, tol))
        if abs(delta) <= tol:
            matched.append(item)
        else:
            mismatched.append(item)

    if include_periods and (ref.get("periods") or ref.get("period_results")):
        eng_p = extract_period_map(fin)
        ref_p_raw = ref.get("periods") or ref.get("period_results") or {}
        ref_p: dict[str, float] = {}
        for k, v in ref_p_raw.items():
            try:
                ref_p[str(k)] = float(v)
            except (TypeError, ValueError):
                continue
        for key in sorted(set(eng_p) | set(ref_p)):
            locator = f"period.{key}"
            if key not in eng_p:
                missing.append({"locator": locator, "side": "engine", "ref_value": ref_p.get(key)})
                continue
            if key not in ref_p:
                # reference may be partial — only flag if reference claimed the key
                continue
            ev, rv = eng_p[key], ref_p[key]
            tol = _tol_for(key, amount_tol, ratio_tol)
            delta = ev - rv
            item = {
                "locator": locator,
                "engine_value": ev,
                "ref_value": rv,
                "delta": delta,
                "tol": tol,
            }
            item.update(_deviation_grade(ev, rv, delta, tol))
            if abs(delta) <= tol:
                matched.append(item)
            else:
                mismatched.append(item)

    red_flags = [item for item in mismatched if item.get("red_flag")]
    needs_explanation = [item for item in mismatched if item.get("verdict") == "explain"]
    return {
        "ok": len(mismatched) == 0 and not any(m.get("side") == "engine" for m in missing),
        "matched": matched,
        "mismatched": mismatched,
        "missing": missing,
        "red_flags": red_flags,
        "needs_explanation": needs_explanation,
        "summary": {
            "matched": len(matched),
            "mismatched": len(mismatched),
            "missing": len(missing),
            "red_flags": len(red_flags),
            "needs_explanation": len(needs_explanation),
        },
    }


def build_reference_from_engine(fin: dict[str, Any], *, include_periods: bool = False) -> dict[str, Any]:
    """Helper to snapshot engine outputs as a dual-track reference fixture."""
    ref: dict[str, Any] = {"indicators": extract_engine_indicators(fin)}
    if include_periods:
        ref["periods"] = extract_period_map(fin)
    return ref
