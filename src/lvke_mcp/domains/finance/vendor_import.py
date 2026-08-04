"""Vendor financial-workbook import and cleanup detection.

The imported workbook is a *reference*, never a calculation source of truth.  This
module extracts cached values and formulas, maps heterogeneous vendor sheets to the
13-table business dictionary, and produces deterministic inputs for our own engine.
It never writes back to the vendor workbook.
"""

from __future__ import annotations

import hashlib
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Optional


REFERENCE_PACK_VERSION = "vendor_reference.v2"
MAX_ROWS = 500
MAX_COLS = 60
IRR_RESIDUAL_TOL_WAN = 0.1


class VendorImportError(RuntimeError):
    """Raised when a vendor workbook cannot be read as a reference pack."""


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _col_letter(index: int) -> str:
    out = ""
    value = int(index)
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        out = chr(65 + remainder) + out
    return out


_CELL_RE = re.compile(r"^([A-Z]{1,3})(\d+)$")


def _col_index(letter: str) -> int:
    out = 0
    for ch in letter:
        out = out * 26 + ord(ch) - 64
    return out


def _cell_parts(cell: str) -> tuple[int, int]:
    match = _CELL_RE.match(str(cell).replace("$", ""))
    if not match:
        return 0, 0
    return int(match.group(2)), _col_index(match.group(1))


def _to_float(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _norm(value: Any) -> str:
    return re.sub(r"[\s　]+", "", str(value or "")).replace("：", ":")


def _read_value_sheets(path: Path, sheet_names: list[str]) -> tuple[dict[str, Any], str]:
    """Read cached values once with openpyxl, falling back to excel_bridge reader."""
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            result: dict[str, Any] = {}
            for name in workbook.sheetnames:
                worksheet = workbook[name]
                values: dict[str, Any] = {}
                row_limit = min(int(worksheet.max_row or 0), MAX_ROWS)
                col_limit = min(int(worksheet.max_column or 0), MAX_COLS)
                for row in worksheet.iter_rows(
                    min_row=1, max_row=row_limit, min_col=1, max_col=col_limit
                ):
                    for cell in row:
                        value = _jsonable(cell.value)
                        if value is not None:
                            values[cell.coordinate] = value
                result[name] = {
                    "values": values,
                    "max_row": int(worksheet.max_row or 0),
                    "max_col": int(worksheet.max_column or 0),
                    "value_backend": "openpyxl",
                }
            return result, "openpyxl"
        finally:
            workbook.close()
    except Exception as openpyxl_error:  # noqa: BLE001
        try:
            from lvke_mcp.servers.excel_bridge.reader import pick_backend

            backend = pick_backend()
            result = {}
            for name in sheet_names:
                read = backend.read_sheet(path, name, MAX_ROWS, MAX_COLS)
                values = {}
                for row_index, row in enumerate(read.rows, 1):
                    for col_index, value in enumerate(row, 1):
                        if value is not None:
                            values[f"{_col_letter(col_index)}{row_index}"] = _jsonable(value)
                result[name] = {
                    "values": values,
                    "max_row": read.row_count,
                    "max_col": read.col_count,
                    "value_backend": read.backend,
                }
            return result, str(getattr(backend, "name", "excel_bridge-reader"))
        except Exception as fallback_error:  # noqa: BLE001
            raise VendorImportError(
                "无法读取甲方工作簿缓存值："
                f"openpyxl={type(openpyxl_error).__name__}; "
                f"fallback={type(fallback_error).__name__}: {fallback_error}"
            ) from fallback_error


def _header_text(sheet: dict[str, Any], max_rows: int = 4) -> str:
    parts = []
    for cell, value in (sheet.get("values") or {}).items():
        row, _ = _cell_parts(cell)
        if row <= max_rows and isinstance(value, str):
            parts.append(value)
    return " ".join(parts)


def _sheet_is_nonempty(sheet: dict[str, Any]) -> bool:
    """Return whether a vendor sheet contains any business-visible content.

    Formula-only sheets are non-empty even when the reader cannot obtain cached
    values.  This deliberately errs on the side of requiring an explicit review
    decision instead of silently dropping a vendor worksheet.
    """

    for value in (sheet.get("values") or {}).values():
        if value is not None and (not isinstance(value, str) or value.strip()):
            return True
    return bool(sheet.get("formulas"))


def _infer_business(sheet_name: str, sheet: dict[str, Any]) -> str:
    """Infer canonical business name; business meaning wins over bare appendix no.

    表名 + 表头(前 4 行)联合判定。**强业务特征**(投资复核 / 利润分配 / 还本付息 /
    现金流)先判定：利润表、现金流表的表头会同时出现"营业收入""税金"列标题，若把
    收入税金规则放前面会把它们误判为收入税金表(FIN-MAP-001，房地产利润表实测错配为
    附表5)。因此把这些强特征提到收入税金规则之前，并给收入税金规则加"非利润分配"
    排除项。**注意**："总成本费用"必须保持在工资/折旧/摊销/原材料/燃料之前，否则总成本
    表会被其列标题(工资及福利费/折旧费…)误判为工资/折旧子表。
    """
    name = _norm(sheet_name)
    raw = f"{sheet_name} {_header_text(sheet)}"
    text = _norm(raw)
    # ── 强业务特征优先(表名/表头前 4 行出现即高置信) ──
    if "投资复核" in text or "投资估算复核" in text:
        return "投资估算复核表"
    if name == "项目财务分析" or "净现值计算方式" in text:
        return "投资估算复核表"
    if "利润" in text and "分配" in text:
        return "利润与利润分配表"
    if "还本付息" in text or "还款付息" in text:
        return "还款付息测算表"
    if "项目投资现金流" in text:
        return "项目投资现金流量表"
    if "项目资本金" in text and ("现金流" in text or name == "项目资本金"):
        return "项目资本金流量表"
    if "敏感度分析" in text or "敏感性分析" in text:
        return "单因素敏感性分析表"
    if "主要经济指标汇总" in text or "指标名称" in text:
        return "主要技术经济指标汇总表"
    if "结构方案" in text and "综合判断" in text:
        return "建筑/技术方案比选表"
    # ── 投资/融资侧 ──
    if "固定资产投资估算" in text:
        return "固定资产投资估算表"
    if "建设期" in text and "利息" in text:
        return "建设期贷款利息表"
    if "流动资金估算" in text or "流动资金测算" in text:
        return "流动资金估算表"
    if "资金筹措" in text and ("使用计划" in text or "总投资" in text):
        return "投资使用计划与资金筹措表"
    # ── 收入税金表：放在利润/现金流之后，显式排除利润分配特征 ──
    if (
        "营业收入" in text
        and ("税金" in text or "增值税" in text)
        and not ("利润" in text and "分配" in text)
    ):
        return "营业收入、税金及附加和增值税估算表"
    # ── 成本侧：总成本费用必须先于工资/折旧/摊销/原材料/燃料子表判定 ──
    if "总成本费用" in text:
        return "总成本费用估算表"
    if "工资" in text and ("福利" in text or "附加" in text):
        return "工资及附加估算表"
    if "折旧" in text:
        return "固定资产折旧费估算表"
    if "摊销" in text:
        return "无形资产和其他资产摊销估算表"
    if "原材料" in text:
        return "外购原材料费估算表"
    if "燃料" in text and "动力" in text:
        return "外购燃料和动力费估算表"
    return ""


def _rows(sheet: dict[str, Any]) -> dict[int, dict[int, tuple[str, Any]]]:
    grouped: dict[int, dict[int, tuple[str, Any]]] = defaultdict(dict)
    for cell, value in (sheet.get("values") or {}).items():
        row, col = _cell_parts(cell)
        if row and col:
            grouped[row][col] = (cell, value)
    return grouped


def _row_label(row: dict[int, tuple[str, Any]], max_label_col: int = 4) -> str:
    return " / ".join(
        str(value).strip()
        for col, (_, value) in sorted(row.items())
        if col <= max_label_col and isinstance(value, str) and str(value).strip()
    )


def _longest_period_sequence(row: dict[int, tuple[str, Any]]) -> dict[int, int]:
    points = []
    for col, (_, value) in sorted(row.items()):
        number = _to_float(value)
        if number is None or abs(number - round(number)) > 1e-9:
            continue
        integer = int(round(number))
        if 0 <= integer <= 100:
            points.append((col, integer))
    best: list[tuple[int, int]] = []
    current: list[tuple[int, int]] = []
    for point in points:
        if current and (point[0] != current[-1][0] + 1 or point[1] != current[-1][1] + 1):
            if len(current) > len(best):
                best = current
            current = []
        current.append(point)
    if len(current) > len(best):
        best = current
    return dict(best) if len(best) >= 3 else {}


def _period_headers(sheet: dict[str, Any], before_row: int) -> dict[int, int]:
    best: dict[int, int] = {}
    for row_index, row in _rows(sheet).items():
        if row_index >= before_row:
            continue
        candidate = _longest_period_sequence(row)
        if len(candidate) > len(best):
            best = candidate
    return best


def _series_for_row(sheet: dict[str, Any], row_index: int) -> list[dict[str, Any]]:
    row = _rows(sheet).get(row_index) or {}
    headers = _period_headers(sheet, row_index)
    result = []
    for col, period in sorted(headers.items()):
        cell_value = row.get(col)
        if not cell_value:
            continue
        value = _to_float(cell_value[1])
        if value is None:
            continue
        result.append({"period": period, "cell": cell_value[0], "value": value})
    return result


def _find_row(
    sheet: Optional[dict[str, Any]],
    keywords: Iterable[str],
    *,
    exclude: Iterable[str] = (),
) -> tuple[int, str]:
    if not sheet:
        return 0, ""
    wanted = [_norm(x) for x in keywords]
    excluded = [_norm(x) for x in exclude]
    for row_index, row in sorted(_rows(sheet).items()):
        label = _row_label(row)
        normalized = _norm(label)
        if wanted and all(word in normalized for word in wanted) and not any(
            word in normalized for word in excluded
        ):
            return row_index, label
    return 0, ""


def _series_by_label(
    sheet: Optional[dict[str, Any]],
    keywords: Iterable[str],
    *,
    exclude: Iterable[str] = (),
) -> list[dict[str, Any]]:
    row_index, _ = _find_row(sheet, keywords, exclude=exclude)
    return _series_for_row(sheet, row_index) if row_index else []


def _total_for_row(sheet: dict[str, Any], row_index: int) -> Optional[float]:
    rows = _rows(sheet)
    row = rows.get(row_index) or {}
    total_columns: list[int] = []
    for candidate_row, cells in rows.items():
        if candidate_row >= row_index:
            continue
        for col, (_, value) in cells.items():
            if isinstance(value, str) and "合计" in _norm(value):
                total_columns.append(col)
    for col in sorted(set(total_columns), reverse=True):
        if col in row:
            value = _to_float(row[col][1])
            if value is not None:
                return value
    numbers = []
    for col, (_, raw) in row.items():
        if col <= 2:
            continue
        value = _to_float(raw)
        if value is not None:
            numbers.append(value)
    return max(numbers, key=lambda item: abs(item)) if numbers else None


def _total_by_label(
    sheet: Optional[dict[str, Any]],
    keywords: Iterable[str],
    *,
    exclude: Iterable[str] = (),
) -> Optional[float]:
    if not sheet:
        return None
    row_index, _ = _find_row(sheet, keywords, exclude=exclude)
    return _total_for_row(sheet, row_index) if row_index else None


def _find_mapped_sheet(reference_pack: dict[str, Any], business: str) -> Optional[dict[str, Any]]:
    for sheet in (reference_pack.get("sheets") or {}).values():
        if sheet.get("business") == business:
            return sheet
    return None


def _review_trial_rates(review_sheet: Optional[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    if not review_sheet:
        return {}
    rows = _rows(review_sheet)
    formulas = review_sheet.get("formulas") or {}
    result: dict[str, dict[str, Any]] = {}
    capital_marker = 0
    trial_index = 0
    for row_index, row in sorted(rows.items()):
        if "资本金内部收益率" in _norm(_row_label(row)):
            capital_marker = row_index
        if "试算使净现值之和为0" not in _norm(_row_label(row)):
            continue
        candidate = None
        for col, (cell, raw) in sorted(row.items()):
            if col < 3:
                continue
            value = _to_float(raw)
            if value is not None:
                candidate = (cell, value)
                break
        if not candidate:
            continue
        trial_index += 1
        next_label = _norm(_row_label(rows.get(row_index + 1) or {}))
        # Several vendor templates omit the capital-IRR section title entirely;
        # their third trial block is nevertheless the capital cash-flow trial.
        if (capital_marker and row_index > capital_marker) or trial_index >= 3:
            key = "capital_after_tax"
        elif "税前" in next_label:
            key = "project_pre_tax"
        else:
            key = "project_after_tax"
        result[key] = {
            "rate": candidate[1],
            "rate_pct": candidate[1] * 100.0 if abs(candidate[1]) <= 1.0 else candidate[1],
            "cell": candidate[0],
            "hardcoded": candidate[0] not in formulas,
        }
    return result


def _first_numeric_on_row(sheet: Optional[dict[str, Any]], row_index: int) -> Optional[float]:
    if not sheet or not row_index:
        return None
    for col, (_, raw) in sorted((_rows(sheet).get(row_index) or {}).items()):
        if col < 3:
            continue
        value = _to_float(raw)
        if value is not None:
            return value
    return None


def _row_value(
    sheet: dict[str, Any], row_index: int, col_index: int,
) -> tuple[str, Any]:
    return (_rows(sheet).get(row_index) or {}).get(col_index, ("", None))


def _extract_indicator_summary(reference_pack: dict[str, Any]) -> dict[str, Any]:
    """Structure the vendor's key-indicator sheet with cell-level provenance.

    The summary is a read-only reference surface.  Its values never replace an
    approved engine run, but they are consumed by reference-track comparison and
    reverse-reconciliation checks instead of being left as an unmapped worksheet.
    """

    sheet = _find_mapped_sheet(reference_pack, "主要技术经济指标汇总表")
    if not sheet:
        return {"available": False, "rows": [], "normalized": {}, "warnings": ["未找到指标汇总表"]}
    sheet_name = str((sheet.get("mapping") or {}).get("vendor_sheet_actual") or "")
    rows = _rows(sheet)
    header_row = 0
    for row_index, cells in sorted(rows.items()):
        text = _norm(" ".join(str(value) for _, value in cells.values() if value is not None))
        if "指标名称" in text and ("数据" in text or "数值" in text):
            header_row = row_index
            break
    if not header_row:
        return {
            "available": False,
            "sheet_name": sheet_name,
            "rows": [],
            "normalized": {},
            "warnings": ["指标汇总表缺少可识别表头"],
        }

    structured: list[dict[str, Any]] = []
    normalized: dict[str, dict[str, Any]] = {}

    def bind(key: str, value: Any, locator: str, label: str, unit: str) -> None:
        number = _to_float(value)
        if number is None or key in normalized:
            return
        normalized[key] = {
            "value": number,
            "locator": locator,
            "label": label,
            "unit": unit,
        }

    for row_index, cells in sorted(rows.items()):
        if row_index <= header_row:
            continue
        label_cell, label_raw = cells.get(2, ("", None))
        label = str(label_raw or "").strip()
        if not label:
            continue
        unit_cell, unit_raw = cells.get(3, ("", None))
        unit = str(unit_raw or "").strip()
        seq_cell, seq_raw = cells.get(1, ("", None))
        remark_cell, remark_raw = cells.get(6, ("", None))
        values = []
        for col_index, scope in ((4, "primary"), (5, "secondary")):
            cell, raw = cells.get(col_index, ("", None))
            if _to_float(raw) is None:
                continue
            values.append({
                "scope": scope,
                "value": _to_float(raw),
                "cell": cell,
                "locator": f"{sheet_name}!{cell}",
            })
        structured.append({
            "row": row_index,
            "seq": seq_raw,
            "indicator": label,
            "unit": unit,
            "values": values,
            "remark": remark_raw,
            "locators": {
                "seq": f"{sheet_name}!{seq_cell}" if seq_cell else "",
                "indicator": f"{sheet_name}!{label_cell}" if label_cell else "",
                "unit": f"{sheet_name}!{unit_cell}" if unit_cell else "",
                "remark": f"{sheet_name}!{remark_cell}" if remark_cell else "",
            },
        })
        primary = values[0] if values else None
        secondary = values[1] if len(values) > 1 else None
        text = _norm(label)
        if primary:
            if "项目总投资" in text and "资金来源" not in text:
                bind("total_investment", primary["value"], primary["locator"], label, unit)
            elif "固定资产投资" in text and "资金来源" not in text:
                bind("construction_investment", primary["value"], primary["locator"], label, unit)
            elif "流动资金" in text and ("新增" in text or "铺底" in text):
                bind("working_capital", primary["value"], primary["locator"], label, unit)
            elif "申请贷款" in text or "银行长期贷款" in text:
                bind("loan", primary["value"], primary["locator"], label, unit)
            elif ("自筹资金" in text or "企业自筹" in text) and "流动资金" not in text:
                bind("equity_capital", primary["value"], primary["locator"], label, unit)
            elif text == "建设期":
                bind("build_period_years", primary["value"], primary["locator"], label, unit)
            elif text in {"经营收入", "销售收入(不含税)", "营业收入"}:
                bind("revenue", primary["value"], primary["locator"], label, unit)
            elif "资本金内部收益率" in text:
                bind("capital_irr_pct", primary["value"], primary["locator"], label, unit)
            elif "项目财务内部收益率" in text:
                bind("project_irr_pre_tax_pct", primary["value"], primary["locator"], label, unit)
                if secondary:
                    bind("project_irr_pct", secondary["value"], secondary["locator"], label, unit)
            elif "项目财务净现值" in text:
                bind("npv_pre_tax_wan", primary["value"], primary["locator"], label, unit)
                if secondary:
                    bind("npv_wan", secondary["value"], secondary["locator"], label, unit)
            elif "项目投资回收期" in text:
                bind("payback_pre_tax_years", primary["value"], primary["locator"], label, unit)
                if secondary:
                    bind("static_payback_years", secondary["value"], secondary["locator"], label, unit)

    return {
        "available": bool(structured),
        "sheet_name": sheet_name,
        "header_row": header_row,
        "rows": structured,
        "normalized": normalized,
        "warnings": [],
    }


def _sensitivity_factor_code(label: str) -> str:
    text = _norm(label)
    if "建设投资" in text or "固定资产投资" in text:
        return "construction_investment"
    if "服务价格" in text:
        return "service_price"
    if "产品价格" in text or "销售价格" in text:
        return "product_price"
    if "直接成本" in text:
        return "direct_cost"
    if "运营负荷" in text or "生产负荷" in text or "经营负荷" in text:
        return "operating_load"
    return re.sub(r"[^0-9A-Za-z_]+", "_", text).strip("_").lower() or "unknown"


def _extract_sensitivity_summary(
    reference_pack: dict[str, Any],
    indicator_summary: dict[str, Any],
) -> dict[str, Any]:
    """Extract the real factory/cemetery sensitivity overview and provenance."""

    overview: Optional[dict[str, Any]] = None
    sheet_name = ""
    supporting: dict[str, str] = {}
    for name, candidate in (reference_pack.get("sheets") or {}).items():
        if candidate.get("business") != "单因素敏感性分析表":
            continue
        header = _norm(_header_text(candidate, max_rows=2))
        if "不确定性因数" in header and "内部收益率" in header:
            overview = candidate
            sheet_name = str(name)
            continue
        code = _sensitivity_factor_code(str(name))
        if code != "unknown":
            supporting.setdefault(code, str(name))
    if not overview:
        return {"available": False, "factors": [], "warnings": ["未找到敏感性汇总页"]}

    factors: list[dict[str, Any]] = []
    current: Optional[dict[str, Any]] = None
    base: dict[str, Any] = {}
    for row_index, cells in sorted(_rows(overview).items()):
        factor_cell, factor_raw = cells.get(1, ("", None))
        factor_label = str(factor_raw or "").strip()
        change_cell, change_raw = cells.get(2, ("", None))
        irr_cell, irr_raw = cells.get(3, ("", None))
        coefficient_cell, coefficient_raw = cells.get(4, ("", None))
        critical_cell, critical_raw = cells.get(5, ("", None))
        irr = _to_float(irr_raw)
        if factor_label and "基本方案" in _norm(factor_label) and irr is not None:
            base = {
                "irr_pct": irr,
                "row": row_index,
                "factor_locator": f"{sheet_name}!{factor_cell}",
                "irr_locator": f"{sheet_name}!{irr_cell}",
            }
            current = None
            continue
        change = _to_float(change_raw)
        if factor_label and change is not None and irr is not None:
            code = _sensitivity_factor_code(factor_label)
            current = {
                "factor": factor_label,
                "factor_code": code,
                "factor_locator": f"{sheet_name}!{factor_cell}",
                "supporting_sheet": supporting.get(code, ""),
                "critical_point_pct": None,
                "critical_point_locator": "",
                "scenarios": [],
            }
            factors.append(current)
        if current is None or change is None or irr is None:
            continue
        critical = _to_float(critical_raw)
        if critical is not None:
            current["critical_point_pct"] = critical
            current["critical_point_locator"] = f"{sheet_name}!{critical_cell}"
        current["scenarios"].append({
            "change_ratio": change,
            "change_pct": change * 100.0 if abs(change) <= 1.0 else change,
            "irr_pct": irr,
            "sensitivity_coefficient": _to_float(coefficient_raw),
            "row": row_index,
            "locators": {
                "change": f"{sheet_name}!{change_cell}",
                "irr": f"{sheet_name}!{irr_cell}",
                "coefficient": f"{sheet_name}!{coefficient_cell}" if coefficient_cell else "",
            },
        })

    expected = ((indicator_summary.get("normalized") or {}).get("project_irr_pre_tax_pct") or {})
    expected_value = _to_float(expected.get("value"))
    base_value = _to_float(base.get("irr_pct"))
    reconciliation = {
        "summary_value": expected_value,
        "summary_locator": expected.get("locator") or "",
        "sensitivity_value": base_value,
        "sensitivity_locator": base.get("irr_locator") or "",
        "tolerance_percentage_points": 0.01,
        "delta_percentage_points": (
            abs(base_value - expected_value)
            if base_value is not None and expected_value is not None else None
        ),
    }
    reconciliation["within_tolerance"] = (
        reconciliation["delta_percentage_points"] is not None
        and reconciliation["delta_percentage_points"] <= 0.01
    )
    warnings = []
    if not base:
        warnings.append("敏感性汇总页缺少基本方案IRR")
    if factors and any(len(item.get("scenarios") or []) < 2 for item in factors):
        warnings.append("至少一个敏感性因素缺少完整变动档位")
    if expected_value is not None and not reconciliation["within_tolerance"]:
        warnings.append("敏感性基准IRR与主要指标汇总表不一致")
    return {
        "available": bool(base and factors),
        "sheet_name": sheet_name,
        "base": base,
        "factors": factors,
        "supporting_sheets": supporting,
        "base_reconciliation": reconciliation,
        "warnings": warnings,
    }


def _extract_reference_data(reference_pack: dict[str, Any]) -> None:
    indicator_summary = _extract_indicator_summary(reference_pack)
    reference_pack["indicator_summary"] = indicator_summary
    reference_pack["sensitivity_summary"] = _extract_sensitivity_summary(
        reference_pack, indicator_summary
    )
    funding = _find_mapped_sheet(reference_pack, "投资使用计划与资金筹措表")
    project_cf = _find_mapped_sheet(reference_pack, "项目投资现金流量表")
    capital_cf = _find_mapped_sheet(reference_pack, "项目资本金流量表")
    debt = _find_mapped_sheet(reference_pack, "还款付息测算表")
    review = _find_mapped_sheet(reference_pack, "投资估算复核表")

    project_pre = _series_by_label(
        project_cf, ["所得税前净现金流量"], exclude=["累计"]
    )
    project_after = _series_by_label(
        project_cf, ["所得税后净现金流量"], exclude=["累计"]
    )
    capital_after = _series_by_label(
        capital_cf, ["所得税后净现金流量"], exclude=["累计"]
    )
    project_revenue = _series_by_label(project_cf, ["营业收入"])
    project_cost = _series_by_label(project_cf, ["经营成本"])
    icr = _series_by_label(debt, ["利息备付率"])
    dscr = _series_by_label(debt, ["偿债备付率"])
    principal = _series_by_label(debt, ["其中:还本", "还本"], exclude=["还本付息", "当期还本付息"])

    cashflows = {
        "project_pre_tax": project_pre,
        "project_after_tax": project_after,
        "capital_after_tax": capital_after,
        "project_revenue": project_revenue,
        "project_operating_cost": project_cost,
        "debt_icr": icr,
        "debt_dscr": dscr,
        "debt_principal": principal,
    }
    reference_pack["cashflows"] = cashflows

    total = _total_by_label(funding, ["总投资"], exclude=["使用计划", "资金筹措表"])
    interest = _total_by_label(funding, ["建设期", "利息"])
    working = _total_by_label(funding, ["流动资金"])
    capital = _total_by_label(funding, ["项目资本金"])
    debt_funding_total = _total_by_label(funding, ["债务资金"])
    long_term_loan = _total_by_label(funding, ["银行长期借款"])
    short_term_loan_schedule = _series_by_label(funding, ["银行短期借款"])
    short_term_loan = next(
        (float(item["value"]) for item in short_term_loan_schedule if float(item.get("value") or 0.0) > 0),
        None,
    )
    loan = debt_funding_total if debt_funding_total is not None else long_term_loan
    loan_selection = "debt_funding_total"
    # Some vendor sheets put revolving working-capital borrowing below the
    # project-funding total.  In that layout the row labelled “债务资金” can be
    # larger than total investment minus equity, while the explicitly labelled
    # long-term bank loan reconciles exactly.  Keep every source value, but use
    # the evidenced long-term principal for the deterministic project model.
    if (
        total is not None and capital is not None and long_term_loan is not None
        and abs((float(capital) + float(long_term_loan)) - float(total))
        <= max(abs(float(total)) * 1e-6, 0.01)
        and (
            debt_funding_total is None
            or abs((float(capital) + float(debt_funding_total)) - float(total))
            > max(abs(float(total)) * 1e-6, 0.01)
        )
    ):
        loan = long_term_loan
        loan_selection = "long_term_loan_reconciled_to_total"

    trials = _review_trial_rates(review)
    benchmark = None
    if review:
        row_index, _ = _find_row(review, ["根据行业确定", "收益率"])
        benchmark = _first_numeric_on_row(review, row_index)
    vendor_indicators: dict[str, Any] = {
        "total_investment": total,
        "interest_during_construction": interest or 0.0,
        "working_capital": working or 0.0,
        "equity_capital": capital,
        "loan": loan,
        "debt_funding_total": debt_funding_total,
        "long_term_loan": long_term_loan,
        "short_term_working_capital_borrowing": short_term_loan,
        "short_term_working_capital_borrowing_schedule": short_term_loan_schedule,
        "loan_selection": loan_selection,
        "benchmark_rate_pct": benchmark * 100.0 if benchmark is not None and abs(benchmark) <= 1 else benchmark,
        "trial_rates": trials,
    }
    summary_values = indicator_summary.get("normalized") or {}

    def summary_number(key: str) -> Optional[float]:
        return _to_float((summary_values.get(key) or {}).get("value"))

    # A summary sheet is a reference/reconciliation source, never model truth.
    # It may fill a missing vendor-reference field but cannot silently overwrite
    # a value already extracted from the underlying financing/cash-flow sheets.
    for key in (
        "total_investment", "construction_investment", "working_capital",
        "equity_capital", "loan", "project_irr_pre_tax_pct",
        "project_irr_pct", "capital_irr_pct", "npv_wan",
        "static_payback_years", "build_period_years", "revenue",
    ):
        value = summary_number(key)
        if value is not None and vendor_indicators.get(key) is None:
            vendor_indicators[key] = value
    if total is not None:
        vendor_indicators["construction_investment"] = round(
            total - float(interest or 0.0) - float(working or 0.0), 6
        )
    if trials.get("project_pre_tax"):
        vendor_indicators["project_irr_pre_tax_pct"] = trials["project_pre_tax"]["rate_pct"]
    if trials.get("project_after_tax"):
        vendor_indicators["project_irr_pct"] = trials["project_after_tax"]["rate_pct"]
    if trials.get("capital_after_tax"):
        vendor_indicators["capital_irr_pct"] = trials["capital_after_tax"]["rate_pct"]

    if project_after and benchmark is not None:
        try:
            from lvke_mcp.servers.finance_calc.calculations import npv
            from lvke_mcp.domains.finance.reference_track import payback_from_period_rows

            values = [float(item["value"]) for item in project_after]
            vendor_indicators["npv_wan"] = npv([0.0, *values], float(benchmark))
            payback = payback_from_period_rows(project_after, rate=float(benchmark))
            vendor_indicators["static_payback_years"] = payback["static_years"]
            vendor_indicators["dynamic_payback_years"] = payback["dynamic_years"]
        except Exception:  # noqa: BLE001
            pass

    comparable_keys = (
        "total_investment", "construction_investment", "working_capital",
        "interest_during_construction", "equity_capital", "loan",
        "project_irr_pct", "capital_irr_pct", "npv_wan", "static_payback_years",
    )
    reference_pack["vendor_indicators"] = vendor_indicators
    reference_pack["indicators"] = {
        key: vendor_indicators[key]
        for key in comparable_keys
        if vendor_indicators.get(key) is not None
    }
    reconciliations = []
    for key in (
        "total_investment", "construction_investment", "working_capital",
        "equity_capital", "loan", "project_irr_pct", "capital_irr_pct",
        "npv_wan", "static_payback_years",
    ):
        summary_item = summary_values.get(key) or {}
        summary_value = _to_float(summary_item.get("value"))
        extracted_value = _to_float(vendor_indicators.get(key))
        if summary_value is None or extracted_value is None:
            continue
        tolerance = 0.01 if "irr" in key else max(abs(summary_value) * 1e-6, 0.01)
        delta = abs(summary_value - extracted_value)
        reconciliations.append({
            "field": key,
            "summary_value": summary_value,
            "summary_locator": summary_item.get("locator") or "",
            "underlying_value": extracted_value,
            "delta": delta,
            "tolerance": tolerance,
            "within_tolerance": delta <= tolerance,
        })
    reference_pack["summary_reconciliation"] = reconciliations
    mismatched_summary = [row for row in reconciliations if not row["within_tolerance"]]
    if mismatched_summary:
        reference_pack.setdefault("warnings", []).append(
            f"主要指标汇总表存在 {len(mismatched_summary)} 项反向勾稽差异"
        )

    periods: dict[str, float] = {}
    for index, item in enumerate(project_after):
        periods[f"project_cashflow.net_cashflow.Y{index}"] = float(item["value"])
    for index, item in enumerate(capital_after):
        periods[f"capital_cashflow.net_cashflow.Y{index}"] = float(item["value"])
    for index, item in enumerate(icr, 1):
        periods[f"debt_service.icr.Y{index}"] = float(item["value"])
    for index, item in enumerate(dscr, 1):
        periods[f"debt_service.dscr.Y{index}"] = float(item["value"])
    reference_pack["periods"] = periods


def build_reference_pack(xlsx_path: str | Path) -> dict[str, Any]:
    """Read one vendor xlsx into an immutable, JSON-serializable reference pack."""
    path = Path(xlsx_path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(str(path))
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise VendorImportError(f"不支持的甲方工作簿类型：{path.suffix}")

    try:
        from lvke_mcp.servers.excel_bridge.formulas import (
            FormulaBackend,
            FormulaBackendUnavailable,
        )
        from lvke_mcp.servers.lvke_templates.catalog import map_vendor_sheet
    except Exception as exc:  # noqa: BLE001
        raise VendorImportError(f"缺少 Excel/附表桥接模块：{exc}") from exc

    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    formula_status = "available"
    formula_warning = ""
    formula_sheets: dict[str, dict[str, Any]] = {}
    cross_sheet = {"cross_sheet_refs": {}, "matrix": {}, "total": 0}
    sheet_names: list[str] = []
    backend = None
    try:
        backend = FormulaBackend(str(path))
        sheet_names = backend.sheet_names()
        for name in sheet_names:
            raw = backend.read_formulas(name, max_rows=MAX_ROWS, max_cols=MAX_COLS)
            formula_sheets[name] = {
                item["cell"]: {
                    "formula": item.get("formula") or "",
                    "cached_value": item.get("cached_value"),
                    "references": list(item.get("references") or []),
                }
                for item in raw.get("cells") or []
            }
        cross_sheet = backend.cross_sheet_refs()
    except FormulaBackendUnavailable as exc:
        formula_status = "unavailable"
        formula_warning = str(exc)
        try:
            from lvke_mcp.servers.excel_bridge.reader import pick_backend

            sheet_names = pick_backend().list_sheets(path)
        except Exception as read_exc:  # noqa: BLE001
            raise VendorImportError(str(read_exc)) from read_exc
    except Exception as exc:  # noqa: BLE001
        raise VendorImportError(f"公式解析失败：{type(exc).__name__}: {exc}") from exc
    finally:
        if backend is not None:
            backend.close()

    value_sheets, value_backend = _read_value_sheets(path, sheet_names)
    sheet_map: dict[str, Any] = {}
    sheets: dict[str, Any] = {}
    for name in sheet_names:
        sheet = dict(value_sheets.get(name) or {"values": {}, "max_row": 0, "max_col": 0})
        sheet["formulas"] = formula_sheets.get(name) or {}
        sheet["formula_count"] = len(sheet["formulas"])
        business = _infer_business(name, sheet)
        by_business = map_vendor_sheet(business=business) if business else None
        by_name = map_vendor_sheet(sheet_name=name.strip())
        conflicts = []
        if by_business and by_name and by_business.get("business") != by_name.get("business"):
            conflicts.append(
                "业务语义候选与裸工作表名登记冲突："
                f"{by_business.get('business')} != {by_name.get('business')}"
            )
        mapped = by_business or by_name
        mapping = dict(mapped or {})
        resolved_business = business or mapping.get("business") or ""
        formula_cross_refs = sum(
            1
            for item in (sheet.get("formulas") or {}).values()
            for reference in (item.get("references") or [])
            if "!" in str(reference)
        )
        key_rows = []
        for row in _rows(sheet).values():
            label = _row_label(row)
            normalized = _norm(label)
            if any(
                token in normalized
                for token in (
                    "总投资", "营业收入", "净现金流量", "内部收益率",
                    "还本付息", "不确定性因数", "指标名称",
                )
            ):
                key_rows.append(label)
            if len(key_rows) >= 8:
                break
        rule_hits = []
        if business:
            rule_hits.append("sheet_name+header_semantics")
        if by_business:
            rule_hits.append("canonical_business_dictionary")
        if by_name:
            rule_hits.append("registered_vendor_sheet_name")
        if key_rows:
            rule_hits.append("key_business_rows")
        if formula_cross_refs:
            rule_hits.append("cross_sheet_formula_references")
        non_empty = _sheet_is_nonempty(sheet)
        confidence = 0.0
        if mapped and resolved_business:
            confidence = 0.97 if by_business and by_name and not conflicts else 0.90 if by_business else 0.72
            if key_rows:
                confidence = min(0.99, confidence + 0.02)
            if formula_cross_refs:
                confidence = min(0.99, confidence + 0.01)
            if conflicts:
                confidence = min(confidence, 0.49)
        mapping.update({
            "vendor_sheet_actual": name,
            "business": resolved_business,
            "candidate_business": resolved_business,
            "mapped": bool(mapped and not conflicts),
            "mapping_rule": "vendor_business_semantics.v2" if mapped else "unmapped",
            "confidence": confidence,
            "rule_hits": rule_hits,
            "key_row_signals": key_rows,
            "cross_sheet_formula_reference_count": formula_cross_refs,
            "conflict_reasons": conflicts,
            "non_empty": non_empty,
            # Mechanical inference is only a candidate.  Every non-empty sheet
            # must receive an append-only human mapped/ignored decision before
            # the bound reference track can be approved.
            "decision_status": "pending" if non_empty else "not_required",
            "review_required": non_empty,
        })
        sheet["business"] = mapping["business"]
        sheet["mapping"] = mapping
        sheets[name] = sheet
        sheet_map[name] = mapping

    dependency_graph: dict[str, list[str]] = {}
    for name, sheet in sheets.items():
        for cell, item in (sheet.get("formulas") or {}).items():
            qualified = []
            for reference in item.get("references") or []:
                target = str(reference).replace("$", "")
                qualified.append(target if "!" in target else f"{name}!{target}")
            dependency_graph[f"{name}!{cell}"] = qualified

    reference_pack: dict[str, Any] = {
        "version": REFERENCE_PACK_VERSION,
        "source_type": "vendor_reference",
        "reliability_grade": "C",
        "read_only": True,
        "source": {
            "path": str(path),
            "workbook_name": path.name,
            "workbook_sha256": digest,
        },
        "formula_status": formula_status,
        "formula_warning": formula_warning,
        "value_backend": value_backend,
        "sheet_map": sheet_map,
        "sheets": sheets,
        "formulas": {name: sheet.get("formulas") or {} for name, sheet in sheets.items()},
        "dependency_graph": dependency_graph,
        "cross_sheet_refs": cross_sheet,
        "warnings": [formula_warning] if formula_warning else [],
    }
    unmapped = [name for name, mapping in sheet_map.items() if not mapping.get("mapped")]
    if unmapped:
        reference_pack["warnings"].append(f"未映射工作表：{', '.join(unmapped)}")
    pending_decisions = [
        name for name, mapping in sheet_map.items()
        if mapping.get("non_empty") and mapping.get("decision_status") == "pending"
    ]
    if pending_decisions:
        reference_pack["warnings"].append(
            f"{len(pending_decisions)} 张非空工作表待人工映射/忽略裁决"
        )
    _extract_reference_data(reference_pack)
    return reference_pack


_CONST_ARITH = re.compile(
    r"^\s*[+-]?\d+(?:\.\d+)?\s*[+\-*/]\s*[+-]?\d+(?:\.\d+)?\s*$"
)


def _constant_formula_frequency(packs: Iterable[dict[str, Any]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for pack in packs:
        signatures = set()
        for formulas in (pack.get("formulas") or {}).values():
            for item in (formulas or {}).values():
                formula = re.sub(r"\s+", "", str(item.get("formula") or "").lstrip("="))
                if _CONST_ARITH.fullmatch(formula):
                    signatures.add(formula)
        counter.update(signatures)
    return counter


def _orphan_constant_formulas(reference_pack: dict[str, Any]) -> list[dict[str, Any]]:
    dependents: dict[str, set[str]] = defaultdict(set)
    for sheet_name, formulas in (reference_pack.get("formulas") or {}).items():
        for cell, item in (formulas or {}).items():
            dependent_id = f"{sheet_name}!{cell}"
            for reference in item.get("references") or []:
                target = str(reference).replace("$", "").split(":", 1)[0]
                if "!" not in target:
                    target = f"{sheet_name}!{target}"
                dependents[target].add(dependent_id)
    result = []
    for sheet_name, formulas in (reference_pack.get("formulas") or {}).items():
        for cell, item in (formulas or {}).items():
            formula = re.sub(r"\s+", "", str(item.get("formula") or "").lstrip("="))
            if not _CONST_ARITH.fullmatch(formula):
                continue
            locator = f"{sheet_name}!{cell}"
            direct = dependents.get(locator) or set()
            # A copied scratch block often has one terminal dependent (for
            # example J16 constant subtraction -> K16 ratio).  Treat the whole
            # disconnected two-cell island as orphan, but retain parameters
            # such as 3/100 that feed a real formula chain or multiple sheets.
            if len(direct) > 1 or any(dependents.get(item_id) for item_id in direct):
                continue
            result.append({"locator": locator, "formula": formula, "cached_value": item.get("cached_value")})
    return result


def _label_locators(sheet: Optional[dict[str, Any]], keyword: str) -> list[str]:
    if not sheet:
        return []
    result = []
    for row in _rows(sheet).values():
        for col, (cell, value) in row.items():
            if col <= 4 and isinstance(value, str) and _norm(keyword) in _norm(value):
                result.append(cell)
    return result


def detect_cleanup_issues(
    reference_pack: dict[str, Any],
    *,
    cohort_reference_packs: Optional[list[dict[str, Any]]] = None,
    irr_residual_tolerance_wan: float = IRR_RESIDUAL_TOL_WAN,
) -> list[dict[str, Any]]:
    """Detect F1/F2/F3 without modifying any vendor value."""
    findings: list[dict[str, Any]] = []
    project_cf = _find_mapped_sheet(reference_pack, "项目投资现金流量表")
    investment_labels = []
    financing_labels = []
    for keyword in ("项目资本金", "固定资产投资", "建设投资"):
        investment_labels.extend(_label_locators(project_cf, keyword))
    for keyword in ("借款本金偿还", "借款利息支付"):
        financing_labels.extend(_label_locators(project_cf, keyword))
    if investment_labels and financing_labels:
        sheet_name = str((project_cf or {}).get("mapping", {}).get("vendor_sheet_actual") or "")
        findings.append({
            "code": "F1",
            "type": "project_cashflow_financing_duplication",
            "severity": "high",
            "blocking": False,
            "locator": ", ".join(f"{sheet_name}!{cell}" for cell in financing_labels),
            "vendor_value": "项目投资现金流同时列示投资支出与还本付息",
            "engine_suggestion": "全投资口径移除借款本金偿还和利息支付；融资项仅进入资本金现金流",
            "detail": "附表9混入融资现金流，可能造成投资本金重复计列并压低项目IRR",
        })

    trials = (reference_pack.get("vendor_indicators") or {}).get("trial_rates") or {}
    cashflows = reference_pack.get("cashflows") or {}
    try:
        from lvke_mcp.servers.finance_calc.calculations import irr, npv
    except Exception:  # noqa: BLE001
        irr = npv = None  # type: ignore
    for key, label in (
        ("project_pre_tax", "项目税前IRR"),
        ("project_after_tax", "项目税后IRR"),
        ("capital_after_tax", "资本金税后IRR"),
    ):
        trial = trials.get(key) or {}
        series = cashflows.get(key) or []
        if not trial or not series or not trial.get("hardcoded"):
            continue
        values = [float(item["value"]) for item in series]
        rate = float(trial.get("rate") or 0.0)
        residual = None
        solved_pct = None
        solve_error = ""
        if npv is not None:
            try:
                residual = float(npv([0.0, *values], rate))
            except Exception as exc:  # noqa: BLE001
                solve_error = f"NPV复算失败：{type(exc).__name__}: {exc}"
        if irr is not None:
            try:
                solved_pct = float(irr([0.0, *values], guess=rate) * 100.0)
            except Exception as exc:  # noqa: BLE001
                solve_error = f"IRR求解失败：{type(exc).__name__}: {exc}"
        residual_exceeded = residual is None or abs(residual) > float(irr_residual_tolerance_wan)
        negative_irr = (
            float(trial.get("rate_pct") or 0.0) < 0
            or (solved_pct is not None and solved_pct < 0)
        )
        findings.append({
            "code": "F2",
            "type": "hardcoded_irr_trial",
            "severity": "high" if residual_exceeded else "medium",
            "blocking": negative_irr,
            "locator": f"{_find_mapped_sheet(reference_pack, '投资估算复核表').get('mapping', {}).get('vendor_sheet_actual', '投资复核')}!{trial.get('cell')}",
            "vendor_value": float(trial.get("rate_pct") or 0.0),
            "npv_residual_wan": residual,
            "tolerance_wan": float(irr_residual_tolerance_wan),
            "engine_suggestion": solved_pct,
            "negative_irr": negative_irr,
            "detail": (
                f"{label}为硬编码试算值，代回NPV残差="
                f"{residual if residual is not None else '不可得'}万元；应使用确定性IRR求解"
                + (f"；{solve_error}" if solve_error else "")
            ),
        })

    cohort = [reference_pack, *(cohort_reference_packs or [])]
    frequency = _constant_formula_frequency(cohort)
    for item in _orphan_constant_formulas(reference_pack):
        count = int(frequency.get(item["formula"], 1))
        findings.append({
            "code": "F3",
            "type": "orphan_constant_formula",
            "severity": "high" if count >= 2 else "medium",
            "blocking": False,
            "locator": item["locator"],
            "vendor_value": f"={item['formula']}",
            "cached_value": item.get("cached_value"),
            "cohort_occurrences": count,
            "engine_suggestion": "删除无下游引用的常量试算式；如属有效参数，应改为有来源的独立输入",
            "detail": (
                "发现纯常量运算且无下游引用的孤儿公式"
                + (f"，在{count}套参考表重复出现" if count >= 2 else "")
            ),
        })
    return findings


def _extract_loan_rate(reference_pack: dict[str, Any]) -> Optional[float]:
    relevant = {
        "建设期贷款利息表", "还款付息测算表",
    }
    pattern = re.compile(r"年利率\s*([0-9]+(?:\.[0-9]+)?)\s*%")
    for sheet in (reference_pack.get("sheets") or {}).values():
        if sheet.get("business") not in relevant:
            continue
        for value in (sheet.get("values") or {}).values():
            if not isinstance(value, str):
                continue
            match = pattern.search(_norm(value))
            if match:
                return float(match.group(1)) / 100.0
        for item in (sheet.get("formulas") or {}).values():
            formula = str(item.get("formula") or "")
            for match in re.finditer(r"\*\s*(0\.\d+)", formula):
                rate = float(match.group(1))
                if 0 < rate < 0.2:
                    return rate
    return None


def _build_period_years(reference_pack: dict[str, Any]) -> int:
    project = (reference_pack.get("cashflows") or {}).get("project_after_tax") or []
    if project:
        return max(int(item.get("period") or 0) for item in project)
    return max(len(project), 1)


def _build_years(reference_pack: dict[str, Any]) -> int:
    revenue = (reference_pack.get("cashflows") or {}).get("project_revenue") or []
    positive = [int(item["period"]) for item in revenue if float(item.get("value") or 0.0) > 0]
    if positive:
        return max(min(positive) - 1, 1)
    cash = (reference_pack.get("cashflows") or {}).get("project_after_tax") or []
    first_positive = next(
        (index for index, item in enumerate(cash) if float(item.get("value") or 0.0) > 0),
        1,
    )
    return max(first_positive, 1)


def _extract_depreciation_years(reference_pack: dict[str, Any]) -> Optional[int]:
    """Read useful life from the depreciation appendix instead of deriving it.

    ``calculation period - construction period`` is not an accounting useful
    life and was the source of a material five-workbook divergence.
    """

    sheet = _find_mapped_sheet(reference_pack, "固定资产折旧费估算表")
    if not sheet:
        return None
    for row_index, row in sorted(_rows(sheet).items()):
        label = _norm(_row_label(row))
        if not any(token in label for token in ("折旧年限", "使用年限", "折旧期")):
            continue
        candidates = []
        for _col, (_cell, value) in sorted(row.items()):
            number = _to_float(value)
            if number is not None and 1 <= number <= 80:
                candidates.append(int(round(number)))
        if candidates:
            return candidates[-1]
    return None


def _extract_depreciation_classes(reference_pack: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract class-specific original value, useful life and implied salvage.

    Real feasibility workbooks often place ``折旧`` and ``年限`` on two header
    rows above a life column, then repeat ``类别 / 原值 / 当期折旧费 / 净值``
    blocks.  Keeping those classes avoids replacing 30/10/5-year assets with a
    synthetic project-period life.
    """

    sheet = _find_mapped_sheet(reference_pack, "固定资产折旧费估算表")
    if not sheet:
        return []
    rows = _rows(sheet)
    column_headers: dict[int, str] = defaultdict(str)
    for row_index, row in rows.items():
        if row_index > 8:
            continue
        for col, (_cell, raw) in row.items():
            if isinstance(raw, str):
                column_headers[col] += _norm(raw)
    life_columns = [
        col for col, text in column_headers.items()
        if "年限" in text and ("折旧" in text or "使用" in text)
    ]
    if not life_columns:
        return []

    descriptors = ("原值", "当期折旧费", "折旧费", "净值", "合计")
    total_columns = [
        col for row_index, row in rows.items() if row_index <= 8
        for col, (_cell, raw) in row.items()
        if isinstance(raw, str) and "合计" in _norm(raw)
    ]
    result: list[dict[str, Any]] = []
    for row_index, row in sorted(rows.items()):
        label = _norm(_row_label(row))
        if "原值" not in label or "合计" in label:
            continue
        name = ""
        for previous in range(row_index - 1, max(row_index - 4, 0), -1):
            prior = rows.get(previous) or {}
            candidate = next(
                (
                    str(raw).strip() for col, (_cell, raw) in sorted(prior.items())
                    if col <= 2 and isinstance(raw, str) and str(raw).strip()
                ),
                "",
            )
            if candidate and not any(token in _norm(candidate) for token in descriptors):
                name = candidate
                break
        if not name:
            continue
        original_cell = ""
        original: Optional[float] = None
        for col in total_columns:
            item = row.get(col)
            value = _to_float(item[1]) if item else None
            if value is not None:
                original = value
                original_cell = item[0]
                break
        if original is None:
            candidates = [
                (cell, value) for col, (cell, raw) in sorted(row.items())
                if col > 2 and col not in life_columns
                and (value := _to_float(raw)) is not None
            ]
            if candidates:
                original_cell, original = max(candidates, key=lambda item: abs(item[1]))
        if original is None or original <= 0:
            continue
        life_cell = ""
        life: Optional[int] = None
        for col in life_columns:
            item = row.get(col)
            value = _to_float(item[1]) if item else None
            if value is not None and 1 <= value <= 100:
                life = int(round(value))
                life_cell = item[0]
                break
        if life is None:
            continue
        annual_row_index, annual_label = _find_row(
            sheet, ["当期折旧费"], exclude=(),
        )
        # _find_row returns the first block; select this class's following row.
        if annual_row_index <= row_index:
            annual_row_index = next(
                (
                    candidate for candidate in range(row_index + 1, min(row_index + 4, max(rows) + 1))
                    if "当期折旧费" in _norm(_row_label(rows.get(candidate) or {}))
                ),
                0,
            )
        if annual_row_index and annual_row_index > row_index + 3:
            annual_row_index = 0
        annual_series = _series_for_row(sheet, annual_row_index) if annual_row_index else []
        annual_item = next(
            (item for item in annual_series if float(item.get("value") or 0.0) > 0),
            None,
        )
        annual = float(annual_item["value"]) if annual_item else None
        implied_salvage = 0.05
        if annual is not None and original > 0:
            candidate_salvage = 1.0 - annual * life / float(original)
            if -0.001 <= candidate_salvage <= 0.5:
                implied_salvage = round(max(candidate_salvage, 0.0), 6)
        result.append({
            "name": name,
            "original_value_wan": round(float(original), 6),
            "depreciation_years": life,
            "salvage_rate": implied_salvage,
            "vendor_annual_depreciation_wan": round(annual, 6) if annual is not None else None,
            "source_locators": {
                "original_value": original_cell,
                "depreciation_years": life_cell,
                "annual_depreciation": str((annual_item or {}).get("cell") or ""),
            },
        })
    return result


def _cost_items(reference_pack: dict[str, Any]) -> dict[str, float]:
    sheet = _find_mapped_sheet(reference_pack, "总成本费用估算表")
    if not sheet:
        values = [
            float(item["value"])
            for item in (reference_pack.get("cashflows") or {}).get("project_operating_cost") or []
            if float(item.get("value") or 0.0) > 0
        ]
        return {"经营成本": max(values)} if values else {}
    aliases = (
        "外购原材料", "燃料", "动力", "工资", "福利", "修理费",
        "营业费用", "销售费用", "其他费用", "经营成本",
    )
    result: dict[str, float] = {}
    for row_index, row in sorted(_rows(sheet).items()):
        label = _row_label(row)
        normalized = _norm(label)
        if not any(alias in normalized for alias in aliases):
            continue
        # 摊销/折旧非现金经营成本；不得并入 cost_items 达产现金口径（B-2 房地产簿误吸）
        if any(token in normalized for token in ("摊销", "折旧", "摊提")):
            continue
        if "经营成本" in normalized and result:
            continue
        series = _series_for_row(sheet, row_index)
        positives = [float(item["value"]) for item in series if float(item["value"]) > 0]
        if not positives:
            continue
        name = next(
            (str(value).strip() for col, (_, value) in sorted(row.items()) if col <= 4 and isinstance(value, str)),
            label,
        )
        result[name] = max(positives)
    # 经营成本合计行与明细（如销售费用）同值时去重，避免 period_opex 双倍（B-2 房地产簿）
    if result:
        detail_sum = sum(
            float(v) for k, v in result.items() if "经营成本" not in _norm(k)
        )
        for key in list(result):
            if "经营成本" not in _norm(key):
                continue
            total = float(result[key] or 0.0)
            if detail_sum > 0 and abs(total - detail_sum) <= max(0.02, abs(total) * 1e-6):
                result.pop(key, None)
            elif detail_sum > 0 and total > 0:
                # 有明细时优先明细，丢弃合计行
                result.pop(key, None)
    if not result:
        values = [
            float(item["value"])
            for item in (reference_pack.get("cashflows") or {}).get("project_operating_cost") or []
            if float(item.get("value") or 0.0) > 0
        ]
        if values:
            result["经营成本"] = max(values)
    return result


def infer_vendor_project_context(reference_pack: dict[str, Any]) -> dict[str, Any]:
    name = str((reference_pack.get("source") or {}).get("workbook_name") or "")
    normalized = _norm(name)
    revenue_sheet = _find_mapped_sheet(
        reference_pack, "营业收入、税金及附加和增值税估算表"
    )
    revenue_labels = [
        _norm(_row_label(row)) for row in _rows(revenue_sheet).values()
    ] if revenue_sheet else []
    has_product_detail = (
        sum(1 for label in revenue_labels if "单价" in label) >= 1
        and sum(1 for label in revenue_labels if "数量" in label) >= 1
    )
    if "房地产" in normalized or any(
        sheet.get("business") == "营业收入、税金及附加和增值税估算表"
        and "不动产" in _norm(" ".join(map(str, (sheet.get("values") or {}).values())))
        for sheet in (reference_pack.get("sheets") or {}).values()
    ):
        industry = "房地产"
        revenue_model = "property_sales"
    elif "产品出售" in normalized:
        industry = "制造业"
        revenue_model = "product_sales"
    elif "厂房出租" in normalized:
        industry = "产业园区租赁"
        revenue_model = "lease_portfolio"
    elif "墓地" in normalized:
        industry = "殡葬服务"
        revenue_model = "inventory_sales"
    elif any(keyword in normalized for keyword in ("石灰岩", "骨料", "机制砂")) or has_product_detail:
        industry = "建材深加工"
        revenue_model = "product_sales"
    else:
        industry = ""
        revenue_model = "flat"
    return {
        "invest_type": "enterprise",
        "industry": industry,
        "build_period_months": _build_years(reference_pack) * 12,
        "revenue_model": revenue_model,
        "source": "vendor_reference_C",
    }



def _series_values(series: list[dict[str, Any]] | None) -> list[float]:
    """Flatten period series to a 0-based year vector (period 1 -> index 0).

    Sparse period maps (e.g. loan only in year 2) are preserved as zeros in
    earlier years — critical for non-uniform construction/loan phasing (B-2).
    """
    if not series:
        return []
    by_period: dict[int, float] = {}
    order: list[float] = []
    for item in series:
        if not isinstance(item, dict):
            continue
        value = _to_float(item.get("value"))
        if value is None:
            continue
        period = item.get("period")
        try:
            p = int(period) if period is not None else None
        except (TypeError, ValueError):
            p = None
        if p is not None and p >= 1:
            by_period[p] = round(float(value), 6)
        else:
            order.append(round(float(value), 6))
    if by_period:
        max_p = max(by_period)
        return [float(by_period.get(i, 0.0)) for i in range(1, max_p + 1)]
    return order


def _funding_year_schedules(reference_pack: dict[str, Any]) -> dict[str, list[float]]:
    """Read construction/loan/equity year plans from 投资使用计划与资金筹措表.

    These are *input* parameters (how much is spent/borrowed each year), not
    engine outputs. Used to replace uniform build-period averaging (B-2).
    """
    funding = _find_mapped_sheet(reference_pack, "投资使用计划与资金筹措表")
    if not funding:
        return {}
    construction = _series_values(
        _series_by_label(funding, ["建设投资"], exclude=["总投资", "使用计划"])
    )
    idc = _series_values(
        _series_by_label(funding, ["建设期", "利息"], exclude=["合计表"])
    )
    if not idc:
        idc = _series_values(_series_by_label(funding, ["建设期贷款利息"]))
    loan_draw = _series_values(
        _series_by_label(funding, ["申请银行长期借款"])
    )
    if not loan_draw:
        loan_draw = _series_values(
            _series_by_label(funding, ["银行长期借款"], exclude=["短期"])
        )
    equity = _series_values(
        _series_by_label(funding, ["项目资本金"], exclude=["用于"])
    )
    wc = _series_values(
        _series_by_label(funding, ["新增铺底流动资金"])
    )
    if not wc:
        wc = _series_values(_series_by_label(funding, ["新增流动资金"]))
    if not wc:
        wc = _series_values(_series_by_label(funding, ["流动资金"], exclude=["用于", "借款"]))

    # Build-year cash outlay: choose construction-only vs construction+IDC by
    # matching the vendor project CF build-year negatives when available.
    outlay: list[float] = []
    cand_c: list[float] = list(construction)
    cand_ci: list[float] = []
    if construction or idc:
        n = max(len(construction), len(idc))
        for i in range(n):
            c = construction[i] if i < len(construction) else 0.0
            d = idc[i] if i < len(idc) else 0.0
            cand_ci.append(round(float(c) + float(d), 6))
        while cand_ci and cand_ci[-1] == 0.0:
            cand_ci.pop()
        while cand_c and cand_c[-1] == 0.0:
            cand_c.pop()

    project_after = _series_values(
        (reference_pack.get("cashflows") or {}).get("project_after_tax")
    )
    # Build years are leading non-positive project CF years.
    build_ref: list[float] = []
    for value in project_after:
        if value <= 0:
            build_ref.append(round(abs(float(value)), 6))
        else:
            break

    def _l1(a: list[float], b: list[float]) -> float:
        n = max(len(a), len(b))
        s = 0.0
        for i in range(n):
            av = a[i] if i < len(a) else 0.0
            bv = b[i] if i < len(b) else 0.0
            s += abs(av - bv)
        return s

    if build_ref and (cand_c or cand_ci):
        err_c = _l1(cand_c, build_ref) if cand_c else float("inf")
        err_ci = _l1(cand_ci, build_ref) if cand_ci else float("inf")
        outlay = cand_ci if err_ci <= err_c else cand_c
    elif cand_ci:
        outlay = cand_ci
    elif cand_c:
        outlay = cand_c
    else:
        total_plan = _series_values(
            _series_by_label(funding, ["总投资"], exclude=["使用计划", "资金筹措表"])
        )
        if total_plan:
            outlay = list(total_plan)
            if wc and outlay and abs(outlay[-1] - wc[-1]) <= max(0.02, abs(outlay[-1]) * 1e-6):
                outlay = outlay[:-1]

    result: dict[str, list[float]] = {}
    if outlay and any(v > 0 for v in outlay):
        result["construction_outlay_by_year"] = outlay
    if construction and any(v > 0 for v in construction):
        result["construction_invest_by_year"] = construction
    if idc and any(v > 0 for v in idc):
        result["idc_by_year"] = idc
    if loan_draw and any(v > 0 for v in loan_draw):
        result["loan_draw_by_year"] = loan_draw
    if equity and any(v > 0 for v in equity):
        result["equity_inject_by_year"] = equity
    if wc and any(v > 0 for v in wc):
        result["working_capital_by_year"] = wc
    return result



def _construction_items_from_investment_sheet(reference_pack: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract qty×indicator engineering lines from 固定资产投资估算表 (input side).

    Only rows with unit + positive quantity + positive indicator are kept so
    formal 附表1 can show reference-grade quantity/indicator without inventing
    reverse-engineered quantities (B-4).
    """
    sheet = _find_mapped_sheet(reference_pack, "固定资产投资估算表")
    if not sheet:
        return []
    other_section_row, _ = _find_row(sheet, ["工程建设其他费用"])
    items: list[dict[str, Any]] = []
    for row_index, row in sorted(_rows(sheet).items()):
        if other_section_row and row_index >= other_section_row:
            break
        label = _row_label(row)
        if not label:
            continue
        normalized = _norm(label)
        # skip pure section headers without qty
        unit = None
        quantity = None
        indicator = None
        # Total columns vary across vendor templates.  Resolve the header-backed
        # total instead of assuming I/J; direct-amount rows may legitimately
        # omit a unit indicator (for example, an installation-material lump sum).
        amount = _total_for_row(sheet, row_index)
        for col, cell in sorted(row.items()):
            if col == 3 and isinstance(cell[1], str) and str(cell[1]).strip():
                unit = str(cell[1]).strip().replace("\n", "")
            if col == 4:
                quantity = _to_float(cell[1])
            if col == 5:
                indicator = _to_float(cell[1])
        sequence = row.get(1, ("", None))[1]
        sequence_text = str(sequence or "").strip()
        if (
            (indicator is None or indicator <= 0)
            and sequence_text
            and "." not in sequence_text
        ):
            # Integer-level rows are section subtotals (for example
            # “1 建筑工程”); their children carry the actual input lines.
            continue
        if quantity is None or quantity <= 0:
            continue
        if not unit:
            continue
        if (indicator is None or indicator <= 0) and (amount is None or amount <= 0):
            continue
        # category heuristic
        category = "other"
        if any(k in normalized for k in ("车间", "楼", "门卫", "道路", "绿化", "广场", "建筑", "土建")):
            category = "civil"
        elif any(k in normalized for k in ("设备", "电梯", "吊车", "光伏", "家具", "充电")):
            category = "equipment"
        elif any(k in normalized for k in ("安装", "给排水", "供配电", "弱电", "消防", "通风", "安防")):
            category = "installation"
        name = next(
            (
                str(cell[1]).strip()
                for col, cell in sorted(row.items())
                if col == 2 and isinstance(cell[1], str) and str(cell[1]).strip()
            ),
            label,
        )
        items.append({
            "name": name,
            "unit": unit,
            "quantity": round(float(quantity), 6),
            **(
                {"indicator_yuan": round(float(indicator), 6)}
                if indicator is not None and indicator > 0
                else {}
            ),
            "amount_wan": (
                round(float(amount), 6)
                if amount is not None and amount > 0
                else round(float(quantity) * float(indicator) / 10000.0, 6)
            ),
            "category": category,
            "source_row": row_index,
        })
    return items


def _construction_detail_from_items(items: list[dict[str, Any]]) -> dict[str, float]:
    """Map qty×indicator items into engine construction_detail three buckets.

    Engine / quality_audit expect civil+equipment+installation (=建设投资工程费用三段).
    Uncategorized amounts fold into civil rather than inventing a fourth key.
    """
    civil = equipment = installation = 0.0
    for item in items:
        amount = float(item.get("amount_wan") or 0.0)
        cat = str(item.get("category") or "other")
        if cat == "equipment":
            equipment += amount
        elif cat == "installation":
            installation += amount
        else:
            civil += amount
    return {
        "civil_wan": round(civil, 6),
        "equipment_wan": round(equipment, 6),
        "installation_wan": round(installation, 6),
    }



def _wc_turnover_from_sheet(reference_pack: dict[str, Any]) -> dict[str, Any] | None:
    """Extract min turnover days from 流动资金估算表 (input parameters).

    保留存货子树（原材料/燃料/在产品/产成品），供 reference 结构门禁使用；
    同时提供兼容字段 inventory（优先原材料天数）给引擎既有周转法。
    """
    sheet = _find_mapped_sheet(reference_pack, "流动资金估算表")
    if not sheet:
        return None
    mapping = {
        "应收账款": "receivable",
        "存货": "inventory",
        "原材料": "raw",
        "燃料": "fuel",
        "动力": "fuel",
        "在产品": "wip",
        "产成品": "finished",
        "现金": "cash",
        "应付账款": "payable",
    }
    out: dict[str, Any] = {}
    inv_detail: dict[str, float] = {}
    for row_index, row in sorted(_rows(sheet).items()):
        label = _row_label(row)
        normalized = _norm(label)
        for zh, key in mapping.items():
            if zh not in normalized:
                continue
            days = None
            for col, cell in sorted(row.items()):
                if col == 3:
                    days = _to_float(cell[1])
                    break
            if days is None or days <= 0:
                continue
            days_f = round(float(days), 6)
            series = _series_for_row(sheet, row_index)
            mature = max(
                (float(item.get("value") or 0.0) for item in series),
                default=0.0,
            )
            component: Any = days_f
            if mature > 0:
                mature_item = max(
                    series,
                    key=lambda item: float(item.get("value") or 0.0),
                )
                component = {
                    "days": days_f,
                    "annual_base_wan": round(mature * 360.0 / days_f, 6),
                    "base_source": (
                        f"{sheet.get('name') or '流动资金估算表'}!"
                        f"{mature_item.get('cell') or ''}"
                    ).rstrip("!"),
                }
            if key in {"raw", "fuel", "wip", "finished"}:
                inv_detail.setdefault(key, component)
            elif key not in out:
                out[key] = component
    if inv_detail:
        out["inventory_detail"] = inv_detail
        # 兼容引擎：inventory 天数优先用原材料，否则用已有 inventory
        if "inventory" not in out:
            first = inv_detail.get("raw") or inv_detail.get("finished") or next(iter(inv_detail.values()))
            out["inventory"] = (
                first.get("days") if isinstance(first, dict) else first
            )
    return out or None


def _full_working_capital_profile(reference_pack: dict[str, Any]) -> dict[str, Any]:
    """Return full turnover working capital, distinct from initial/铺底 funding."""

    sheet = _find_mapped_sheet(reference_pack, "流动资金估算表")
    if not sheet:
        return {}
    total_series: list[dict[str, Any]] = []
    increase_series: list[dict[str, Any]] = []
    excluded = tuple(_norm(value) for value in ("铺底", "当期", "来源", "借款", "自筹"))
    for row_index, row in sorted(_rows(sheet).items()):
        normalized = _norm(_row_label(row))
        series = _series_for_row(sheet, row_index)
        if not series:
            continue
        if "流动资金" in normalized and "当期增加额" in normalized:
            increase_series = series
        elif not total_series and "流动资金" in normalized and not any(
            token in normalized for token in excluded
        ):
            total_series = series
    positive = [
        item for item in total_series if float(item.get("value") or 0.0) > 0
    ]
    if not positive:
        return {}
    peak_item = max(positive, key=lambda item: float(item.get("value") or 0.0))
    return {
        "total_wan": round(float(peak_item.get("value") or 0.0), 6),
        "total_locator": (
            f"{sheet.get('name') or '流动资金估算表'}!"
            f"{peak_item.get('cell') or ''}"
        ).rstrip("!"),
        "increase_by_year": _series_values(increase_series),
    }


def _investment_segment_totals(reference_pack: dict[str, Any]) -> dict[str, float]:
    sheet = _find_mapped_sheet(reference_pack, "固定资产投资估算表")
    if not sheet:
        return {}
    values = {
        "engineering_wan": _total_by_label(
            sheet,
            ["工程费用"],
            exclude=["工程建设其他费用"],
        ),
        "other_wan": _total_by_label(sheet, ["工程建设其他费用"]),
        "reserve_wan": _total_by_label(sheet, ["预备费"]),
    }
    return {
        key: round(float(value), 6)
        for key, value in values.items()
        if value is not None and value >= 0
    }


def _staff_detail_from_wage_sheet(reference_pack: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract labor headcount × average wage from 工资及福利费估算表."""
    sheet = (
        _find_mapped_sheet(reference_pack, "工资及福利费估算表")
        or _find_mapped_sheet(reference_pack, "工资及附加估算表")
    )
    if not sheet:
        return []
    rows = _rows(sheet)
    headcount = None
    avg_wage = None
    for _idx, row in sorted(rows.items()):
        label = _norm(_row_label(row))
        if "人数" in label or "劳动定员" in label:
            for col, cell in sorted(row.items()):
                if col >= 4:
                    val = _to_float(cell[1])
                    if val is not None and val > 0:
                        headcount = val
                        break
        if "人均" in label and "工资" in label:
            for col, cell in sorted(row.items()):
                if col >= 4:
                    val = _to_float(cell[1])
                    if val is not None and val > 0:
                        avg_wage = val
                        break
    if headcount and avg_wage:
        return [{
            "category": "劳动定员",
            "headcount": int(round(headcount)),
            "avg_wage_yuan": round(float(avg_wage), 2),
        }]
    return []


def build_finance_input_from_vendor(reference_pack: dict[str, Any]) -> dict[str, Any]:
    """Build deterministic engine inputs from vendor *input parameters*, not outputs."""
    indicators = reference_pack.get("vendor_indicators") or {}
    vendor_total = _to_float(indicators.get("total_investment"))
    if vendor_total is None or vendor_total <= 0:
        return {
            "is_operating": True,
            "input_sources": {},
            "_missing_inputs": ["total_investment_wan"],
        }
    interest = max(_to_float(indicators.get("interest_during_construction")) or 0.0, 0.0)
    initial_working = max(_to_float(indicators.get("working_capital")) or 0.0, 0.0)
    construction = round(vendor_total - interest - initial_working, 6)
    full_wc_profile = _full_working_capital_profile(reference_pack)
    working = max(
        _to_float(full_wc_profile.get("total_wan")) or initial_working,
        0.0,
    )
    total = round(construction + interest + working, 6)
    capital = _to_float(indicators.get("equity_capital"))
    loan = _to_float(indicators.get("loan"))
    short_working_loan = max(
        (
            float(item.get("value") or 0.0)
            for item in indicators.get("short_term_working_capital_borrowing_schedule") or []
            if isinstance(item, dict)
        ),
        default=0.0,
    )
    if working > initial_working and short_working_loan > 0:
        loan = round(float(loan or 0.0) + short_working_loan, 6)
    revenue_series = (reference_pack.get("cashflows") or {}).get("project_revenue") or []
    revenue_values = [float(item["value"]) for item in revenue_series if float(item.get("value") or 0.0) > 0]
    annual_revenue = max(revenue_values) if revenue_values else None
    principal = (reference_pack.get("cashflows") or {}).get("debt_principal") or []
    loan_years = sum(1 for item in principal if float(item.get("value") or 0.0) > 0)
    # Prefer full 还本 series including zeros only if positive exists
    _principal_vals = [
        round(float(item.get("value") or 0.0), 6)
        for item in principal
        if isinstance(item, dict)
    ]
    context = infer_vendor_project_context(reference_pack)
    source_ref = str((reference_pack.get("source") or {}).get("path") or "")
    source_fields = (
        "total_investment_wan", "invest_breakdown", "capital_own_wan", "loan_wan",
        "loan_rate", "loan_years", "calc_period_years", "annual_revenue_wan",
        "cost_items", "operating_cost_by_year",
        "construction_outlay_by_year", "loan_draw_by_year",
    )
    input_sources = {
        field: {
            "note": "从甲方计算表输入参数抽取；仅C级参考，输出由我方引擎重算",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
        for field in source_fields
    }
    # 逐年经营成本（附表成本/现金流经营成本行）：属输入侧序列，供引擎按年使用，
    # 禁止把达产峰值按营收比例回推（B-2 amount-bridge 主因之一）。
    opex_series_raw = (reference_pack.get("cashflows") or {}).get("project_operating_cost") or []
    operating_cost_by_year: list[float] = []
    for item in opex_series_raw:
        if not isinstance(item, dict):
            continue
        value = _to_float(item.get("value"))
        if value is None:
            continue
        operating_cost_by_year.append(round(float(value), 6))
    result: dict[str, Any] = {
        "total_investment_wan": round(total, 6),
        "invest_breakdown": {
            "construction_wan": construction,
            "interest_wan": round(interest, 6),
            "working_capital_wan": round(working, 6),
        },
        "is_operating": bool(revenue_values),
        "annual_revenue_wan": round(annual_revenue, 6) if annual_revenue is not None else None,
        "loan_rate": _extract_loan_rate(reference_pack),
        "loan_years": max(loan_years, 1),
        "calc_period_years": max(_build_period_years(reference_pack), 1),
        "build_period_months": context["build_period_months"],
        "cost_items": _cost_items(reference_pack),
        "input_sources": input_sources,
    }
    if operating_cost_by_year and any(v > 0 for v in operating_cost_by_year):
        result["operating_cost_by_year"] = operating_cost_by_year
        result["input_sources"]["operating_cost_by_year"] = {
            "note": "从甲方表经营成本逐年序列抽取（输入参数）；引擎按年使用，不抄净现金流输出",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    # S2 T-debt: 还本/付息序列作为输入（来自还款付息表行，非输出指标抄写）
    if _principal_vals and any(v > 0 for v in _principal_vals):
        if working > initial_working and short_working_loan > 0:
            _principal_vals[-1] = round(
                _principal_vals[-1] + short_working_loan,
                6,
            )
        result["loan_principal_by_year"] = _principal_vals
        result["loan_repay_method"] = "principal_schedule"
        result["input_sources"]["loan_principal_by_year"] = {
            "note": "从还款付息表「还本」行抽取；引擎按序列还本并重算余额/利息备付",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    debt_sheet = _find_mapped_sheet(reference_pack, "还款付息测算表")
    if debt_sheet is None:
        for _n, _sh in (reference_pack.get("sheets") or {}).items():
            if "还本付息" in str(_n) or "还款付息" in str(_n):
                debt_sheet = _sh
                break
    if debt_sheet is not None:
        _interest_pay = _series_by_label(debt_sheet, ["付息"], exclude=["还本付息"])
        if not _interest_pay:
            _interest_pay = _series_by_label(debt_sheet, ["其中:付息", "支付利息"])
        _int_vals = [
            round(float(item.get("value") or 0.0), 6)
            for item in _interest_pay
            if isinstance(item, dict)
        ]
        if (
            _int_vals
            and any(v > 0 for v in _int_vals)
            and not (working > initial_working and short_working_loan > 0)
        ):
            result["loan_interest_by_year"] = _int_vals
            result["input_sources"]["loan_interest_by_year"] = {
                "note": "从还款付息表「付息」行抽取（输入）；缺省时引擎按余额×利率重算",
                "source_ref": source_ref,
                "evidence_level": "C",
            }
        # improve principal extract with 其中:还本 if debt_principal empty
        if not result.get("loan_principal_by_year"):
            _prin2 = _series_by_label(debt_sheet, ["其中:还本", "还本"], exclude=["还本付息", "当期还本付息"])
            _pv = [round(float(x.get("value") or 0.0), 6) for x in _prin2 if isinstance(x, dict)]
            if _pv and any(v > 0 for v in _pv):
                result["loan_principal_by_year"] = _pv
                result["loan_repay_method"] = "principal_schedule"
                result["loan_years"] = max(sum(1 for v in _pv if v > 0), 1)
    # B-2: construction / loan draw schedules from 资金筹措表 (input phasing)
    for key, series in _funding_year_schedules(reference_pack).items():
        result[key] = series
        result["input_sources"][key] = {
            "note": "从投资使用计划与资金筹措表抽取的分年投入/提款（输入参数，非输出指标）",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    if full_wc_profile.get("increase_by_year"):
        result["working_capital_by_year"] = list(
            full_wc_profile["increase_by_year"]
        )
        result["input_sources"]["working_capital_by_year"] = {
            "note": "从流动资金估算表的全额流动资金当期增加额抽取；不使用铺底资金替代全额流资",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    # B-2 terminal recovery inputs from 项目投资现金流量表 (row present ⇒ authoritative)
    project_cf = _find_mapped_sheet(reference_pack, "项目投资现金流量表")
    if project_cf:
        fa_rec = _series_values(
            _series_by_label(project_cf, ["回收固定资产余值"])
        )
        wc_rec = _series_values(
            _series_by_label(
                project_cf, ["回收", "流动资金"], exclude=["固定资产"]
            )
        )
        # Row exists with no positives → explicit 0 (common vendor simplification)
        fa_row, _ = _find_row(project_cf, ["回收固定资产余值"])
        wc_row, _ = _find_row(project_cf, ["回收", "流动资金"], exclude=["固定资产"])
        if fa_row:
            peak_fa = max(fa_rec) if fa_rec else 0.0
            result["terminal_fixed_asset_recover_wan"] = round(float(peak_fa or 0.0), 6)
            result["input_sources"]["terminal_fixed_asset_recover_wan"] = {
                "note": "从项目投资现金流量表「回收固定资产余值」抽取；无正值则按 0 回收",
                "source_ref": source_ref,
                "evidence_level": "C",
            }
        if wc_row:
            peak_wc = max(wc_rec) if wc_rec else 0.0
            result["terminal_working_capital_recover_wan"] = round(
                float(working if working > initial_working else peak_wc or 0.0),
                6,
            )
            result["input_sources"]["terminal_working_capital_recover_wan"] = {
                "note": "从项目投资现金流量表「回收流动资金」抽取；无正值则按 0 回收",
                "source_ref": source_ref,
                "evidence_level": "C",
            }
    depreciation_classes = _extract_depreciation_classes(reference_pack)
    depreciation_years = _extract_depreciation_years(reference_pack)
    if depreciation_classes:
        result["depreciation_classes"] = depreciation_classes
        result["input_sources"]["depreciation_classes"] = {
            "note": "从固定资产折旧费估算表逐类抽取原值、折旧年限、残值率和来源单元格",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    elif depreciation_years is not None:
        result["depreciation_years"] = depreciation_years
        result["input_sources"]["depreciation_years"] = {
            "note": "从固定资产折旧费估算表的折旧/使用年限字段提取",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    else:
        result.setdefault("_missing_inputs", []).append("depreciation_years")
    if capital is not None:
        result["capital_own_wan"] = round(capital, 6)
    if loan is not None:
        result["loan_wan"] = round(loan, 6)
    if result.get("loan_rate") is None:
        result.pop("loan_rate", None)
    if result.get("annual_revenue_wan") is None:
        result.pop("annual_revenue_wan", None)
    wc_turn = _wc_turnover_from_sheet(reference_pack)
    if wc_turn:
        result["wc_turnover"] = wc_turn
        result["input_sources"]["wc_turnover"] = {
            "note": "从流动资金估算表最低周转天数抽取（含存货分项，输入参数）",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    construction_items = _construction_items_from_investment_sheet(reference_pack)
    investment_segments = _investment_segment_totals(reference_pack)
    if construction_items:
        bd = dict(result.get("invest_breakdown") or {})
        if investment_segments.get("other_wan") is not None:
            bd["other_wan"] = investment_segments["other_wan"]
        if investment_segments.get("reserve_wan") is not None:
            bd["reserve_wan"] = investment_segments["reserve_wan"]
        bd["construction_items"] = construction_items
        detail = _construction_detail_from_items(construction_items)
        detail_sum = sum(float(detail.get(k) or 0.0) for k in (
            "civil_wan", "equipment_wan", "installation_wan",
        ))
        construction_total = float((bd.get("construction_wan") or construction or 0.0))
        engineering_target = float(
            investment_segments.get("engineering_wan")
            or max(
                construction_total
                - float(bd.get("other_wan") or 0.0)
                - float(bd.get("reserve_wan") or 0.0),
                0.0,
            )
        )
        # Only reconcile against the evidenced engineering subtotal.  Other
        # costs and contingency remain separate segments and are never folded
        # into civil works.
        if engineering_target > 0 and detail_sum > 0:
            residual = round(engineering_target - detail_sum, 6)
            if residual > 0:
                detail["civil_wan"] = round(float(detail.get("civil_wan") or 0.0) + residual, 6)
                detail_sum = engineering_target
            if detail_sum + 1.0 >= engineering_target * 0.95:
                if detail_sum > engineering_target + 1.0:
                    scale = engineering_target / detail_sum
                    for k in ("civil_wan", "equipment_wan", "installation_wan"):
                        detail[k] = round(float(detail.get(k) or 0.0) * scale, 6)
                bd["construction_detail"] = detail
                result["construction_detail"] = detail
        result["invest_breakdown"] = bd
        result["input_sources"]["construction_items"] = {
            "note": "从固定资产投资估算表抽取工程量×估算指标明细（输入参数）",
            "source_ref": source_ref,
            "evidence_level": "C",
            "detail_sum_wan": round(sum(float(detail.get(k) or 0.0) for k in ("civil_wan","equipment_wan","installation_wan")), 6),
            "construction_wan": construction_total,
            "engineering_target_wan": engineering_target,
            "detail_attached": "construction_detail" in bd,
        }
    staff_detail = _staff_detail_from_wage_sheet(reference_pack)
    if staff_detail:
        result["staff_detail"] = staff_detail
        result["input_sources"]["staff_detail"] = {
            "note": "从工资及福利费估算表抽取劳动定员×人均年工资",
            "source_ref": source_ref,
            "evidence_level": "C",
        }
    wage = next(
        (value for key, value in (result.get("cost_items") or {}).items() if "工资" in key),
        None,
    )
    if wage is not None:
        result["wage_wan"] = wage
    return result


def _product_lines_from_revenue_sheet(reference_pack: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract named price/quantity/ramp lines from a real product revenue table."""

    sheet = _find_mapped_sheet(
        reference_pack, "营业收入、税金及附加和增值税估算表"
    )
    if not sheet:
        return []
    rows = _rows(sheet)
    products: list[dict[str, Any]] = []
    for row_index, row in sorted(rows.items()):
        price_row = rows.get(row_index + 1) or {}
        quantity_row = rows.get(row_index + 2) or {}
        if "单价" not in _norm(_row_label(price_row)) or "数量" not in _norm(_row_label(quantity_row)):
            continue
        name = next(
            (str(raw).strip() for col, (_cell, raw) in sorted(row.items()) if col == 2 and str(raw).strip()),
            "",
        )
        if not name:
            continue
        price_values = [
            float(value) for col, (_cell, raw) in sorted(price_row.items())
            if col >= 5 and (value := _to_float(raw)) is not None and value > 0
        ]
        quantities = [
            float(value) for col, (_cell, raw) in sorted(quantity_row.items())
            if col >= 5 and (value := _to_float(raw)) is not None and value >= 0
        ]
        if not price_values or not quantities or max(quantities) <= 0:
            continue
        capacity = max(quantities)
        unit = next(
            (str(raw).strip() for col, (_cell, raw) in sorted(quantity_row.items()) if col == 3 and str(raw).strip()),
            "项",
        )
        products.append({
            "name": name,
            "unit": unit,
            "price_per_unit": round(max(price_values), 6),
            "price_unit": "yuan",
            "capacity": round(capacity, 6),
            "ramp": [round(value / capacity, 8) for value in quantities],
            "var_cost_rate": 0.0,
        })
    return products


def build_vendor_finance_spec(
    reference_pack: dict[str, Any],
    finance_input: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Select the required revenue model while keeping vendor evidence at grade C."""
    inputs = finance_input or build_finance_input_from_vendor(reference_pack)
    context = infer_vendor_project_context(reference_pack)
    revenue_series = [
        float(item["value"])
        for item in (reference_pack.get("cashflows") or {}).get("project_revenue") or []
        if float(item.get("value") or 0.0) > 0
    ]
    peak = max(revenue_series) if revenue_series else float(inputs.get("annual_revenue_wan") or 0.0)
    model = context["revenue_model"]
    if model == "product_sales" and peak > 0:
        product_lines = _product_lines_from_revenue_sheet(reference_pack)
        ramp = [round(value / peak, 8) for value in revenue_series]
        revenue = {
            "model": "product_sales",
            "annual_revenue_wan": peak,
            "products": product_lines or [{
                "name": "甲方表综合产品组合",
                "unit": "项",
                "price_per_unit": round(peak * 10000.0, 6),
                "price_unit": "yuan",
                "capacity": 1.0,
                "ramp": ramp,
                "var_cost_rate": 0.0,
            }],
        }
    elif model == "property_sales" and revenue_series:
        total_revenue = sum(revenue_series)
        revenue_sheet = _find_mapped_sheet(
            reference_pack, "营业收入、税金及附加和增值税估算表"
        )
        area = 0.0
        if revenue_sheet:
            for row_index, row in _rows(revenue_sheet).items():
                if "销售面积" not in _norm(_row_label(row)):
                    continue
                value = _total_for_row(revenue_sheet, row_index)
                if value and value > 0:
                    area += value
        if area <= 0:
            area = 1.0
        revenue = {
            "model": "property_sales",
            "annual_revenue_wan": peak,
            "saleable_area": round(area, 6),
            "price_per_sqm": round(total_revenue * 10000.0 / area, 6),
            "absorption": [round(value / total_revenue, 8) for value in revenue_series],
        }
    elif model in {"lease_portfolio", "inventory_sales"} and revenue_series:
        revenue = {
            "model": model,
            "annual_revenue_wan": peak,
            "annual_schedule_wan": [round(value, 6) for value in revenue_series],
            "inventory_total": round(sum(revenue_series), 6) if model == "inventory_sales" else 0.0,
            "sales_schedule": (
                [round(value / sum(revenue_series), 8) for value in revenue_series]
                if model == "inventory_sales" and sum(revenue_series) > 0 else []
            ),
        }
    else:
        revenue = {"model": "flat", "annual_revenue_wan": peak}
    return {
        "version": "finance_spec.v2",
        "confirmation_status": "candidate",
        "industry": context["industry"],
        "invest_type": context["invest_type"],
        "source_hint": "vendor_reference_C",
        "revenue": revenue,
        "cost": {},
        "tax": {},
        "assumptions": [
            "甲方工作簿仅作C级输入与公式参考；所有指标由本系统确定性引擎重算",
        ],
    }
