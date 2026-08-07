"""工作簿读取与业务类型推断：值表读取、表头文本与非空判定。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from .base import (
    MAX_COLS,
    MAX_ROWS,
    VendorImportError,
    _cell_parts,
    _col_letter,
    _jsonable,
    _norm,
)


def _read_value_sheets(path: Path, sheet_names: list[str]) -> tuple[dict[str, Any], str]:
    """Read cached values once with openpyxl, falling back to excel_bridge reader."""
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            result: dict[str, Any] = {}
            for name in workbook.sheetnames:
                worksheet = workbook[name]
                values: dict[str, Any] = {}
                row_limit = min(int(worksheet.max_row or 0), MAX_ROWS)
                col_limit = min(int(worksheet.max_column or 0), MAX_COLS)
                for row in worksheet.iter_rows(
                    min_row=1, max_row=row_limit, min_col=1, max_col=col_limit
                ):
                    for cell in row:
                        value = _jsonable(cell.value)
                        if value is not None:
                            values[cell.coordinate] = value
                result[name] = {
                    "values": values,
                    "max_row": int(worksheet.max_row or 0),
                    "max_col": int(worksheet.max_column or 0),
                    "value_backend": "openpyxl",
                }
            return result, "openpyxl"
        finally:
            workbook.close()
    except Exception as openpyxl_error:  # noqa: BLE001
        try:
            from lvke_mcp.adapters.spreadsheets.reader import pick_backend

            backend = pick_backend()
            result = {}
            for name in sheet_names:
                read = backend.read_sheet(path, name, MAX_ROWS, MAX_COLS)
                values = {}
                for row_index, row in enumerate(read.rows, 1):
                    for col_index, value in enumerate(row, 1):
                        if value is not None:
                            values[f"{_col_letter(col_index)}{row_index}"] = _jsonable(value)
                result[name] = {
                    "values": values,
                    "max_row": read.row_count,
                    "max_col": read.col_count,
                    "value_backend": read.backend,
                }
            return result, str(getattr(backend, "name", "excel_bridge-reader"))
        except Exception as fallback_error:  # noqa: BLE001
            raise VendorImportError(
                "无法读取甲方工作簿缓存值："
                f"openpyxl={type(openpyxl_error).__name__}; "
                f"fallback={type(fallback_error).__name__}: {fallback_error}"
            ) from fallback_error


def _header_text(sheet: dict[str, Any], max_rows: int = 4) -> str:
    parts = []
    for cell, value in (sheet.get("values") or {}).items():
        row, _ = _cell_parts(cell)
        if row <= max_rows and isinstance(value, str):
            parts.append(value)
    return " ".join(parts)


def _sheet_is_nonempty(sheet: dict[str, Any]) -> bool:
    """Return whether a vendor sheet contains any business-visible content.

    Formula-only sheets are non-empty even when the reader cannot obtain cached
    values.  This deliberately errs on the side of requiring an explicit review
    decision instead of silently dropping a vendor worksheet.
    """

    for value in (sheet.get("values") or {}).values():
        if value is not None and (not isinstance(value, str) or value.strip()):
            return True
    return bool(sheet.get("formulas"))


def _infer_business(sheet_name: str, sheet: dict[str, Any]) -> str:
    """Infer canonical business name; business meaning wins over bare appendix no.

    表名 + 表头(前 4 行)联合判定。**强业务特征**(投资复核 / 利润分配 / 还本付息 /
    现金流)先判定：利润表、现金流表的表头会同时出现"营业收入""税金"列标题，若把
    收入税金规则放前面会把它们误判为收入税金表(FIN-MAP-001，房地产利润表实测错配为
    附表5)。因此把这些强特征提到收入税金规则之前，并给收入税金规则加"非利润分配"
    排除项。**注意**："总成本费用"必须保持在工资/折旧/摊销/原材料/燃料之前，否则总成本
    表会被其列标题(工资及福利费/折旧费…)误判为工资/折旧子表。
    """
    name = _norm(sheet_name)
    raw = f"{sheet_name} {_header_text(sheet)}"
    text = _norm(raw)
    # ── 强业务特征优先(表名/表头前 4 行出现即高置信) ──
    if "投资复核" in text or "投资估算复核" in text:
        return "投资估算复核表"
    if name == "项目财务分析" or "净现值计算方式" in text:
        return "投资估算复核表"
    if "利润" in text and "分配" in text:
        return "利润与利润分配表"
    if "还本付息" in text or "还款付息" in text:
        return "还款付息测算表"
    if "项目投资现金流" in text:
        return "项目投资现金流量表"
    if "项目资本金" in text and ("现金流" in text or name == "项目资本金"):
        return "项目资本金流量表"
    if "敏感度分析" in text or "敏感性分析" in text:
        return "单因素敏感性分析表"
    if "主要经济指标汇总" in text or "指标名称" in text:
        return "主要技术经济指标汇总表"
    if "结构方案" in text and "综合判断" in text:
        return "建筑/技术方案比选表"
    # ── 投资/融资侧 ──
    if "固定资产投资估算" in text:
        return "固定资产投资估算表"
    if "建设期" in text and "利息" in text:
        return "建设期贷款利息表"
    if "流动资金估算" in text or "流动资金测算" in text:
        return "流动资金估算表"
    if "资金筹措" in text and ("使用计划" in text or "总投资" in text):
        return "投资使用计划与资金筹措表"
    # ── 收入税金表：放在利润/现金流之后，显式排除利润分配特征 ──
    if (
        "营业收入" in text
        and ("税金" in text or "增值税" in text)
        and not ("利润" in text and "分配" in text)
    ):
        return "营业收入、税金及附加和增值税估算表"
    # ── 成本侧：总成本费用必须先于工资/折旧/摊销/原材料/燃料子表判定 ──
    if "总成本费用" in text:
        return "总成本费用估算表"
    if "工资" in text and ("福利" in text or "附加" in text):
        return "工资及附加估算表"
    if "折旧" in text:
        return "固定资产折旧费估算表"
    if "摊销" in text:
        return "无形资产和其他资产摊销估算表"
    if "原材料" in text:
        return "外购原材料费估算表"
    if "燃料" in text and "动力" in text:
        return "外购燃料和动力费估算表"
    return ""


def _find_mapped_sheet(reference_pack: dict[str, Any], business: str) -> Optional[dict[str, Any]]:
    for sheet in (reference_pack.get("sheets") or {}).values():
        if sheet.get("business") == business:
            return sheet
    return None
