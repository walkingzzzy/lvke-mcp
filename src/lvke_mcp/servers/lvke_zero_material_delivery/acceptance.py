"""Real stdio MCP acceptance for zero-material delivery; does not use pytest."""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook

from lvke_mcp.testing.protocol_testkit import (
    initialize_message,
    initialized_notification,
    run_raw,
    tool_call,
)

MODULE = "lvke_mcp.servers.lvke_zero_material_delivery.server"
PROTOCOL_VERSION = "2025-11-25"
WORKSPACE = "zmd-acceptance"
OTHER_WORKSPACE = "zmd-other-workspace"


def _request(method: str, params: dict[str, Any], *, timeout: float = 180) -> dict[str, Any]:
    responses, stderr = run_raw(
        MODULE,
        [
            initialize_message(1, PROTOCOL_VERSION),
            initialized_notification(),
            {"jsonrpc": "2.0", "id": 2, "method": method, "params": params},
        ],
        timeout=timeout,
    )
    if stderr:
        print(stderr, file=sys.stderr, end="")
    response = responses[-1]
    if "error" in response:
        raise AssertionError(f"{method} protocol error: {response['error']}")
    return dict(response.get("result") or {})


def _call(name: str, arguments: dict[str, Any], *, timeout: float = 180) -> dict[str, Any]:
    responses, stderr = run_raw(
        MODULE,
        [
            initialize_message(1, PROTOCOL_VERSION),
            initialized_notification(),
            tool_call(2, name, arguments),
        ],
        timeout=timeout,
    )
    if stderr:
        print(stderr, file=sys.stderr, end="")
    response = responses[-1]
    if "error" in response:
        raise AssertionError(f"{name} protocol error: {response['error']}")
    result = dict(response.get("result") or {})
    structured = result.get("structuredContent")
    if isinstance(structured, dict):
        return dict(structured)
    content = result.get("content") or []
    if not content or content[0].get("type") != "text":
        raise AssertionError(f"{name} returned no JSON tool content")
    return json.loads(content[0]["text"])


def _read_standard(uri: str) -> tuple[bytes, str]:
    result = _request("resources/read", {"uri": uri})
    contents = result.get("contents") or []
    if len(contents) != 1:
        raise AssertionError(f"resources/read returned {len(contents)} contents for {uri}")
    item = contents[0]
    mime_type = str(item.get("mimeType") or "")
    if "blob" in item:
        return base64.b64decode(item["blob"]), mime_type
    return str(item.get("text") or "").encode("utf-8"), mime_type


def _assert_fail_closed(payload: dict[str, Any]) -> None:
    assert payload.get("validation_complete") is False
    assert payload.get("input_evidence_complete") is False


def _find_uri(uris: list[str], suffix: str) -> str:
    return next(uri for uri in uris if uri.endswith(suffix))


def _run() -> dict[str, Any]:
    listed = _request("tools/list", {})
    tool_names = {item["name"] for item in listed.get("tools") or []}
    assert len(tool_names) == 11
    assert {"delivery_create_from_sentence", "delivery_start", "delivery_confirm_assumptions"} <= tool_names

    ambiguous = _call(
        "delivery_create_from_sentence",
        {
            "workspace_id": "zmd-ambiguous",
            "sentence": "建设制造业文旅综合项目",
            "idempotency_key": "ambiguous-create-v1",
        },
    )
    assert ambiguous.get("status") == "missing_inputs"
    assert ambiguous.get("code") == "ambiguous_route"
    _assert_fail_closed(ambiguous)

    create_args = {
        "workspace_id": WORKSPACE,
        "sentence": "在湖北建设一座儿童游乐园并编制可行性研究报告",
        "project_name": "湖北儿童游乐园零材料基准",
        "region": "湖北省",
        "idempotency_key": "create-v1",
    }
    created = _call("delivery_create_from_sentence", create_args)
    replayed = _call("delivery_create_from_sentence", create_args)
    assert replayed.get("idempotent_replay") is True
    assert replayed["delivery_intent"]["delivery_intent_id"] == created["delivery_intent"]["delivery_intent_id"]
    assert replayed["delivery_run"]["delivery_run_id"] == created["delivery_run"]["delivery_run_id"]
    _assert_fail_closed(created)

    conflict = _call(
        "delivery_create_from_sentence",
        {**create_args, "sentence": "在湖北建设制造业工厂并编制可行性研究报告"},
    )
    assert conflict.get("code") == "idempotency_conflict"

    initial_run_id = created["delivery_run"]["delivery_run_id"]
    started = _call(
        "delivery_start",
        {
            "workspace_id": WORKSPACE,
            "delivery_run_id": initial_run_id,
            "idempotency_key": "start-v1",
        },
        timeout=300,
    )
    preview_run = started["delivery_run"]
    assert preview_run["stage"] == "preview_ready"
    assert preview_run["assurance_level"] == "estimate_preview"
    assert preview_run["domain_results"]["technical_preview_ready"] is True
    _assert_fail_closed(started)

    preview_run_id = preview_run["delivery_run_id"]
    restarted_status = _call(
        "delivery_status",
        {"workspace_id": WORKSPACE, "delivery_run_id": preview_run_id},
    )
    assert restarted_status["delivery_run"]["content_hash"] == preview_run["content_hash"]
    assert restarted_status["stage"] == "preview_ready"

    assumptions = _call(
        "delivery_list_assumptions",
        {
            "workspace_id": WORKSPACE,
            "assumption_package_id": preview_run["assumption_package_id"],
            "limit": 10,
        },
    )
    fields = assumptions["assumptions"]
    assert len(fields) == 6
    assert all(item["source_type"] == "controlled_assumption" for item in fields)
    scores = [item["confirmation_priority_score"] for item in assumptions["confirmation_items"]]
    assert scores == sorted(scores, reverse=True)
    prior_revenue = next(item["value"] for item in fields if item["name"] == "annual_revenue_wan")

    confirmed = _call(
        "delivery_confirm_assumptions",
        {
            "workspace_id": WORKSPACE,
            "assumption_package_id": preview_run["assumption_package_id"],
            "confirmations": [
                {
                    "name": "annual_revenue_wan",
                    "value": round(float(prior_revenue) * 1.1, 2),
                    "source_ref": "acceptance-user-confirmation",
                }
            ],
            "idempotency_key": "confirm-v2",
        },
        timeout=300,
    )
    assert confirmed.get("automatic_recalculation") is True
    confirmed_run = confirmed["delivery_run"]
    confirmed_package = confirmed["assumption_package"]
    assert confirmed_package["revision"] == 2
    assert confirmed_package["previous_assumption_package_id"] == preview_run["assumption_package_id"]
    assert confirmed_package["assumption_package_id"] != preview_run["assumption_package_id"]
    assert confirmed_run["stage"] == "preview_ready"
    assert confirmed_run["object_refs"]["finance_run_id"] != preview_run["object_refs"]["finance_run_id"]
    assert confirmed_run["previous_run_id"] == confirmed["confirmation_run"]["delivery_run_id"]
    _assert_fail_closed(confirmed)

    old_package = _call(
        "delivery_get",
        {"workspace_id": WORKSPACE, "object_id": preview_run["assumption_package_id"]},
    )["object"]
    old_revenue = next(item["value"] for item in old_package["fields"] if item["name"] == "annual_revenue_wan")
    assert old_package["revision"] == 1
    assert old_revenue == prior_revenue

    artifacts = _call(
        "delivery_get_artifacts",
        {"workspace_id": WORKSPACE, "delivery_run_id": confirmed_run["delivery_run_id"]},
    )
    uris = list(artifacts["artifacts"])
    manifest_uri = artifacts["manifest_uri"]
    report_md_uri = _find_uri(uris, "/files/report.md")
    report_docx_uri = _find_uri(uris, "/files/report.docx")
    xlsx_uri = _find_uri(uris, "/xlsx")
    csv_uris = [uri for uri in uris if "/csv/" in uri and not uri.endswith("/00-lineage")]
    assert len(csv_uris) == 13

    manifest_bytes, manifest_mime = _read_standard(manifest_uri)
    manifest_record = json.loads(manifest_bytes)
    manifest = manifest_record["payload"]
    assert manifest_mime == "application/json"
    assert manifest["status"] == "estimate_preview"
    assert manifest["object_refs"]["finance_run_id"] == confirmed_run["object_refs"]["finance_run_id"]
    _assert_fail_closed(manifest)

    markdown_bytes, markdown_mime = _read_standard(report_md_uri)
    assert markdown_mime.startswith("text/markdown")
    assert "技术预估版，非正式发布" in markdown_bytes.decode("utf-8")
    docx_bytes, docx_mime = _read_standard(report_docx_uri)
    assert "wordprocessingml.document" in docx_mime
    document = Document(io.BytesIO(docx_bytes))
    assert any("技术预估版" in paragraph.text for paragraph in document.paragraphs)

    xlsx_bytes, xlsx_mime = _read_standard(xlsx_uri)
    assert "spreadsheetml.sheet" in xlsx_mime
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    assert len(workbook.sheetnames) >= 13
    workbook.close()

    csv_bytes, csv_mime = _read_standard(csv_uris[0])
    assert "csv" in csv_mime
    csv_rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    assert len(csv_rows) >= 2

    expected_file_hashes = {item["resource_uri"]: item["content_hash"] for item in manifest["files"]}
    for uri, expected_hash in expected_file_hashes.items():
        content, _mime = _read_standard(uri)
        assert "sha256:" + hashlib.sha256(content).hexdigest() == expected_hash

    resources = _request("resources/list", {}).get("resources") or []
    listed_uris = {str(item["uri"]) for item in resources}
    assert {manifest_uri, report_md_uri, report_docx_uri, xlsx_uri, csv_uris[0]} <= listed_uris

    wrong_workspace = _call(
        "delivery_read_resource",
        {"workspace_id": OTHER_WORKSPACE, "uri": manifest_uri},
    )
    assert wrong_workspace.get("code") == "resource_scope_mismatch"

    cancelled = _call(
        "delivery_cancel",
        {
            "workspace_id": WORKSPACE,
            "delivery_run_id": confirmed_run["delivery_run_id"],
            "reason": "acceptance cancellation",
            "idempotency_key": "cancel-v1",
        },
    )
    cancelled_run_id = cancelled["delivery_run"]["delivery_run_id"]
    blocked_start = _call(
        "delivery_start",
        {
            "workspace_id": WORKSPACE,
            "delivery_run_id": cancelled_run_id,
            "idempotency_key": "start-cancelled-v1",
        },
    )
    assert blocked_start.get("code") == "delivery_run_cancelled"
    resumed = _call(
        "delivery_resume",
        {
            "workspace_id": WORKSPACE,
            "delivery_run_id": cancelled_run_id,
            "reason": "acceptance resume",
            "idempotency_key": "resume-v1",
        },
    )
    assert resumed["delivery_run"]["stage"] == "preview_ready"
    assert resumed["delivery_run"]["previous_run_id"] == cancelled_run_id

    return {
        "ok": True,
        "server": "lvke-zero-material-delivery",
        "tool_count": len(tool_names),
        "workspace_id": WORKSPACE,
        "initial_run_id": initial_run_id,
        "preview_run_id": preview_run_id,
        "confirmed_run_id": confirmed_run["delivery_run_id"],
        "finance_run_changed": True,
        "csv_count": len(csv_uris),
        "xlsx_sheet_count": len(load_workbook(io.BytesIO(xlsx_bytes), read_only=True).sheetnames),
        "standard_resource_count": len(resources),
        "validation_complete": False,
        "input_evidence_complete": False,
    }


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="lvke-zero-material-acceptance-") as home:
        os.environ["LVKE_HOME"] = str(Path(home))
        result = _run()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
