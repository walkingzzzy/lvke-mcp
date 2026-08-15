"""Versioned rule-pack registry and deterministic first-wave checks."""

from __future__ import annotations

import json
import hashlib
import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Iterable

from lvke_mcp.runtime.storage import sha256_json

ENGINE_VERSION = "lvke-deliverable-review-engine.v1.2"
RECALC_ENV_VERSION = "libreoffice-headless.required.v1"
REPO_ROOT = Path(__file__).resolve().parents[2]

_DEFINITIONS: dict[str, dict[str, Any]] = {
    "core-deliverable": {
        "version": "1.0.0", "standards": ["PKG-STD-001"],
        "rules": ["CORE.TARGET.RESOLVED", "CORE.UPSTREAM.COMPLETE", "CORE.STANDARDS.LOCKED"],
    },
    "finance-core": {
        "version": "1.2.0", "standards": [
            "PKG-STD-001", "PKG-STD-011", "PKG-STD-012", "PKG-STD-013", "PKG-STD-014",
            "PKG-STD-015", "PKG-STD-016", "PKG-STD-017", "PKG-STD-018", "PKG-STD-019", "PKG-STD-020",
        ],
        "rules": [
            "FIN.EXISTING.CHECKS", "FIN.INVESTMENT.BALANCE", "FIN.FUNDING.BALANCE",
            "FIN.DEPRECIATION.RECALC", "FIN.TAX.RECALC", "FIN.DEBT.ROLLFORWARD",
            "FIN.DEBT.COVERAGE", "FIN.WORKING_CAPITAL.DRIVER",
            "FIN.PERIOD.RECONCILIATION", "FIN.SENSITIVITY.RERUN",
            "FIN.XLSX.INTEGRITY", "FIN.XLSX.RECALC",
        ],
    },
    "report-core": {
        "version": "1.2.0", "standards": ["PKG-STD-001", "PKG-STD-013"],
        "rules": [
            "REPORT.EXISTING.VALIDATION", "REPORT.SECTIONS.COMPLETE",
            "REPORT.PLACEHOLDER", "REPORT.DUPLICATE.PARAGRAPH",
            "REPORT.CLAIM.EVIDENCE", "REPORT.NUMBERS.BOUND",
            "REPORT.INTERNAL.CONSISTENCY", "REPORT.REFERENCES.FRESH",
        ],
    },
    "combined-core": {
        "version": "1.1.0", "standards": ["PKG-STD-001", "PKG-STD-013"],
        "rules": [
            "COMBINED.BINDINGS.COMPLETE", "COMBINED.UPSTREAM.VERDICTS",
            "COMBINED.NUMBERS.MATCH", "COMBINED.CONCLUSIONS.MATCH",
        ],
    },
    "generic-feasibility": {
        "version": "1.0.0", "standards": ["PKG-STD-001"], "rules": ["PROJECT.METADATA.COMPLETE"],
    },
    "amusement-feasibility": {
        "version": "1.0.0", "standards": ["PKG-STD-001", "PKG-STD-002", "PKG-STD-003"],
        "rules": ["PROJECT.METADATA.COMPLETE"],
    },
    "asset-acquisition": {
        "version": "1.0.0", "standards": ["PKG-STD-006", "PKG-STD-018", "PKG-STD-019"], "rules": ["ACQ.TRANSACTION.INPUTS"],
    },
    "hotel-acquisition": {
        "version": "1.2.0", "standards": ["PKG-STD-021"],
        "rules": [
            "HOTEL.RIGHTS.LICENSES", "HOTEL.OPERATING_MODEL",
            "HOTEL.ROOM_COUNT.CONFLICT", "HOTEL.AREA.CONFLICT",
            "HOTEL.LAND_USE.COMPLIANCE", "HOTEL.LEASE.TERMS.CONFLICT",
        ],
    },
    "solar-acquisition": {
        "version": "1.0.0", "standards": ["PKG-STD-006", "PKG-STD-013", "PKG-STD-014"],
        "rules": ["ACQ.TRANSACTION.INPUTS"],
    },
    "mineral-processing": {
        "version": "1.2.0", "standards": ["PKG-STD-010", "PKG-STD-022"],
        "rules": [
            "MINERAL.PERMITS", "MINERAL.MARKET.RADIUS",
            "MINERAL.OPERATING.DRIVERS", "MINERAL.CONTRACT.FIELDS",
        ],
    },
}

_COMPONENT_RULE_SOURCES: dict[str, tuple[str, ...]] = {
    "finance-core": ("finance-report-core", "accounting-tax-core"),
    "report-core": ("finance-report-core", "accounting-tax-core"),
    "asset-acquisition": ("finance-report-core", "accounting-tax-core"),
    "hotel-acquisition": ("hotel-mining-core",),
    "mineral-processing": ("hotel-mining-core",),
}

_CONTEXTUAL_RULE_PACKS = {
    "amusement-feasibility", "asset-acquisition", "hotel-acquisition",
    "solar-acquisition", "mineral-processing",
}
_ACQUISITION_TARGETS = {"acquisition_run", "acquisition_tables_package"}
_ACQUISITION_TRANSACTIONS = {"asset_acquisition", "equity_acquisition"}


def _context_flags(
    target_type: str,
    component_types: Iterable[str],
    project_context: dict[str, Any],
) -> dict[str, Any]:
    components = {str(item) for item in component_types if str(item)}
    project_type = str(project_context.get("project_type") or "")
    transaction = str(project_context.get("transaction_structure") or "")
    acquisition = (
        target_type in _ACQUISITION_TARGETS
        or bool(components.intersection(_ACQUISITION_TARGETS))
        or project_type == "asset_acquisition"
        or transaction in _ACQUISITION_TRANSACTIONS
    )
    return {
        "components": components,
        "project_type": project_type,
        "transaction": transaction,
        "asset_type": str(project_context.get("asset_type") or "general"),
        "acquisition": acquisition,
        "lease": transaction == "operation_lease",
    }


def _pack_exclusion_reason(pack_id: str, flags: dict[str, Any]) -> str:
    asset_type = flags["asset_type"]
    if pack_id == "amusement-feasibility" and asset_type != "amusement_park":
        return "asset_type_not_amusement_park"
    if pack_id == "asset-acquisition" and not flags["acquisition"] and not flags["lease"]:
        return "transaction_structure_not_acquisition_or_lease"
    if pack_id == "hotel-acquisition" and not (
        asset_type == "hotel_lease" and (flags["acquisition"] or flags["lease"])
    ):
        return "asset_type_or_transaction_not_hotel_acquisition"
    if pack_id == "solar-acquisition" and not (
        asset_type == "solar_power" and flags["acquisition"]
    ):
        return "asset_type_or_transaction_not_solar_acquisition"
    if pack_id == "mineral-processing" and asset_type != "mineral_processing":
        return "asset_type_not_mineral_processing"
    return ""


def _source_rule_exclusion_reason(
    rule: dict[str, Any], flags: dict[str, Any],
) -> str:
    rule_id = str(rule.get("rule_id") or "")
    if rule_id in {"AT-DEED-001", "AT-LVAT-001"} and not flags["acquisition"]:
        return "requires_asset_or_equity_acquisition"
    if rule_id == "AT-LEASE-001" and not flags["lease"]:
        return "requires_operation_lease"
    return ""


def _rule_source_catalog() -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    directory = REPO_ROOT / "config" / "review_rule_sources"
    for path in sorted(directory.glob("*.json")):
        try:
            document = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        source_pack = str(document.get("rule_pack_id") or "")
        for row in document.get("rules") or []:
            if not isinstance(row, dict) or not row.get("rule_id"):
                continue
            record = {
                **row, "source_rule_pack_id": source_pack,
                "source_rule_pack_version": document.get("version"),
                "source_file": path.name,
            }
            catalog[str(row["rule_id"])] = record
    return catalog


def _source_rule_selected(rule: dict[str, Any], components: list[str]) -> bool:
    source_pack = str(rule.get("source_rule_pack_id") or "")
    selected_sources = {
        source for component in components for source in _COMPONENT_RULE_SOURCES.get(component, ())
    }
    if source_pack not in selected_sources:
        return False
    rule_id = str(rule.get("rule_id") or "")
    if "hotel-acquisition" in components and rule_id.startswith("HM-HOTEL-"):
        return True
    if "mineral-processing" in components and (
        rule_id.startswith("HM-MINE-") or rule_id.startswith("HM-FACTORY-")
    ):
        return True
    if source_pack == "hotel-mining-core":
        return False
    return True

_DEFAULTS = {
    "finance_run": ["core-deliverable", "finance-core", "generic-feasibility"],
    "finance_tables_package": ["core-deliverable", "finance-core", "generic-feasibility"],
    "finance_xlsx": ["core-deliverable", "finance-core"],
    "finance_xlsx_source": ["core-deliverable", "finance-core"],
    "acquisition_run": ["core-deliverable", "finance-core", "asset-acquisition"],
    "acquisition_tables_package": ["core-deliverable", "finance-core", "asset-acquisition"],
    "report_revision": ["core-deliverable", "report-core", "generic-feasibility"],
    "report_artifact": ["core-deliverable", "report-core"],
    "combined_deliverable": ["core-deliverable", "finance-core", "report-core", "combined-core"],
}

_RULE_TARGETS: dict[str, set[str]] = {
    "FIN.EXISTING.CHECKS": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.INVESTMENT.BALANCE": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.FUNDING.BALANCE": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.DEPRECIATION.RECALC": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.TAX.RECALC": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.DEBT.ROLLFORWARD": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.DEBT.COVERAGE": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.WORKING_CAPITAL.DRIVER": {"finance_run", "finance_tables_package"},
    "FIN.PERIOD.RECONCILIATION": {"finance_run", "finance_tables_package"},
    "FIN.SENSITIVITY.RERUN": {"finance_run", "finance_tables_package", "acquisition_run", "acquisition_tables_package", "combined_deliverable"},
    "FIN.XLSX.INTEGRITY": {"finance_xlsx", "finance_xlsx_source", "combined_deliverable"},
    "FIN.XLSX.RECALC": {"finance_xlsx", "finance_xlsx_source", "combined_deliverable"},
    "REPORT.EXISTING.VALIDATION": {"report_revision", "combined_deliverable"},
    "REPORT.PLACEHOLDER": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.DUPLICATE.PARAGRAPH": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.SECTIONS.COMPLETE": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.CLAIM.EVIDENCE": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.NUMBERS.BOUND": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.INTERNAL.CONSISTENCY": {"report_revision", "report_artifact", "combined_deliverable"},
    "REPORT.REFERENCES.FRESH": {"report_revision", "report_artifact", "combined_deliverable"},
    "COMBINED.BINDINGS.COMPLETE": {"combined_deliverable"},
    "COMBINED.UPSTREAM.VERDICTS": {"combined_deliverable"},
    "COMBINED.NUMBERS.MATCH": {"combined_deliverable"},
    "COMBINED.CONCLUSIONS.MATCH": {"combined_deliverable"},
    "PROJECT.METADATA.COMPLETE": {"finance_run", "finance_tables_package", "report_revision", "combined_deliverable"},
    "ACQ.TRANSACTION.INPUTS": {"acquisition_run", "acquisition_tables_package", "report_revision", "combined_deliverable"},
    "HOTEL.RIGHTS.LICENSES": {"acquisition_run", "report_revision", "report_artifact", "combined_deliverable"},
    "HOTEL.OPERATING_MODEL": {"acquisition_run", "report_revision", "report_artifact", "combined_deliverable"},
    "HOTEL.ROOM_COUNT.CONFLICT": {"report_revision", "report_artifact", "combined_deliverable"},
    "HOTEL.AREA.CONFLICT": {"report_revision", "report_artifact", "combined_deliverable"},
    "HOTEL.LAND_USE.COMPLIANCE": {"report_revision", "report_artifact", "combined_deliverable"},
    "HOTEL.LEASE.TERMS.CONFLICT": {"report_revision", "report_artifact", "combined_deliverable"},
    "MINERAL.PERMITS": {"report_revision", "report_artifact", "combined_deliverable"},
    "MINERAL.MARKET.RADIUS": {"report_revision", "report_artifact", "combined_deliverable"},
    "MINERAL.OPERATING.DRIVERS": {"report_revision", "report_artifact", "combined_deliverable"},
    "MINERAL.CONTRACT.FIELDS": {"report_revision", "report_artifact", "combined_deliverable"},
}


def registry() -> list[dict[str, Any]]:
    rows = []
    for pack_id, definition in _DEFINITIONS.items():
        source_rules = [
            row for row in _rule_source_catalog().values()
            if _source_rule_selected(row, [pack_id])
        ]
        body = {"rule_pack_id": pack_id, **definition, "source_rules": source_rules}
        rows.append({**body, "content_hash": sha256_json(body)})
    return rows


def compose(
    target_type: str,
    requested: Iterable[str] = (),
    overlays: Iterable[str] = (),
    component_types: Iterable[str] = (),
    project_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    # Mandatory layers cannot be removed by supplying a narrow requested list.
    # Requested packs extend the target defaults; industry packs are the final
    # overlays in the deterministic composition order.
    ids = list(_DEFAULTS.get(target_type, ["core-deliverable"]))
    actual_components = {str(item) for item in component_types if str(item)}
    if target_type == "combined_deliverable":
        if actual_components.intersection({
            "finance_run", "finance_tables_package", "report_revision",
        }):
            ids.append("generic-feasibility")
        if actual_components.intersection({
            "acquisition_run", "acquisition_tables_package",
        }):
            ids.append("asset-acquisition")
    context = dict(project_context or {})
    flags = _context_flags(target_type, actual_components, context)
    asset_type = str(context.get("asset_type") or "general")
    project_type = str(context.get("project_type") or "")
    transaction = str(context.get("transaction_structure") or "")
    automatic: list[str] = []
    if asset_type == "amusement_park":
        automatic.append("amusement-feasibility")
    if asset_type == "hotel_lease":
        automatic.extend(["asset-acquisition", "hotel-acquisition"])
    if asset_type == "mineral_processing":
        automatic.append("mineral-processing")
    if asset_type == "solar_power" and project_type == "asset_acquisition":
        automatic.extend(["asset-acquisition", "solar-acquisition"])
    if project_type == "asset_acquisition" or transaction in {
        "asset_acquisition", "equity_acquisition", "operation_lease",
    }:
        automatic.append("asset-acquisition")
    ids.extend(str(item) for item in requested if str(item))
    ids.extend(automatic)
    ids.extend(str(item) for item in overlays if str(item))
    ordered: list[str] = []
    excluded_rule_packs: list[dict[str, str]] = []
    for item in ids:
        if item not in _DEFINITIONS:
            raise ValueError(f"rule_pack_not_found:{item}")
        exclusion_reason = _pack_exclusion_reason(item, flags)
        if item in _CONTEXTUAL_RULE_PACKS and exclusion_reason:
            exclusion = {"rule_pack_id": item, "reason": exclusion_reason}
            if exclusion not in excluded_rule_packs:
                excluded_rule_packs.append(exclusion)
            continue
        if item not in ordered:
            ordered.append(item)
    definitions = [{"rule_pack_id": item, **_DEFINITIONS[item]} for item in ordered]
    source_target_types = {target_type}
    if target_type == "combined_deliverable":
        source_target_types.update(str(item) for item in component_types if str(item))
    source_rules = [
        row for row in _rule_source_catalog().values()
        if _source_rule_selected(row, ordered)
        and source_target_types.intersection(set(row.get("target_kinds") or []))
    ]
    excluded_rules = [
        {
            "rule_id": str(row.get("rule_id") or ""),
            "reason": _source_rule_exclusion_reason(row, flags),
            "source_rule_pack_id": str(row.get("source_rule_pack_id") or ""),
        }
        for row in source_rules
        if _source_rule_exclusion_reason(row, flags)
    ]
    source_rules = [
        row for row in source_rules
        if not _source_rule_exclusion_reason(row, flags)
    ]
    enabled_source_rules = [row for row in source_rules if row.get("enabled", True) is not False]
    all_rules = [
        *[rule for item in ordered for rule in _DEFINITIONS[item]["rules"]],
        *[str(row.get("rule_id") or "") for row in enabled_source_rules],
    ]
    applicable_rules = [
        rule for rule in all_rules
        if rule not in _RULE_TARGETS
        or target_type in _RULE_TARGETS[rule]
        or (target_type == "combined_deliverable" and bool(
            set(component_types).intersection(_RULE_TARGETS[rule])
        ))
    ]
    body = {
        "rule_pack_id": "+".join(ordered),
        "version": ".".join(_DEFINITIONS[item]["version"] for item in ordered),
        "components": definitions,
        "rule_sources": enabled_source_rules,
        "disabled_rules": [row for row in source_rules if row.get("enabled", True) is False],
        "applicable_rules": applicable_rules,
        "standard_package_ids": sorted({sid for item in ordered for sid in _DEFINITIONS[item]["standards"]}),
        "project_context": context,
        "selected_rule_packs": ordered,
        "excluded_rule_packs": excluded_rule_packs,
        "excluded_rules": excluded_rules,
    }
    return {**body, "content_hash": sha256_json(body)}


def standards_snapshot(
    repo_root: Path,
    package_ids: Iterable[str],
    *,
    review_purpose: str = "project_delivery",
) -> dict[str, Any]:
    path = repo_root / "config" / "review_standards.lock.json"
    if not path.is_file():
        return _standards_snapshot_from_materials(
            package_ids,
            review_purpose=review_purpose,
        )
    try:
        lock = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"available": False, "content_hash": None, "packages": [], "incomplete": list(package_ids)}
    wanted = set(package_ids)
    packages = [
        {
            "package_id": row.get("package_id"), "title": row.get("title"),
            "gate_status": row.get("gate_status"), "source_manifest_sha256": row.get("source_manifest_sha256"),
            "review_findings_required": list(row.get("review_findings_required") or []),
            "scope_limitations": list(row.get("scope_limitations") or []),
            "artifacts": [
                {
                    "artifact_id": artifact.get("artifact_id"),
                    "publisher": artifact.get("publisher"),
                    "document_number": artifact.get("document_number"),
                    "publication_date": artifact.get("publication_date"),
                    "source_url": artifact.get("source_url"),
                    "official_page_url": artifact.get("official_page_url"),
                    "sha256": artifact.get("sha256"),
                    "page_count": artifact.get("page_count"),
                }
                for artifact in (row.get("artifacts") or [])
            ],
        }
        for row in (lock.get("packages") or []) if row.get("package_id") in wanted
    ]
    found = {str(row.get("package_id")) for row in packages}
    incomplete = sorted((wanted - found) | {str(row["package_id"]) for row in packages if row.get("gate_status") != "passed"})
    snapshot = {
        "available": True, "schema_version": lock.get("schema_version"),
        "packages": packages, "incomplete": incomplete,
    }
    return _apply_standard_review_purpose(snapshot, review_purpose)


def _standards_snapshot_from_materials(
    package_ids: Iterable[str],
    *,
    review_purpose: str = "project_delivery",
) -> dict[str, Any]:
    """Build the immutable review basis from the checked-in standard packages."""

    import os

    wanted = sorted(set(str(item) for item in package_ids if str(item)))
    configured = str(os.environ.get("LVKE_GOLDEN_DATA_ROOT") or "").strip()
    docs_root = (
        Path(configured)
        if configured
        else Path(__file__).resolve().parents[4] / "docs"
    )
    package_root = docs_root / "研报资料库" / "交付型资料源" / "06_标准方法包"
    packages: list[dict[str, Any]] = []
    incomplete: list[str] = []
    for package_id in wanted:
        root = package_root / package_id
        manifest_path = root / "source_manifest.json"
        conclusion_path = root / "review" / "review_conclusion.json"
        try:
            manifest_bytes = manifest_path.read_bytes()
            manifest = json.loads(manifest_bytes.decode("utf-8"))
            conclusion = json.loads(conclusion_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            incomplete.append(package_id)
            continue
        manifest_hash = hashlib.sha256(manifest_bytes).hexdigest()
        gate_passed = (
            str(manifest.get("package_id") or "") == package_id
            and str(conclusion.get("package_id") or "") == package_id
            and str(manifest.get("gate_status") or "") == "passed"
            and str(conclusion.get("gate_status") or "") == "passed"
            and str(conclusion.get("source_manifest_sha256") or "") == manifest_hash
        )
        if not gate_passed:
            incomplete.append(package_id)
        packages.append({
            "package_id": package_id,
            "title": manifest.get("title"),
            "gate_status": "passed" if gate_passed else "incomplete",
            "source_manifest_sha256": manifest_hash,
            "review_findings_required": list(conclusion.get("review_findings_required") or []),
            "scope_limitations": [str(manifest.get("reuse_boundary") or "")],
            "artifacts": [
                {
                    "artifact_id": artifact.get("artifact_id"),
                    "publisher": artifact.get("publisher"),
                    "document_number": artifact.get("document_number"),
                    "publication_date": artifact.get("publication_date"),
                    "source_url": artifact.get("source_url"),
                    "official_page_url": artifact.get("official_page_url"),
                    "sha256": artifact.get("sha256"),
                    "page_count": artifact.get("page_count"),
                    "status": artifact.get("status"),
                    "source_is_official": artifact.get("source_is_official"),
                }
                for artifact in manifest.get("artifacts") or []
                if isinstance(artifact, dict)
            ],
        })
    snapshot = {
        "available": bool(packages),
        "schema_version": "review_standards.material_snapshot.v1",
        "packages": packages,
        "incomplete": sorted(set(incomplete) | (set(wanted) - {str(row.get("package_id")) for row in packages})),
    }
    return _apply_standard_review_purpose(snapshot, review_purpose)


_FAILED_ARTIFACT_STATUSES = frozenset({
    "download_or_extract_failed",
    "download_failed",
    "extract_failed",
    "missing",
    "unavailable",
})


def _is_official_framework_artifact(artifact: Any) -> bool:
    """Return whether an artifact really stands as official framework basis."""

    if not isinstance(artifact, dict):
        return False
    if not artifact.get("sha256"):
        return False
    if artifact.get("source_is_official") is not True:
        return False
    status = str(artifact.get("status") or "").strip()
    return status not in _FAILED_ARTIFACT_STATUSES


def _apply_standard_review_purpose(
    snapshot: dict[str, Any],
    review_purpose: str,
) -> dict[str, Any]:
    """Apply the explicit PKG-STD-011 process-acceptance boundary.

    S001/S002 establish the public framework and revision status.  They do not
    replace the restricted S003 methodology text.  Therefore the package may
    support a framework-only process review, while project delivery remains
    incomplete and must not claim full-methodology conformance.
    """

    purpose = str(review_purpose or "project_delivery")
    result = {
        **snapshot,
        "packages": [dict(row) for row in snapshot.get("packages") or []],
        "incomplete": list(snapshot.get("incomplete") or []),
        "framework_only": [],
        "review_purpose": purpose,
    }
    if purpose == "process_acceptance":
        for package in result["packages"]:
            if str(package.get("package_id") or "") != "PKG-STD-011":
                continue
            artifacts = {
                str(item.get("artifact_id") or ""): item
                for item in package.get("artifacts") or []
                if isinstance(item, dict)
            }
            # 「官方框架依据」不能只看 sha256 非空：还要求来源确为官方发布、
            # 且该条目自身不处于抓取/解析失败态，否则一份下载失败的占位记录
            # 也会被当成框架已具备。
            public_framework_present = all(
                _is_official_framework_artifact(artifacts.get(item_id))
                for item_id in ("S001", "S002")
            )
            methodology_present = bool(artifacts.get("S003", {}).get("sha256"))
            if public_framework_present and not methodology_present:
                package["process_acceptance_status"] = "framework_only"
                package["framework_only"] = True
                result["framework_only"].append("PKG-STD-011")
                result["incomplete"] = [
                    item for item in result["incomplete"]
                    if item != "PKG-STD-011"
                ]
    selected = {
        "schema_version": result.get("schema_version"),
        "review_purpose": purpose,
        "packages": result["packages"],
        "incomplete": sorted(set(result["incomplete"])),
        "framework_only": sorted(set(result["framework_only"])),
    }
    result["incomplete"] = selected["incomplete"]
    result["framework_only"] = selected["framework_only"]
    result["content_hash"] = sha256_json(selected) if result["packages"] else None
    return result


def finding(rule_id: str, severity: str, message: str, *, category: str, blocking: bool | None = None,
            expected: Any = None, actual: Any = None, difference: Any = None, tolerance: Any = None,
            target_location: dict[str, Any] | None = None, evidence: list[dict[str, Any]] | None = None,
            standard_basis: list[dict[str, Any]] | None = None, review_area: str = "",
            remediation: str = "", confidence: float = 1.0, source_issue_id: str = "") -> dict[str, Any]:
    stable = {"rule_id": rule_id, "category": category, "target_location": target_location or {}, "source_issue_id": source_issue_id}
    return {
        "finding_id": "fnd_" + sha256_json(stable).removeprefix("sha256:")[:24],
        "rule_id": rule_id, "category": category, "severity": severity,
        "blocking": severity == "P0" or (severity == "P1" if blocking is None else bool(blocking)),
        "confidence": max(0.0, min(float(confidence), 1.0)), "message": message,
        "expected": expected, "actual": actual, "difference": difference, "tolerance": tolerance,
        "calculation_trace": [], "target_location": target_location or {},
        "evidence": evidence or [], "standard_basis": standard_basis or [],
        "remediation": remediation, "review_area": review_area,
        "status": "open", "source_issue_id": source_issue_id, "history": [],
    }


def _translated_formula(source: Any, destination: str) -> str | None:
    value = source.value
    if not isinstance(value, str) or not value.startswith("="):
        return None
    try:
        from openpyxl.formula.translate import Translator

        return Translator(value, origin=source.coordinate).translate_formula(destination)
    except Exception:  # noqa: BLE001
        return None


def _formula_shape(formula: str) -> tuple[tuple[str, str, str], ...] | None:
    """Retain operators/literals while removing cell-reference locations."""

    try:
        from openpyxl.formula import Tokenizer

        return tuple(
            (
                "<REFERENCE>" if token.type == "OPERAND" and token.subtype == "RANGE" else token.value,
                token.type,
                token.subtype,
            )
            for token in Tokenizer(formula).items
        )
    except Exception:  # noqa: BLE001
        return None


def _copied_formula_findings(book: Any, path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Find only anomalies independently corroborated from both neighbours."""

    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    metrics = {
        "hardcoded_calculation_findings": 0,
        "formula_copy_inconsistencies": 0,
    }
    for sheet in book.worksheets:
        populated = {
            (cell.row, cell.column): cell
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        for (row_index, column_index), cell in populated.items():
            directions = (
                (
                    "horizontal",
                    (
                        (row_index, column_index - 2),
                        (row_index, column_index - 1),
                        (row_index, column_index + 1),
                        (row_index, column_index + 2),
                    ),
                ),
                (
                    "vertical",
                    (
                        (row_index - 2, column_index),
                        (row_index - 1, column_index),
                        (row_index + 1, column_index),
                        (row_index + 2, column_index),
                    ),
                ),
            )
            for direction, neighbour_keys in directions:
                neighbours = [populated.get(key) for key in neighbour_keys]
                if any(neighbour is None for neighbour in neighbours):
                    continue
                if any(neighbour.style_id != cell.style_id for neighbour in neighbours):
                    continue
                expected_candidates = {
                    _translated_formula(neighbour, cell.coordinate)
                    for neighbour in neighbours
                }
                if None in expected_candidates or len(expected_candidates) != 1:
                    continue
                expected_formula = expected_candidates.pop()
                location = {
                    "workbook": path.name,
                    "sheet": sheet.title,
                    "cell": cell.coordinate,
                    "direction": direction,
                }
                value = cell.value
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    dedupe_key = (sheet.title, cell.coordinate, "hardcoded")
                    if dedupe_key in seen:
                        continue
                    seen.add(dedupe_key)
                    rows.append(finding(
                        "FIN.XLSX.HARDCODED.CALCULATION",
                        "P1",
                        "连续公式计算区被硬编码数值覆盖",
                        category="formula",
                        expected=expected_formula,
                        actual=value,
                        target_location={**location, "expected_formula": expected_formula},
                        review_area="finance",
                        remediation="恢复可追溯公式，核对该期输入与相邻期间后重新计算并复测",
                        confidence=0.95,
                    ))
                    metrics["hardcoded_calculation_findings"] += 1
                    continue
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                if value == expected_formula:
                    continue
                expected_shape = _formula_shape(expected_formula)
                if expected_shape is None or _formula_shape(value) != expected_shape:
                    continue
                dedupe_key = (sheet.title, cell.coordinate, "copy")
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                rows.append(finding(
                    "FIN.XLSX.FORMULA_COPY.INCONSISTENT",
                    "P1",
                    "连续计算区中的复制公式与前后期间模式不一致",
                    category="formula",
                    expected=expected_formula,
                    actual=value,
                    target_location={**location, "formula": value},
                    review_area="finance",
                    remediation="按前后期间一致的引用模式修复公式，并核对是否发生期间错位",
                    confidence=0.95,
                ))
                metrics["formula_copy_inconsistencies"] += 1
    return rows, metrics


def _cleanup_locator(path: Path, locator: str) -> dict[str, Any]:
    raw_locations = [item.strip() for item in locator.split(", ") if item.strip()]
    if len(raw_locations) > 1:
        locations = []
        for raw in raw_locations:
            sheet, separator, cell = raw.rpartition("!")
            locations.append(
                {"sheet": sheet, "cell": cell}
                if separator
                else {"locator": raw}
            )
        return {"workbook": path.name, "locations": locations}
    sheet, separator, cell = locator.rpartition("!")
    if not separator:
        return {"workbook": path.name, "locator": locator}
    return {"workbook": path.name, "sheet": sheet, "cell": cell}


def _vendor_cleanup_findings(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Project the existing read-only vendor checks into unified findings."""

    try:
        from lvke_mcp.domains.finance import vendor_import

        reference_pack = vendor_import.build_reference_pack(path)
        cleanup_rows = vendor_import.detect_cleanup_issues(reference_pack)
    except Exception:  # noqa: BLE001
        return [], ["vendor_cleanup_scan_unavailable"]

    rows: list[dict[str, Any]] = []
    for item in cleanup_rows:
        kind = str(item.get("type") or "")
        locator = str(item.get("locator") or "")
        location = _cleanup_locator(path, locator)
        if kind == "project_cashflow_financing_duplication":
            rows.append(finding(
                "FIN.XLSX.CASHFLOW.FINANCING_DUPLICATION",
                "P0",
                str(item.get("detail") or "项目投资现金流混入融资现金流"),
                category="cashflow",
                actual=item.get("vendor_value"),
                target_location=location,
                review_area="finance",
                remediation=str(item.get("engine_suggestion") or "分离项目投资与资本金现金流口径"),
            ))
            continue
        if kind == "hardcoded_irr_trial":
            residual = item.get("npv_residual_wan")
            tolerance = item.get("tolerance_wan")
            residual_exceeded = (
                residual is None
                or tolerance is None
                or abs(float(residual)) > float(tolerance)
            )
            row = finding(
                "FIN.XLSX.IRR.HARDCODED",
                "P0" if residual_exceeded else "P1",
                str(item.get("detail") or "IRR 使用硬编码试算值"),
                category="profitability",
                expected=item.get("engine_suggestion"),
                actual=item.get("vendor_value"),
                difference=residual,
                tolerance=tolerance,
                target_location=location,
                review_area="finance",
                remediation="使用确定性 IRR 求解器并固化现金流、求解结果与 NPV 残差",
            )
            row["calculation_trace"] = [{
                "operation": "npv_residual_at_vendor_trial_rate",
                "result_wan": residual,
                "tolerance_wan": tolerance,
            }]
            rows.append(row)
            continue
        if kind == "orphan_constant_formula":
            rows.append(finding(
                "FIN.XLSX.CONSTANT_FORMULA.ORPHAN",
                "P2",
                str(item.get("detail") or "发现无下游引用的常量试算公式"),
                category="formula",
                expected="有来源的独立输入或接入有效公式链",
                actual=item.get("vendor_value"),
                target_location={**location, "formula": item.get("vendor_value")},
                review_area="finance",
                remediation=str(item.get("engine_suggestion") or "删除孤儿试算式或补齐输入来源"),
            ))
    return rows, []


def scan_xlsx(path: Path, *, deep: bool) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    incomplete: list[str] = []
    metrics = {
        "sheets": 0,
        "formula_cells": 0,
        "hardcoded_numeric_cells": 0,
        "hardcoded_calculation_findings": 0,
        "formula_copy_inconsistencies": 0,
        "vendor_cleanup_findings": 0,
    }
    try:
        import openpyxl
        book = openpyxl.load_workbook(
            path, data_only=False, read_only=False, keep_links=True,
            keep_vba=path.suffix.lower() == ".xlsm",
        )
    except Exception:  # noqa: BLE001
        return rows, ["xlsx_parser_unavailable"], metrics
    metrics["sheets"] = len(book.sheetnames)
    if getattr(book, "_external_links", None):
        rows.append(finding(
            "FIN.XLSX.EXTERNAL_LINKS.PRESENT", "P1", "工作簿包含外部链接关系",
            category="workbook_integrity", actual=len(book._external_links),  # noqa: SLF001
            target_location={"workbook": path.name}, review_area="finance",
            remediation="移除外链或随审查包固化并校验全部外部依赖",
        ))
    for defined_name in book.defined_names.values():
        attr_text = str(getattr(defined_name, "attr_text", "") or "")
        if "#REF!" in attr_text.upper():
            rows.append(finding(
                "FIN.XLSX.DEFINED_NAME.INVALID", "P1", "命名区域包含失效引用",
                category="workbook_integrity", actual=attr_text,
                target_location={"workbook": path.name, "defined_name": str(getattr(defined_name, "name", ""))},
                review_area="finance", remediation="修复或删除失效命名区域",
            ))
    for sheet in book.worksheets:
        if sheet.sheet_state != "visible":
            rows.append(finding("FIN.XLSX.HIDDEN_SHEET", "P2", f"存在隐藏工作表：{sheet.title}", category="workbook_integrity",
                                target_location={"workbook": path.name, "sheet": sheet.title}, review_area="finance"))
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                location = {"workbook": path.name, "sheet": sheet.title, "cell": cell.coordinate}
                if isinstance(value, str) and value.startswith("="):
                    metrics["formula_cells"] += 1
                    upper = value.upper()
                    if any(token in upper for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                        rows.append(finding("FIN.XLSX.FORMULA_ERROR", "P0", "公式包含错误标记", category="formula", actual=value,
                                            target_location={**location, "formula": value}, review_area="finance", remediation="修复公式引用后重新导出并复测"))
                    if re.search(r"\[[^]]+\](?:[^!]+)!", value):
                        rows.append(finding("FIN.XLSX.EXTERNAL_LINK", "P1", "公式依赖外部工作簿", category="formula", actual=value,
                                            target_location={**location, "formula": value}, review_area="finance"))
                    if any(f"{name}(" in upper for name in ("NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT")):
                        rows.append(finding("FIN.XLSX.VOLATILE_FORMULA", "P2", "使用易变函数，复现性受影响", category="formula", actual=value,
                                            target_location={**location, "formula": value}, review_area="finance"))
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    metrics["hardcoded_numeric_cells"] += 1
                elif isinstance(value, str) and value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!"}:
                    rows.append(finding("FIN.XLSX.CACHED_ERROR", "P0", "单元格为 Excel 错误值", category="formula", actual=value,
                                        target_location=location, review_area="finance"))
    if path.suffix.lower() == ".xlsm" or getattr(book, "vba_archive", None) is not None:
        incomplete.append("xlsx_macro_or_udf_requires_manual_review")
    calculation = getattr(book, "calculation", None)
    if calculation is not None and bool(getattr(calculation, "iterate", False)):
        incomplete.append("xlsx_iterative_calculation_or_circular_reference")
    copied_rows, copied_metrics = _copied_formula_findings(book, path)
    rows.extend(copied_rows)
    metrics.update(copied_metrics)
    book.close()
    cleanup_rows, cleanup_missing = _vendor_cleanup_findings(path)
    rows.extend(cleanup_rows)
    incomplete.extend(cleanup_missing)
    metrics["vendor_cleanup_findings"] = len(cleanup_rows)
    return rows, incomplete, metrics


def recalculate_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    """Recalculate an isolated read-only copy with a pinned LibreOffice worker."""

    findings: list[dict[str, Any]] = []
    incomplete: list[str] = []
    metrics: dict[str, Any] = {
        "available": False, "formula_cells": 0, "empty_formula_caches": 0,
        "formula_errors": 0, "input_unchanged": False, "worker_version": "",
    }
    binary = os.environ.get("LVKE_REVIEW_SOFFICE") or shutil.which("soffice") or shutil.which("libreoffice")
    if not binary:
        return findings, ["libreoffice_recalc_worker_unavailable"], metrics
    original_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    try:
        version_result = subprocess.run(
            [binary, "--version"], check=True, capture_output=True, text=True, timeout=15,
        )
        version = (version_result.stdout or version_result.stderr or "").strip()[:200]
    except (OSError, subprocess.SubprocessError):
        return findings, ["libreoffice_recalc_worker_version_unavailable"], metrics
    metrics["available"] = True
    metrics["worker_version"] = version
    expected_version = str(os.environ.get("LVKE_REVIEW_LIBREOFFICE_VERSION") or "").strip()
    if expected_version and expected_version not in version:
        return findings, ["libreoffice_recalc_worker_version_mismatch"], metrics
    try:
        import openpyxl

        with tempfile.TemporaryDirectory(prefix="lvke-review-recalc-") as temp_dir:
            root = Path(temp_dir)
            source_dir = root / "source"
            output_dir = root / "output"
            profile_dir = root / "profile"
            source_dir.mkdir()
            output_dir.mkdir()
            profile_dir.mkdir()
            isolated = source_dir / ("review-input.xlsm" if path.suffix.lower() == ".xlsm" else "review-input.xlsx")
            shutil.copy2(path, isolated)
            isolated.chmod(0o400)
            profile_uri = profile_dir.resolve().as_uri()
            subprocess.run(
                [
                    binary, "--headless", "--nologo", "--nodefault", "--nofirststartwizard",
                    f"-env:UserInstallation={profile_uri}", "--convert-to", "xlsx",
                    "--outdir", str(output_dir), str(isolated),
                ],
                check=True, capture_output=True, text=True, timeout=180,
            )
            recalculated = output_dir / "review-input.xlsx"
            if not recalculated.is_file():
                return findings, ["libreoffice_recalc_output_missing"], metrics
            formula_book = openpyxl.load_workbook(isolated, data_only=False, read_only=True, keep_links=False)
            value_book = openpyxl.load_workbook(recalculated, data_only=True, read_only=True, keep_links=False)
            for formula_sheet in formula_book.worksheets:
                if formula_sheet.title not in value_book.sheetnames:
                    incomplete.append(f"libreoffice_recalc_sheet_missing:{formula_sheet.title}")
                    continue
                value_sheet = value_book[formula_sheet.title]
                for row in formula_sheet.iter_rows():
                    for formula_cell in row:
                        formula = formula_cell.value
                        if not (isinstance(formula, str) and formula.startswith("=")):
                            continue
                        metrics["formula_cells"] += 1
                        value = value_sheet[formula_cell.coordinate].value
                        location = {
                            "workbook": path.name, "sheet": formula_sheet.title,
                            "cell": formula_cell.coordinate, "formula": formula,
                        }
                        if value is None:
                            metrics["empty_formula_caches"] += 1
                            findings.append(finding(
                                "FIN.XLSX.EMPTY_FORMULA_CACHE", "P1", "隔离重算后公式缓存仍为空",
                                category="formula_recalculation", expected="calculated value", actual=None,
                                target_location=location, review_area="finance",
                                remediation="修复不受支持公式、外链或计算设置后重新导出",
                            ))
                        elif isinstance(value, str) and value.upper().startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!")):
                            metrics["formula_errors"] += 1
                            findings.append(finding(
                                "FIN.XLSX.RECALCULATED.ERROR", "P0", "隔离重算产生公式错误",
                                category="formula_recalculation", expected="valid calculated value", actual=value,
                                target_location=location, review_area="finance",
                                remediation="修复公式、外链或输入后重新导出",
                            ))
            formula_book.close()
            value_book.close()
    except subprocess.TimeoutExpired:
        incomplete.append("libreoffice_recalc_timeout")
    except (OSError, subprocess.SubprocessError, ValueError):
        incomplete.append("libreoffice_recalc_failed")
    metrics["input_unchanged"] = hashlib.sha256(path.read_bytes()).hexdigest() == original_hash
    if not metrics["input_unchanged"]:
        incomplete.append("libreoffice_recalc_mutated_original")
    if metrics["formula_cells"] and metrics["empty_formula_caches"] == metrics["formula_cells"]:
        incomplete.append("libreoffice_recalc_incomplete")
    return findings, sorted(set(incomplete)), metrics
