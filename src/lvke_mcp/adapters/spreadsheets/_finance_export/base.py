"""错误类型、必需公式族、openpyxl 依赖检查与 soffice 重算。"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from lvke_mcp.domains.finance._run_service.base import (
    DELIVERY_TABLE_META,
    delivery_count_semantics,
    delivery_table_contract_hash,
)


logger = logging.getLogger(__name__)


class FinanceExportError(RuntimeError):
    """Raised when export cannot proceed (missing openpyxl / bad path)."""


_REQUIRED_FORMULA_FAMILIES: dict[str, tuple[str, ...]] = {
    "investment": ("quantity_indicator",),
    "interest-during-construction": ("interest_rollforward",),
    "working-capital": (
        "inventory_component_amount", "inventory_component_sum", "working_capital_total",
    ),
    "funding": ("funding_uses_sources",),
    "income-statement": ("product_revenue", "vat_breakdown"),
    "total-cost": ("cost_item_sum", "total_cost_sum"),
    "wage": ("staff_wage", "wage_total"),
    "depreciation": ("asset_class_depreciation", "rollforward"),
    "amortization": ("amortization",),
    "profit-distribution": ("profit_distribution",),
    "debt-service": ("debt_rollforward", "repayment_coverage"),
    "cashflow": ("net_cashflow",),
    "capital-cashflow": ("capital_cashflow_atomic", "net_cashflow"),
}


def _require_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # noqa: BLE001
        raise FinanceExportError(
            "Excel 导出需要 openpyxl，当前环境不可用"
        ) from exc
    return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter


def _recalculate_with_soffice(path: Path) -> dict[str, Any]:
    """Run a real headless spreadsheet recalc and verify cached data-only values."""
    binary = os.environ.get("SOFFICE") or os.environ.get("LIBREOFFICE") or shutil.which("soffice")
    if not binary:
        return {"ok": False, "available": False, "issues": ["soffice/LibreOffice 不可用"]}
    try:
        with tempfile.TemporaryDirectory(prefix="lvke-soffice-") as tmp:
            subprocess.run(
                [binary, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
                check=True, capture_output=True, text=True, timeout=120,
            )
            recalculated = Path(tmp) / path.name
            if not recalculated.is_file():
                return {"ok": False, "available": True, "issues": ["LibreOffice 未生成重算工作簿"]}
            import openpyxl

            wb = openpyxl.load_workbook(recalculated, data_only=True, read_only=True)
            formula_cache_empty = False
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("#"):
                            return {"ok": False, "available": True, "issues": [f"公式错误 {sheet.title}!{cell.coordinate}: {cell.value}"]}
            # At least one calculated numeric cell must exist in every delivery sheet.
            for sheet in _DELIVERY_SHEETS.values():
                values = [cell.value for row in wb[sheet].iter_rows() for cell in row]
                if not any(isinstance(value, (int, float)) for value in values):
                    formula_cache_empty = True
            if formula_cache_empty:
                return {"ok": False, "available": True, "issues": ["重算后 data_only 缓存为空"]}
            return {"ok": True, "available": True, "issues": []}
    except Exception:  # noqa: BLE001
        logger.exception("LibreOffice 工作簿重算失败")
        return {
            "ok": False,
            "available": True,
            "issues": ["LibreOffice 工作簿重算失败"],
        }


# 工作簿交付 sheet 仅投影领域层的唯一十三表契约，不在适配器重复维护成员或顺序。
_DELIVERY_SHEETS = {
    table_code: delivery_no
    for table_code, delivery_no, _title in DELIVERY_TABLE_META
}
