"""Finance workbook export with formulas (P2-2).

Exports a review-friendly xlsx with the 13 delivery sheets, cross-sheet formulas,
cell lineage, template-gap disclosure, Inputs, Checks and Meta.
"""

from __future__ import annotations

import hashlib
import logging
import re
import os
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, Union

from filelock import FileLock


logger = logging.getLogger(__name__)


class FinanceExportError(RuntimeError):
    """Raised when export cannot proceed (missing openpyxl / bad path)."""


_REQUIRED_FORMULA_FAMILIES: dict[str, tuple[str, ...]] = {
    "investment": ("quantity_indicator",),
    "interest-during-construction": ("interest_rollforward",),
    "working-capital": (
        "inventory_component_amount", "inventory_component_sum", "working_capital_total",
    ),
    "funding": ("funding_uses_sources",),
    "income-statement": ("product_revenue", "vat_breakdown"),
    "total-cost": ("cost_item_sum", "total_cost_sum"),
    "wage": ("staff_wage", "wage_total"),
    "depreciation": ("asset_class_depreciation", "rollforward"),
    "amortization": ("amortization",),
    "profit-distribution": ("profit_distribution",),
    "debt-service": ("debt_rollforward", "repayment_coverage"),
    "cashflow": ("net_cashflow",),
    "capital-cashflow": ("capital_cashflow_atomic", "net_cashflow"),
}


def _require_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # noqa: BLE001
        raise FinanceExportError(
            "Excel 导出需要 openpyxl，当前环境不可用"
        ) from exc
    return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter


def _recalculate_with_soffice(path: Path) -> dict[str, Any]:
    """Run a real headless spreadsheet recalc and verify cached data-only values."""
    binary = os.environ.get("SOFFICE") or os.environ.get("LIBREOFFICE") or shutil.which("soffice")
    if not binary:
        return {"ok": False, "available": False, "issues": ["soffice/LibreOffice 不可用"]}
    try:
        with tempfile.TemporaryDirectory(prefix="lvke-soffice-") as tmp:
            subprocess.run(
                [binary, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
                check=True, capture_output=True, text=True, timeout=120,
            )
            recalculated = Path(tmp) / path.name
            if not recalculated.is_file():
                return {"ok": False, "available": True, "issues": ["LibreOffice 未生成重算工作簿"]}
            import openpyxl

            wb = openpyxl.load_workbook(recalculated, data_only=True, read_only=True)
            formula_cache_empty = False
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("#"):
                            return {"ok": False, "available": True, "issues": [f"公式错误 {sheet.title}!{cell.coordinate}: {cell.value}"]}
            # At least one calculated numeric cell must exist in every delivery sheet.
            for sheet in _DELIVERY_SHEETS.values():
                values = [cell.value for row in wb[sheet].iter_rows() for cell in row]
                if not any(isinstance(value, (int, float)) for value in values):
                    formula_cache_empty = True
            if formula_cache_empty:
                return {"ok": False, "available": True, "issues": ["重算后 data_only 缓存为空"]}
            return {"ok": True, "available": True, "issues": []}
    except Exception:  # noqa: BLE001
        logger.exception("LibreOffice 工作簿重算失败")
        return {
            "ok": False,
            "available": True,
            "issues": ["LibreOffice 工作簿重算失败"],
        }


def _input_rows(fin: dict[str, Any]) -> list[tuple[str, Any, str]]:
    schema = fin.get("schema") or {}
    norm = schema.get("normalized_inputs") or {}
    inv = fin.get("investment") or {}
    fund = fin.get("funding") or {}
    params = fin.get("params") or {}
    rows = [
        ("total_investment_wan", inv.get("total") or norm.get("total_investment_wan"), "万元"),
        ("construction_wan", inv.get("construction") or norm.get("construction_wan"), "万元"),
        ("interest_wan", inv.get("interest") or norm.get("interest_wan"), "万元"),
        ("working_capital_wan", inv.get("working_capital") or norm.get("working_capital_wan"), "万元"),
        ("equity_capital_wan", fund.get("capital"), "万元"),
        ("loan_wan", fund.get("loan"), "万元"),
        ("loan_rate", (fin.get("raw") or {}).get("loan_rate") or norm.get("loan_rate"), "小数"),
        ("calc_years", params.get("calc_years"), "年"),
        ("build_years", params.get("build_years") or params.get("build_period_years"), "年"),
        ("finance_schema_version", fin.get("finance_schema_version") or (schema.get("finance_schema_version")), ""),
    ]
    return [(k, v, u) for k, v, u in rows if v is not None and v != ""]


def _write_inputs(ws, fin, Font, PatternFill, Alignment, Border, Side):
    fill = PatternFill("solid", fgColor="DCEAF7")
    header_font = Font(bold=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    ws.append(["key", "value", "unit", "role"])
    for cell in ws[1]:
        cell.font = header_font
    for key, val, unit in _input_rows(fin):
        ws.append([key, val, unit, "input"])
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="left")
    # named-ish map for formulas: column B is value; row index by key
    key_rows = {}
    for r in range(2, ws.max_row + 1):
        key_rows[str(ws.cell(r, 1).value)] = r
    return key_rows


def _write_year_table(ws, rows: list[dict], fields: list[str], header_font):
    if not rows:
        ws.append(["(empty)"])
        return
    headers = ["year"] + fields
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
    for row in rows:
        year = row.get("year") if row.get("year") is not None else row.get("period")
        ws.append([year] + [row.get(f) for f in fields])


def _write_indicators(ws, fin, key_rows, Font):
    inv = fin.get("investment") or {}
    fund = fin.get("funding") or {}
    ind = fin.get("indicators") or {}
    header_font = Font(bold=True)
    ws.append(["metric", "value", "unit", "source"])
    for cell in ws[1]:
        cell.font = header_font

    def add(name, value, unit, source="engine"):
        ws.append([name, value, unit, source])

    # Prefer formulas referencing Inputs sheet when keys exist
    if "total_investment_wan" in key_rows:
        r = key_rows["total_investment_wan"]
        add("total_investment", f"=Inputs!B{r}", "万元", "formula→Inputs")
    else:
        add("total_investment", inv.get("total"), "万元")
    if "equity_capital_wan" in key_rows:
        r = key_rows["equity_capital_wan"]
        add("equity_capital", f"=Inputs!B{r}", "万元", "formula→Inputs")
    else:
        add("equity_capital", fund.get("capital"), "万元")
    if "loan_wan" in key_rows:
        r = key_rows["loan_wan"]
        add("loan", f"=Inputs!B{r}", "万元", "formula→Inputs")
    else:
        add("loan", fund.get("loan"), "万元")

    add("project_irr_pct", ind.get("project_irr_pct"), "%")
    add("npv_wan", ind.get("npv_wan"), "万元")
    add("static_payback_years", ind.get("static_payback_years"), "年")
    add("dynamic_payback_years", ind.get("dynamic_payback_years"), "年")
    add("bep_pct", ind.get("bep_pct"), "%")

    # funding sum formula if components present
    # locate rows just written for equity/loan if formula
    # Optional check row: capital + loan (values may be formulas)
    # Find equity/loan row numbers on this sheet
    equity_row = loan_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "equity_capital":
            equity_row = r
        if ws.cell(r, 1).value == "loan":
            loan_row = r
    if equity_row and loan_row:
        add("funding_sum", f"=B{equity_row}+B{loan_row}", "万元", "formula")


def _write_checks(ws, fin, Font):
    header_font = Font(bold=True)
    ws.append(["rule", "ok", "detail", "blocking"])
    for cell in ws[1]:
        cell.font = header_font
    checks = []
    try:
        from lvke_mcp.domains.finance import finance_model as fm

        checks = fm.check_consistency(fin) or []
    except Exception:  # noqa: BLE001
        checks = []
    for c in checks:
        ws.append([
            c.get("rule"),
            bool(c.get("ok")),
            c.get("detail"),
            bool(c.get("blocking")),
        ])


_DELIVERY_SHEETS = {
    "investment": "附表1",
    "interest-during-construction": "附表2",
    "working-capital": "附表3",
    "funding": "附表4",
    "income-statement": "附表5",
    "total-cost": "附表6",
    "wage": "附表6-1",
    "depreciation": "附表6-2",
    "amortization": "附表6-3",
    "profit-distribution": "附表7",
    "debt-service": "附表8",
    "cashflow": "附表9",
    "capital-cashflow": "附表10",
}


def _write_delivery_tables(wb, fin, Font, PatternFill, Alignment, Border, Side):
    """Write 13 delivery sheets and return formula lineage + quality metadata."""
    from lvke_mcp.domains.finance import table_render

    pack = table_render.build_all_structured(fin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Border(
        left=Side(style="thin", color="A6A6A6"), right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"), bottom=Side(style="thin", color="A6A6A6"),
    )
    lineage: list[dict[str, Any]] = []
    layouts: dict[str, dict[str, Any]] = {}
    product_schedule: dict[str, Any] = {}
    cost_schedule: dict[str, Any] = {}
    formula_families: dict[str, set[str]] = {key: set() for key in _DELIVERY_SHEETS}

    for key, sheet_name in _DELIVERY_SHEETS.items():
        table = pack.get(key) or {}
        ws = wb.create_sheet(sheet_name)
        columns = table.get("columns") or []
        keys = [str(c.get("key") or "") for c in columns]
        labels = [str(c.get("label") or "") for c in columns]
        ws.cell(1, 1, f"{table.get('delivery_no', sheet_name)} {table.get('title', '')}")
        ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=13)
        ws.cell(1, 1).fill = title_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(labels), 1))
        status_text = str(fin.get("calculation_status") or fin.get("assurance_level") or "computed")
        ws.cell(2, 1, f"计算状态：{status_text}；口径：{table.get('grade', 'summary')}；有效：{'是' if table.get('effective') else '否'}")
        for col, label in enumerate(labels, start=1):
            cell = ws.cell(3, col, label)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r_idx, row in enumerate(table.get("rows") or [], start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, value)
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                source_key = keys[c_idx - 1] if c_idx - 1 < len(keys) else f"col_{c_idx}"
                lineage.append({
                    "target": f"{sheet_name}!{cell.coordinate}",
                    "source": f"{table.get('source')}.{source_key}",
                    "method": "engine_value",
                })
        reference_layout = {
            "keys": keys,
            "first": 4,
            "last": 3 + len(table.get("rows") or []),
        }
        engine_columns = list(table.get("engine_columns") or [])
        engine_rows = list(table.get("engine_rows") or [])
        engine_keys = [str(column.get("key") or "") for column in engine_columns if isinstance(column, dict)]
        engine_labels = [str(column.get("label") or "") for column in engine_columns if isinstance(column, dict)]
        if engine_rows and engine_keys:
            ws.cell(ws.max_row + 2, 1, "引擎逐年计算区（供公式与跨表追溯）")
            engine_header = ws.max_row + 1
            for c_idx, label in enumerate(engine_labels, start=1):
                cell = ws.cell(engine_header, c_idx, label)
                cell.font = Font(bold=True)
                cell.fill = header_fill
            engine_first = engine_header + 1
            for row in engine_rows:
                ws.append(list(row))
                for cell in ws[ws.max_row]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
            engine_last = engine_first + len(engine_rows) - 1
        else:
            engine_first = reference_layout["first"]
            engine_last = reference_layout["last"]
            engine_keys = keys
        if key == "income-statement" and table.get("product_tree"):
            ws.cell(ws.max_row + 2, 1, "分产品量价与达产爬坡")
            product_header = ws.max_row + 1
            op_years = len(engine_rows or table.get("rows") or [])
            product_labels = ["产品", "单位", "单价", "达产产能"] + [f"运营年{y + 1}收入" for y in range(op_years)]
            for c_idx, label in enumerate(product_labels, start=1):
                ws.cell(product_header, c_idx, label).font = Font(bold=True)
            product_rows = []
            for product in table.get("product_tree") or []:
                row_idx = ws.max_row + 1
                ws.append([product.get("name"), product.get("unit"), product.get("price_per_unit"), product.get("capacity")])
                ramp = list(product.get("ramp") or [])
                if ramp:
                    ramp += [ramp[-1]] * (op_years - len(ramp))
                else:
                    ramp = [1.0] * op_years
                unit_scale = 10000.0 if str(product.get("unit") or "").startswith("万") else 1.0
                price_divisor = 1.0 if product.get("price_unit") == "wan" else 10000.0
                for y in range(op_years):
                    col_idx = 5 + y
                    formula = f"=$C{row_idx}*$D{row_idx}*{unit_scale:g}/{price_divisor:g}*{float(ramp[y]):g}"
                    ws.cell(row_idx, col_idx, formula)
                    ws.cell(row_idx, col_idx).number_format = '#,##0.00'
                    lineage.append({"target": f"{sheet_name}!{ws.cell(row_idx, col_idx).coordinate}",
                                    "source": "spec.revenue.products.price*capacity*ramp", "method": "excel_formula"})
                    formula_families["income-statement"].add("product_revenue")
                product_rows.append(row_idx)
            product_schedule = {"header": product_header, "rows": product_rows, "first_year_col": 5, "years": op_years}
        if key == "total-cost":
            cost_items = (fin.get("input_revision") or fin.get("finance_inputs") or {}).get("cost_items") or {}
            if cost_items:
                ws.cell(ws.max_row + 2, 1, "成本费用明细树")
                cost_header = ws.max_row + 1
                op_rows = engine_rows or table.get("rows") or []
                op_keys = engine_keys or keys
                operating_cost_index = op_keys.index("operating_cost") if "operating_cost" in op_keys else -1
                cost_labels = ["成本项目"] + [f"运营年{y + 1}" for y in range(len(op_rows))]
                for c_idx, label in enumerate(cost_labels, start=1):
                    ws.cell(cost_header, c_idx, label).font = Font(bold=True)
                total_input = sum(float(v or 0.0) for v in cost_items.values()) or 1.0
                item_rows = []
                for name, amount in cost_items.items():
                    row_idx = ws.max_row + 1
                    ws.cell(row_idx, 1, str(name))
                    for y, annual_row in enumerate(op_rows):
                        operating_cost = (
                            annual_row[operating_cost_index]
                            if operating_cost_index >= 0 and len(annual_row) > operating_cost_index
                            else 0.0
                        )
                        allocated = round(float(operating_cost or 0.0) * float(amount or 0.0) / total_input, 2)
                        ws.cell(row_idx, 2 + y, allocated)
                        ws.cell(row_idx, 2 + y).number_format = '#,##0.00'
                    item_rows.append(row_idx)
                cost_schedule = {"header": cost_header, "rows": item_rows, "first_year_col": 2, "years": len(op_rows)}
        for idx, label in enumerate(table.get("notes") or [], start=1):
            ws.cell(ws.max_row + 1, 1, f"说明{idx}：{label}")
        for col in range(1, max(len(labels), 1) + 1):
            ws.column_dimensions[ws.cell(3, col).column_letter].width = 16
        ws.freeze_panes = "A4"
        layouts[key] = {
            "ws": ws,
            "keys": engine_keys,
            "first": engine_first,
            "last": engine_last,
            "reference": reference_layout,
            "reference_keys": keys,
        }

    def col(key: str, field: str) -> Optional[int]:
        keys = layouts.get(key, {}).get("keys") or []
        try:
            return keys.index(field) + 1
        except ValueError:
            return None

    def set_formula(
        target_key: str,
        row: int,
        field: str,
        formula: str,
        source: str,
        *,
        family: str = "",
    ) -> None:
        c = col(target_key, field)
        if not c:
            return
        ws = layouts[target_key]["ws"]
        cell = ws.cell(row, c, formula)
        cell.number_format = '#,##0.00'
        lineage.append({"target": f"{ws.title}!{cell.coordinate}", "source": source, "method": "excel_formula"})
        if family:
            formula_families[target_key].add(family)

    def row_by_label(key: str, label: str, field: str = "name") -> Optional[int]:
        c = col(key, field)
        if not c:
            return None
        layout = layouts.get(key) or {}
        ws = layout.get("ws")
        for r in range(layout.get("first", 4), layout.get("last", 3) + 1):
            if str(ws.cell(r, c).value or "").strip() == label:
                return r
        return None

    def reference_row_by_label(key: str, label: str) -> Optional[int]:
        layout = layouts.get(key) or {}
        reference = layout.get("reference") or {}
        reference_keys = list(layout.get("reference_keys") or [])
        label_key = "item" if "item" in reference_keys else ("name" if "name" in reference_keys else "")
        if not label_key:
            return None
        item_col = reference_keys.index(label_key) + 1
        ws = layout.get("ws")
        for row in range(int(reference.get("first") or 4), int(reference.get("last") or 3) + 1):
            if str(ws.cell(row, item_col).value or "").strip() == label:
                return row
        return None

    # 附表4：用途和来源父行均由原子子项逐年求和，禁止把静态投影当公式覆盖。
    funding_layout = layouts.get("funding") or {}
    funding_keys = list(funding_layout.get("reference_keys") or [])
    funding_ws = funding_layout.get("ws")
    if funding_ws is not None:
        use_parent = reference_row_by_label("funding", "项目总投资使用计划")
        source_parent = reference_row_by_label("funding", "资金筹措")
        use_children = [
            reference_row_by_label("funding", label)
            for label in ("建设投资", "建设期贷款利息", "流动资金")
        ]
        source_children = [
            reference_row_by_label("funding", label)
            for label in ("项目资本金（自筹）", "银行贷款", "政府补助/专项债")
        ]
        source_children = [row for row in source_children if row]
        formula_columns = [
            index + 1 for index, key in enumerate(funding_keys)
            if key in {"total", "amount"} or key.startswith("period_") or key.startswith("year_")
        ]
        if use_parent and all(use_children) and source_parent and source_children:
            for column in formula_columns:
                use_refs = [funding_ws.cell(row, column).coordinate for row in use_children]
                source_refs = [funding_ws.cell(row, column).coordinate for row in source_children]
                funding_ws.cell(use_parent, column, "=SUM(" + ",".join(use_refs) + ")")
                funding_ws.cell(source_parent, column, "=SUM(" + ",".join(source_refs) + ")")
            formula_families["funding"].add("funding_uses_sources")

    # 附表1：工程量×估算指标形成金额，合计和比例由公式计算。
    inv_layout = layouts.get("investment") or {}
    for r in range(inv_layout.get("first", 4), inv_layout.get("last", 3) + 1):
        no_col, qty_col, ind_col = col("investment", "no"), col("investment", "quantity"), col("investment", "indicator")
        total_col, pct_col = col("investment", "total"), col("investment", "pct")
        no = str(inv_layout["ws"].cell(r, no_col).value or "") if no_col else ""
        if no.count(".") == 1 and qty_col and ind_col and total_col:
            qty_ref = inv_layout["ws"].cell(r, qty_col).coordinate
            ind_ref = inv_layout["ws"].cell(r, ind_col).coordinate
            formula = f"={qty_ref}*{ind_ref}/10000"
            inv_layout["ws"].cell(r, total_col, formula)
            lineage.append({"target": f"附表1!{inv_layout['ws'].cell(r, total_col).coordinate}",
                            "source": f"附表1!{qty_ref}*{ind_ref}", "method": "excel_formula"})
            formula_families["investment"].add("quantity_indicator")
            for bucket in ("civil", "equipment", "installation", "other"):
                bucket_col = col("investment", bucket)
                if bucket_col and inv_layout["ws"].cell(r, bucket_col).value not in (None, ""):
                    inv_layout["ws"].cell(r, bucket_col, f"={inv_layout['ws'].cell(r, total_col).coordinate}")
            if pct_col:
                total_row = inv_layout.get("last")
                total_ref = inv_layout["ws"].cell(total_row, total_col).coordinate
                inv_layout["ws"].cell(r, pct_col, f"={inv_layout['ws'].cell(r, total_col).coordinate}/{total_ref}*100")

    # 附表5：主表收入由产品量价矩阵逐年汇总，税费由收入/成本公式驱动。
    inc = layouts.get("income-statement") or {}
    if product_schedule:
        raw = fin.get("raw") or {}
        vat_rate = float((fin.get("indicators") or {}).get("vat_rate") or 0.0)
        vat_input_rate = float((fin.get("indicators") or {}).get("vat_input_rate") or 0.0)
        surtax_policy = raw.get("surtax_component_policy") or {}
        surtax_rate = float(
            surtax_policy.get("combined_rate") or raw.get("surtax_vat_rate") or 0.12
        )
        consumption_tax = raw.get("consumption_tax_payable_wan")
        if consumption_tax is None:
            consumption_tax = (fin.get("finance_inputs") or {}).get(
                "consumption_tax_payable_wan"
            )

        def consumption_tax_for_year(index: int) -> float:
            if isinstance(consumption_tax, (list, tuple)):
                if not consumption_tax:
                    return 0.0
                return max(float(consumption_tax[min(index, len(consumption_tax) - 1)] or 0.0), 0.0)
            return max(float(consumption_tax or 0.0), 0.0)

        for y, row in enumerate(range(inc.get("first", 4), inc.get("last", 3) + 1)):
            pcol = product_schedule["first_year_col"] + y
            refs = [f"'{inc['ws'].title}'!{inc['ws'].cell(pr, pcol).coordinate}" for pr in product_schedule["rows"]]
            if refs:
                set_formula(
                    "income-statement", row, "revenue", "=SUM(" + ",".join(refs) + ")",
                    "+".join(refs), family="product_revenue",
                )
            rev_cell = inc["ws"].cell(row, col("income-statement", "revenue")).coordinate
            op_cell = inc["ws"].cell(row, col("income-statement", "operating_cost")).coordinate
            set_formula(
                "income-statement", row, "vat_output", f"={rev_cell}*{vat_rate:g}",
                rev_cell, family="vat_breakdown",
            )
            set_formula(
                "income-statement", row, "vat_input", f"={op_cell}*{vat_input_rate:g}",
                op_cell, family="vat_breakdown",
            )
            vo = inc["ws"].cell(row, col("income-statement", "vat_output")).coordinate
            vi = inc["ws"].cell(row, col("income-statement", "vat_input")).coordinate
            set_formula(
                "income-statement", row, "vat_payable", f"=MAX({vo}-{vi},0)",
                f"{vo}-{vi}", family="vat_breakdown",
            )
            vp = inc["ws"].cell(row, col("income-statement", "vat_payable")).coordinate
            if raw.get("surtax_on_vat"):
                tax_base = vp
                if surtax_policy.get("mode") == "statutory_components":
                    consumption_value = consumption_tax_for_year(y)
                    if consumption_value:
                        tax_base = f"({vp}+{consumption_value:g})"
                set_formula(
                    "income-statement", row, "tax_surtax", f"={tax_base}*{surtax_rate:g}",
                    tax_base, family="vat_breakdown",
                )

    # 附表6：经营成本由成本明细树汇总，总成本由各组成项公式加总。
    cost = layouts.get("total-cost") or {}
    for y, row in enumerate(range(cost.get("first", 4), cost.get("last", 3) + 1)):
        if cost_schedule:
            cidx = cost_schedule["first_year_col"] + y
            refs = [f"'{cost['ws'].title}'!{cost['ws'].cell(cr, cidx).coordinate}" for cr in cost_schedule["rows"]]
            set_formula(
                "total-cost", row, "operating_cost", "=SUM(" + ",".join(refs) + ")",
                "+".join(refs), family="cost_item_sum",
            )
        parts = [col("total-cost", x) for x in ("operating_cost", "depreciation", "amortization", "interest")]
        if all(parts):
            coords = [cost["ws"].cell(row, c).coordinate for c in parts]
            set_formula(
                "total-cost", row, "total_cost", "=SUM(" + ",".join(coords) + ")",
                "+".join(coords), family="total_cost_sum",
            )

    # 附表6-2：累计折旧和净值滚动公式。
    dep = layouts.get("depreciation") or {}
    for offset, row in enumerate(range(dep.get("first", 4), dep.get("last", 3) + 1)):
        dep_cell = dep["ws"].cell(row, col("depreciation", "depreciation")).coordinate
        cum_col = col("depreciation", "cumulative_depreciation")
        net_col = col("depreciation", "net_value")
        orig = dep["ws"].cell(row, col("depreciation", "original_value")).coordinate
        if cum_col:
            formula = f"={dep_cell}" if offset == 0 else f"={dep['ws'].cell(row-1, cum_col).coordinate}+{dep_cell}"
            dep["ws"].cell(row, cum_col, formula)
            formula_families["depreciation"].add("rollforward")
        if net_col and cum_col:
            dep["ws"].cell(row, net_col, f"={orig}-{dep['ws'].cell(row, cum_col).coordinate}")
            formula_families["depreciation"].add("rollforward")

    # 附表2：以不可变 FinanceRun 的分年计息口径为唯一真源。半期计息行可在
    # 表内独立复算；显式利息计划不能被不兼容的提款假设覆盖为另一个结果。
    idc = layouts.get("interest-during-construction") or {}
    idc_rows = (fin.get("annual") or {}).get("interest_during_construction") or []
    for offset, row in enumerate(range(idc.get("first", 4), idc.get("last", 3) + 1)):
        begin_col = col("interest-during-construction", "begin_balance")
        draw_col = col("interest-during-construction", "draw")
        rate_col = col("interest-during-construction", "rate")
        interest_col = col("interest-during-construction", "interest")
        end_col = col("interest-during-construction", "end_balance")
        if all((begin_col, draw_col, rate_col, interest_col, end_col)):
            if offset > 0:
                idc["ws"].cell(row, begin_col, f"={idc['ws'].cell(row-1, end_col).coordinate}")
            begin = idc["ws"].cell(row, begin_col).coordinate
            draw = idc["ws"].cell(row, draw_col).coordinate
            rate = idc["ws"].cell(row, rate_col).coordinate
            schedule_row = idc_rows[offset] if offset < len(idc_rows) else {}
            basis = str(schedule_row.get("calculation_basis") or "")
            stored_interest = float(schedule_row.get("interest") or 0.0)
            if basis == "half_year_average_balance":
                formula = f"=ROUND(({begin}+{draw}/2)*{rate},2)"
                source = f"({begin}+{draw}/2)*{rate}"
            else:
                formula = f"=ROUND({stored_interest:.2f},2)"
                source = (
                    "FinanceRun.annual.interest_during_construction"
                    f"[{offset}].interest"
                )
            idc["ws"].cell(row, interest_col, formula)
            idc["ws"].cell(row, end_col, f"={begin}+{draw}")
            lineage.append({
                "target": f"附表2!{idc['ws'].cell(row, interest_col).coordinate}",
                "source": source,
                "method": basis or "immutable_schedule",
            })
            formula_families["interest-during-construction"].add("interest_rollforward")

    # 附表3：周转天数法在 Excel 内独立复算，并引用收入/成本表达产基数。
    wc_layout = layouts.get("working-capital") or {}
    base_col, days_col, amount_col = (
        col("working-capital", "base"), col("working-capital", "days"), col("working-capital", "amount")
    )
    wc_rows = {}
    if base_col and days_col and amount_col:
        for label in (
            "应收账款", "存货", "原材料", "燃料及动力", "在产品", "产成品",
            "现金", "流动资产小计", "应付账款", "减：应付账款",
            "流动资金", "新增流动资金合计",
        ):
            rr = row_by_label("working-capital", label, field="item")
            if rr:
                wc_rows[label] = rr
        peak_inc_row = inc.get("last", 3)
        peak_cost_row = cost.get("last", 3)
        for label in ("应收账款", "现金", "应付账款", "减：应付账款"):
            rr = wc_rows.get(label)
            if not rr:
                continue
            if label == "应收账款":
                ref = f"'附表5'!{inc['ws'].cell(peak_inc_row, col('income-statement', 'revenue')).coordinate}"
            else:
                ref = f"'附表6'!{cost['ws'].cell(peak_cost_row, col('total-cost', 'operating_cost')).coordinate}"
            wc_layout["ws"].cell(rr, base_col, f"={ref}")
            base = wc_layout["ws"].cell(rr, base_col).coordinate
            days = wc_layout["ws"].cell(rr, days_col).coordinate
            wc_layout["ws"].cell(rr, amount_col, f"={base}*{days}/360")
            lineage.append({"target": f"附表3!{wc_layout['ws'].cell(rr, amount_col).coordinate}",
                            "source": ref, "method": "excel_formula"})
            formula_families["working-capital"].add("turnover_amount")
        component_refs = []
        for label in ("原材料", "燃料及动力", "在产品", "产成品"):
            rr = wc_rows.get(label)
            if not rr:
                continue
            base = wc_layout["ws"].cell(rr, base_col).coordinate
            days = wc_layout["ws"].cell(rr, days_col).coordinate
            wc_layout["ws"].cell(rr, amount_col, f"={base}*{days}/360")
            component_refs.append(wc_layout["ws"].cell(rr, amount_col).coordinate)
            formula_families["working-capital"].add("inventory_component_amount")
        inventory_row = wc_rows.get("存货")
        if inventory_row and len(component_refs) == 4:
            wc_layout["ws"].cell(inventory_row, amount_col, "=SUM(" + ",".join(component_refs) + ")")
            formula_families["working-capital"].add("inventory_component_sum")
        assets = wc_rows.get("流动资产小计")
        if assets:
            refs = [wc_layout["ws"].cell(wc_rows[x], amount_col).coordinate for x in ("应收账款", "存货", "现金") if x in wc_rows]
            wc_layout["ws"].cell(assets, amount_col, "=SUM(" + ",".join(refs) + ")")
            formula_families["working-capital"].add("working_capital_total")
        total_wc = wc_rows.get("新增流动资金合计")
        payable_row = wc_rows.get("应付账款") or wc_rows.get("减：应付账款")
        if total_wc and assets and payable_row:
            wc_layout["ws"].cell(total_wc, amount_col,
                                  f"={wc_layout['ws'].cell(assets, amount_col).coordinate}"
                                  f"-{wc_layout['ws'].cell(payable_row, amount_col).coordinate}")
            formula_families["working-capital"].add("working_capital_total")

    # 附表6-1 / 6-3：工资合计、摊销额由表内组成复算。
    wage = layouts.get("wage") or {}
    for row in range(wage.get("first", 4), wage.get("last", 3) + 1):
        w, welfare, total = col("wage", "wage"), col("wage", "welfare"), col("wage", "total")
        if all((w, welfare, total)):
            wage["ws"].cell(row, total, f"={wage['ws'].cell(row, w).coordinate}+{wage['ws'].cell(row, welfare).coordinate}")
            formula_families["wage"].add("wage_total")
    amort = layouts.get("amortization") or {}
    for row in range(amort.get("first", 4), amort.get("last", 3) + 1):
        base_c, years_c, amount_c = col("amortization", "base"), col("amortization", "amort_years"), col("amortization", "amortization")
        if all((base_c, years_c, amount_c)):
            base = amort["ws"].cell(row, base_c).coordinate
            years = amort["ws"].cell(row, years_c).coordinate
            year_no = amort["ws"].cell(row, col("amortization", "year")).coordinate
            amort["ws"].cell(row, amount_c, f"=IF({year_no}<={years},{base}/{years},0)")
            formula_families["amortization"].add("amortization")

    # 附表6-1：人员类别明细必须由人数×人均工资复算，不能只留工资/福利合计。
    staff_rows = (pack.get("wage") or {}).get("staff_detail") or []
    wage_detail_rows: list[int] = []
    if isinstance(staff_rows, list) and staff_rows:
        ws = wage["ws"]
        start = ws.max_row + 2
        ws.cell(start, 1, "劳动定员与工资复算")
        headers = ["人员类别", "人数", "人均年工资(元)", "工资额(万元)", "福利率", "福利费(万元)", "合计(万元)"]
        for c_idx, label in enumerate(headers, start=1):
            ws.cell(start + 1, c_idx, label).font = Font(bold=True)
        wage_table = pack.get("wage") or {}
        wage_rows = wage_table.get("engine_rows") or wage_table.get("rows") or []
        wage_keys = [
            str(column.get("key") or "")
            for column in (wage_table.get("engine_columns") or wage_table.get("columns") or [])
            if isinstance(column, dict)
        ]
        wage_rate = 0.0
        if wage_rows and "wage" in wage_keys and "welfare" in wage_keys:
            base_wage = float(wage_rows[0][wage_keys.index("wage")] or 0.0)
            welfare_value = float(wage_rows[0][wage_keys.index("welfare")] or 0.0)
            wage_rate = welfare_value / base_wage if base_wage else 0.0
        for fact in staff_rows:
            if not isinstance(fact, dict):
                continue
            row_idx = ws.max_row + 1
            headcount = fact.get("headcount") if fact.get("headcount") is not None else fact.get("人数")
            avg_wage = fact.get("avg_wage_yuan") if fact.get("avg_wage_yuan") is not None else fact.get("人均年工资")
            ws.append([
                fact.get("category") or fact.get("name"), headcount, avg_wage,
                None, wage_rate, None, None,
            ])
            if headcount not in (None, "") and avg_wage not in (None, ""):
                wage_formula = f"=B{row_idx}*C{row_idx}/10000"
                ws.cell(row_idx, 4, wage_formula)
                ws.cell(row_idx, 6, f"=D{row_idx}*E{row_idx}")
                ws.cell(row_idx, 7, f"=D{row_idx}+F{row_idx}")
                wage_detail_rows.append(row_idx)
                formula_families["wage"].update({"staff_wage", "wage_total"})
                lineage.append({
                    "target": f"{ws.title}!D{row_idx}",
                    "source": f"{ws.title}!B{row_idx}*C{row_idx}",
                    "method": "excel_formula",
                })
        if wage_detail_rows:
            wage_refs = ",".join(f"D{row}" for row in wage_detail_rows)
            welfare_refs = ",".join(f"F{row}" for row in wage_detail_rows)
            total_refs = ",".join(f"G{row}" for row in wage_detail_rows)
            for main_row in range(wage.get("first", 4), wage.get("last", 3) + 1):
                set_formula("wage", main_row, "wage", f"=SUM({wage_refs})", wage_refs, family="staff_wage")
                set_formula("wage", main_row, "welfare", f"=SUM({welfare_refs})", welfare_refs, family="wage_total")
                set_formula("wage", main_row, "total", f"=SUM({total_refs})", total_refs, family="wage_total")

    # 附表6-2：分类资产基础块显示原值/年限/残值率和独立年折旧。
    asset_classes = (pack.get("depreciation") or {}).get("asset_classes") or []
    asset_detail_rows: list[int] = []
    if isinstance(asset_classes, list) and asset_classes:
        ws = dep["ws"]
        start = ws.max_row + 2
        ws.cell(start, 1, "资产类别折旧基础")
        headers = ["资产类别", "原值(万元)", "残值率", "折旧年限", "年折旧(万元)", "首年累计折旧", "首年净值"]
        for c_idx, label in enumerate(headers, start=1):
            ws.cell(start + 1, c_idx, label).font = Font(bold=True)
        for fact in asset_classes:
            if not isinstance(fact, dict):
                continue
            original = next((fact.get(field) for field in (
                "original_value_wan", "original_wan", "original_value", "amount_wan",
            ) if fact.get(field) is not None), None)
            years = next((fact.get(field) for field in (
                "depreciation_years", "dep_years", "years", "life",
            ) if fact.get(field) is not None), None)
            salvage = fact.get("salvage_rate")
            row_idx = ws.max_row + 1
            ws.append([fact.get("name") or fact.get("label"), original, salvage, years, None, None, None])
            if original not in (None, "") and years not in (None, "") and salvage is not None:
                ws.cell(row_idx, 5, f"=B{row_idx}*(1-C{row_idx})/D{row_idx}")
                ws.cell(row_idx, 6, f"=E{row_idx}")
                ws.cell(row_idx, 7, f"=B{row_idx}-F{row_idx}")
                asset_detail_rows.append(row_idx)
                formula_families["depreciation"].add("asset_class_depreciation")
                lineage.append({
                    "target": f"{ws.title}!E{row_idx}",
                    "source": f"{ws.title}!B{row_idx}*(1-C{row_idx})/D{row_idx}",
                    "method": "excel_formula",
                })

    # 附表8：利息与期末余额由期初余额、利率和还本计划滚动。
    debt = layouts.get("debt-service") or {}
    profit_layout = layouts.get("profit-distribution") or {}
    for offset, row in enumerate(range(debt.get("first", 4), debt.get("last", 3) + 1)):
        begin_c, rate_c = col("debt-service", "begin"), col("debt-service", "rate")
        principal_c, interest_c, end_c = col("debt-service", "principal"), col("debt-service", "interest"), col("debt-service", "end")
        if all((begin_c, rate_c, principal_c, interest_c, end_c)):
            if offset > 0:
                debt["ws"].cell(row, begin_c, f"={debt['ws'].cell(row-1, end_c).coordinate}")
            begin = debt["ws"].cell(row, begin_c).coordinate
            rate = debt["ws"].cell(row, rate_c).coordinate
            principal = debt["ws"].cell(row, principal_c).coordinate
            debt["ws"].cell(row, interest_c, f"={begin}*{rate}")
            debt["ws"].cell(row, end_c, f"=MAX({begin}-{principal},0)")
            formula_families["debt-service"].add("debt_rollforward")
            debt_service_c = col("debt-service", "debt_service")
            if debt_service_c:
                debt["ws"].cell(
                    row, debt_service_c,
                    f"={debt['ws'].cell(row, principal_c).coordinate}+{debt['ws'].cell(row, interest_c).coordinate}",
                )
                formula_families["debt-service"].add("repayment_coverage")
            source_cols = [
                col("debt-service", field)
                for field in ("repay_source_profit", "repay_source_dep", "repay_source_amort")
            ]
            dscr_c = col("debt-service", "dscr")
            if dscr_c and debt_service_c and all(source_cols):
                source_refs = [debt["ws"].cell(row, c).coordinate for c in source_cols]
                due_ref = debt["ws"].cell(row, debt_service_c).coordinate
                debt["ws"].cell(
                    row, dscr_c,
                    f"=IF({due_ref}>0,SUM({','.join(source_refs)})/{due_ref},\"\")",
                )
                formula_families["debt-service"].add("repayment_coverage")
            icr_c = col("debt-service", "icr")
            ebit_c = col("profit-distribution", "ebit")
            profit_row = profit_layout.get("first", 4) + offset
            if (
                icr_c and ebit_c and profit_row <= profit_layout.get("last", 3)
                and interest_c
            ):
                ebit_ref = f"'附表7'!{profit_layout['ws'].cell(profit_row, ebit_c).coordinate}"
                interest_ref = debt["ws"].cell(row, interest_c).coordinate
                debt["ws"].cell(row, icr_c, f"=IF({interest_ref}>0,{ebit_ref}/{interest_ref},\"\")")
                lineage.append({
                    "target": f"附表8!{debt['ws'].cell(row, icr_c).coordinate}",
                    "source": ebit_ref,
                    "method": "excel_formula",
                })
                formula_families["debt-service"].add("repayment_coverage")

    # Cross-sheet formulas: profit table and cash-flow tables consume their upstream schedules.
    inc = layouts.get("income-statement") or {}
    cost = layouts.get("total-cost") or {}
    profit = layouts.get("profit-distribution") or {}
    for offset, row in enumerate(range(profit.get("first", 4), profit.get("last", 3) + 1)):
        src_inc = inc.get("first", 4) + offset
        src_cost = cost.get("first", 4) + offset
        for field, src_key, src_field in (
            ("revenue", "income-statement", "revenue"),
            ("tax_surtax", "income-statement", "tax_surtax"),
            ("total_cost", "total-cost", "total_cost"),
        ):
            src_row = src_inc if src_key == "income-statement" else src_cost
            src_col = col(src_key, src_field)
            if src_col:
                src_ws = layouts[src_key]["ws"]
                ref = f"'{src_ws.title}'!{src_ws.cell(src_row, src_col).coordinate}"
                set_formula(
                    "profit-distribution", row, field, f"={ref}", ref,
                    family="upstream_links",
                )
        revenue_c = col("profit-distribution", "revenue")
        tax_c = col("profit-distribution", "tax_surtax")
        total_cost_c = col("profit-distribution", "total_cost")
        total_profit_c = col("profit-distribution", "total_profit")
        if all((revenue_c, tax_c, total_cost_c, total_profit_c)):
            revenue_ref = profit["ws"].cell(row, revenue_c).coordinate
            tax_ref = profit["ws"].cell(row, tax_c).coordinate
            cost_ref = profit["ws"].cell(row, total_cost_c).coordinate
            set_formula(
                "profit-distribution", row, "total_profit",
                f"={revenue_ref}-{cost_ref}-{tax_ref}",
                f"{revenue_ref}-{cost_ref}-{tax_ref}", family="profit_distribution",
            )
        profit_ref = profit["ws"].cell(row, total_profit_c).coordinate if total_profit_c else ""
        interest_c = col("total-cost", "interest")
        dep_c = col("total-cost", "depreciation")
        amort_c = col("total-cost", "amortization")
        if profit_ref and interest_c:
            interest_ref = f"'附表6'!{cost['ws'].cell(src_cost, interest_c).coordinate}"
            set_formula(
                "profit-distribution", row, "ebit", f"={profit_ref}+{interest_ref}",
                interest_ref, family="profit_distribution",
            )
        ebit_c = col("profit-distribution", "ebit")
        if ebit_c and dep_c and amort_c:
            ebit_ref = profit["ws"].cell(row, ebit_c).coordinate
            dep_ref = f"'附表6'!{cost['ws'].cell(src_cost, dep_c).coordinate}"
            amort_ref = f"'附表6'!{cost['ws'].cell(src_cost, amort_c).coordinate}"
            set_formula(
                "profit-distribution", row, "ebitda",
                f"={ebit_ref}+{dep_ref}+{amort_ref}",
                f"{dep_ref}+{amort_ref}", family="profit_distribution",
            )
        loss_c = col("profit-distribution", "loss_offset")
        taxable_c = col("profit-distribution", "taxable_income")
        if total_profit_c and loss_c and taxable_c:
            loss_ref = profit["ws"].cell(row, loss_c).coordinate
            set_formula(
                "profit-distribution", row, "taxable_income",
                f"=MAX({profit_ref}-{loss_ref},0)", loss_ref, family="profit_distribution",
            )
        income_tax_c = col("profit-distribution", "income_tax")
        net_c = col("profit-distribution", "net_profit")
        if total_profit_c and income_tax_c and net_c:
            tax_ref = profit["ws"].cell(row, income_tax_c).coordinate
            set_formula(
                "profit-distribution", row, "net_profit", f"={profit_ref}-{tax_ref}",
                tax_ref, family="profit_distribution",
            )
        begin_c = col("profit-distribution", "begin_undistributed")
        undistributed_c = col("profit-distribution", "undistributed")
        if begin_c:
            if offset == 0:
                profit["ws"].cell(row, begin_c, 0.0)
            elif undistributed_c:
                prior = profit["ws"].cell(row - 1, undistributed_c).coordinate
                set_formula(
                    "profit-distribution", row, "begin_undistributed", f"={prior}",
                    prior, family="profit_distribution",
                )
        available_c = col("profit-distribution", "available_distribution")
        if begin_c and net_c and available_c:
            begin_ref = profit["ws"].cell(row, begin_c).coordinate
            net_ref = profit["ws"].cell(row, net_c).coordinate
            set_formula(
                "profit-distribution", row, "available_distribution",
                f"={begin_ref}+{net_ref}", f"{begin_ref}+{net_ref}",
                family="profit_distribution",
            )
        reserve_c = col("profit-distribution", "surplus_reserve")
        if net_c and reserve_c:
            net_ref = profit["ws"].cell(row, net_c).coordinate
            set_formula(
                "profit-distribution", row, "surplus_reserve",
                f"=MAX({net_ref},0)*10%", net_ref, family="profit_distribution",
            )
        distributable_c = col("profit-distribution", "distributable")
        if available_c and reserve_c and distributable_c:
            available_ref = profit["ws"].cell(row, available_c).coordinate
            reserve_ref = profit["ws"].cell(row, reserve_c).coordinate
            set_formula(
                "profit-distribution", row, "distributable",
                f"={available_ref}-{reserve_ref}", f"{available_ref}-{reserve_ref}",
                family="profit_distribution",
            )
        arbitrary_c = col("profit-distribution", "arbitrary_reserve")
        investor_c = col("profit-distribution", "investor_distribution")
        if distributable_c and arbitrary_c and investor_c and undistributed_c:
            distributable_ref = profit["ws"].cell(row, distributable_c).coordinate
            arbitrary_ref = profit["ws"].cell(row, arbitrary_c).coordinate
            investor_ref = profit["ws"].cell(row, investor_c).coordinate
            set_formula(
                "profit-distribution", row, "undistributed",
                f"={distributable_ref}-N({arbitrary_ref})-N({investor_ref})",
                f"{distributable_ref}-{arbitrary_ref}-{investor_ref}",
                family="profit_distribution",
            )

    # 附表4：总投资/资金筹措合计引用附表1项目总投资。
    inv_total_row = row_by_label("investment", "项目总投资合计")
    inv_total_col = col("investment", "total")
    if inv_total_row and inv_total_col:
        inv_ref = f"'附表1'!{layouts['investment']['ws'].cell(inv_total_row, inv_total_col).coordinate}"
        for label in ("项目总投资使用计划", "资金筹措"):
            target_row = row_by_label("funding", label)
            if target_row:
                set_formula(
                    "funding", target_row, "amount", f"={inv_ref}", inv_ref,
                    family="investment_link",
                )

    cash = layouts.get("cashflow") or {}
    for offset, row in enumerate(range(cash.get("first", 4), cash.get("last", 3) + 1)):
        phase_col = col("cashflow", "phase")
        phase = cash["ws"].cell(row, phase_col).value if phase_col else ""
        if phase == "建设期":
            inv_rows = [row_by_label("investment", x) for x in ("工程费用", "工程建设其他费用", "预备费", "建设期利息")]
            inv_total_c = col("investment", "total")
            refs = [f"'附表1'!{layouts['investment']['ws'].cell(r, inv_total_c).coordinate}" for r in inv_rows if r and inv_total_c]
            if refs:
                share = float(cash["ws"].cell(row, col("cashflow", "construction")).value or 0.0)
                cash_table = pack.get("cashflow") or {}
                cash_rows = cash_table.get("engine_rows") or cash_table.get("rows") or []
                cash_columns = cash_table.get("engine_columns") or cash_table.get("columns") or []
                cash_keys = [str(column.get("key") or "") for column in cash_columns if isinstance(column, dict)]
                construction_index = cash_keys.index("construction") if "construction" in cash_keys else -1
                phase_index = cash_keys.index("phase") if "phase" in cash_keys else -1
                denom = sum(
                    float(item[construction_index] or 0.0)
                    for item in cash_rows
                    if construction_index >= 0 and phase_index >= 0
                    and len(item) > max(construction_index, phase_index)
                    and str(item[phase_index]) == "建设期"
                ) or share or 1.0
                set_formula("cashflow", row, "construction",
                            "=SUM(" + ",".join(refs) + f")*{share/denom:g}", "+".join(refs),
                            family="upstream_links")
            continue
        if phase != "运营期":
            continue
        op_idx = offset - max(int((fin.get("params") or {}).get("build_years") or 0), 0)
        src_row = inc.get("first", 4) + max(op_idx, 0)
        for field in ("revenue", "tax_surtax", "income_tax"):
            src_col = col("income-statement", field)
            if src_col and src_row <= inc.get("last", 3):
                src_ws = inc["ws"]
                ref = f"'{src_ws.title}'!{src_ws.cell(src_row, src_col).coordinate}"
                set_formula("cashflow", row, field, f"={ref}", ref, family="upstream_links")
        cost_row = cost.get("first", 4) + max(op_idx, 0)
        cost_col = col("total-cost", "operating_cost")
        if cost_col and cost_row <= cost.get("last", 3):
            ref = f"'附表6'!{cost['ws'].cell(cost_row, cost_col).coordinate}"
            set_formula("cashflow", row, "op_cash_cost", f"={ref}", ref, family="upstream_links")
        wc_total_row = row_by_label("working-capital", "新增流动资金合计", field="item")
        if not wc_total_row:
            wc_total_row = row_by_label("working-capital", "流动资金", field="item")
        if not wc_total_row:
            wc_total_row = row_by_label("working-capital", "流动资金总额（汇总输入）", field="item")
        wc_amount_col = col("working-capital", "amount")
        if op_idx == 0 and wc_total_row and wc_amount_col:
            ref = f"'附表3'!{layouts['working-capital']['ws'].cell(wc_total_row, wc_amount_col).coordinate}"
            set_formula("cashflow", row, "wc_change", f"={ref}", ref, family="upstream_links")
        parts = [
            ("revenue", 1), ("op_cash_cost", -1), ("tax_surtax", -1),
            ("income_tax", -1), ("construction", -1), ("wc_change", -1), ("recover", 1),
        ]
        terms = []
        for field, sign in parts:
            c = col("cashflow", field)
            if c:
                coord = cash["ws"].cell(row, c).coordinate
                terms.append(("+" if sign > 0 else "-") + coord)
        if terms:
            set_formula(
                "cashflow", row, "net_cashflow", "=" + "".join(terms).lstrip("+"),
                "cashflow components", family="net_cashflow",
            )
            cum_col = col("cashflow", "cumulative")
            net_col = col("cashflow", "net_cashflow")
            if cum_col and net_col:
                net_ref = cash["ws"].cell(row, net_col).coordinate
                formula = f"={net_ref}" if row == cash.get("first", 4) else f"={cash['ws'].cell(row-1, cum_col).coordinate}+{net_ref}"
                cash["ws"].cell(row, cum_col, formula)
                formula_families["cashflow"].add("net_cashflow")

    # 附表10：还本付息直接引用附表8债务计划。
    capital = layouts.get("capital-cashflow") or {}
    debt = layouts.get("debt-service") or {}
    for offset, row in enumerate(range(capital.get("first", 4), capital.get("last", 3) + 1)):
        phase_col = col("capital-cashflow", "phase")
        is_operating_row = bool(
            phase_col and capital["ws"].cell(row, phase_col).value == "运营期"
        )
        build_years = max(int((fin.get("params") or {}).get("build_years") or 0), 0)
        op_idx = offset - build_years
        if is_operating_row:
            debt_row = debt.get("first", 4) + max(op_idx, 0)
            for field in ("principal", "interest"):
                debt_col = col("debt-service", field)
                if debt_col and debt_row <= debt.get("last", 3):
                    ref = f"'附表8'!{debt['ws'].cell(debt_row, debt_col).coordinate}"
                    set_formula(
                        "capital-cashflow", row, field, f"={ref}", ref,
                        family="debt_links",
                    )
        inflow_fields = ("revenue", "recover_fixed", "recover_wc")
        outflow_fields = (
            "capital_invest", "op_cash_cost", "tax_surtax", "income_tax",
            "principal", "interest",
        )
        inflow_col = col("capital-cashflow", "cash_inflow")
        outflow_col = col("capital-cashflow", "cash_outflow")
        inflow_parts = [col("capital-cashflow", field) for field in inflow_fields]
        outflow_parts = [col("capital-cashflow", field) for field in outflow_fields]
        if inflow_col and outflow_col and all(inflow_parts) and all(outflow_parts):
            inflow_refs = [capital["ws"].cell(row, c).coordinate for c in inflow_parts]
            outflow_refs = [capital["ws"].cell(row, c).coordinate for c in outflow_parts]
            capital["ws"].cell(row, inflow_col, "=SUM(" + ",".join(inflow_refs) + ")")
            capital["ws"].cell(row, outflow_col, "=SUM(" + ",".join(outflow_refs) + ")")
            formula_families["capital-cashflow"].add("capital_cashflow_atomic")
            formula = (
                f"={capital['ws'].cell(row, inflow_col).coordinate}"
                f"-{capital['ws'].cell(row, outflow_col).coordinate}"
            )
            set_formula(
                "capital-cashflow", row, "net_cashflow", formula,
                "atomic capital cashflow inflow-outflow", family="net_cashflow",
            )

    def append_cashflow_tree(key: str, components: list[tuple[str, list[str]]]) -> None:
        layout = layouts[key]
        ws = layout["ws"]
        start = ws.max_row + 2
        ws.cell(start, 1, "现金流量组成树")
        header = start + 1
        ws.cell(header, 1, "项目")
        main_rows = list(range(layout.get("first", 4), layout.get("last", 3) + 1))
        for idx, main_row in enumerate(main_rows, start=2):
            ws.cell(header, idx, ws.cell(main_row, col(key, "year")).value)
        cursor = header + 1
        for group, fields in components:
            ws.cell(cursor, 1, group)
            cursor += 1
            for field in fields:
                label = next((str(c.get("label")) for c in pack[key].get("columns") or [] if c.get("key") == field), field)
                ws.cell(cursor, 1, f"  {label}")
                source_col = col(key, field)
                if source_col:
                    for idx, main_row in enumerate(main_rows, start=2):
                        ref = f"'{ws.title}'!{ws.cell(main_row, source_col).coordinate}"
                        ws.cell(cursor, idx, f"={ref}")
                        lineage.append({"target": f"{ws.title}!{ws.cell(cursor, idx).coordinate}",
                                        "source": ref, "method": "excel_formula"})
                cursor += 1

    append_cashflow_tree("cashflow", [
        ("现金流入", ["revenue", "recover"]),
        ("现金流出", ["construction", "wc_change", "op_cash_cost", "tax_surtax", "income_tax"]),
        ("现金净额", ["net_cashflow", "cumulative"]),
    ])
    append_cashflow_tree("capital-cashflow", [
        ("现金流入", ["revenue", "recover_fixed", "recover_wc"]),
        ("现金流出", ["capital_invest", "op_cash_cost", "tax_surtax", "income_tax", "principal", "interest"]),
        ("现金净额", ["net_cashflow"]),
    ])

    # 主表是甲方参考行树；逐年计算区只承担可追溯计算。将可直接映射的
    # 参考行用公式回链计算区，避免“上面静态展示、下面才有公式”的假深度。
    for key, sheet_name in _DELIVERY_SHEETS.items():
        table = pack.get(key) or {}
        fields = list(table.get("reference_row_fields") or [])
        if not fields or not table.get("engine_rows"):
            continue
        layout = layouts.get(key) or {}
        ws = layout.get("ws")
        reference = layout.get("reference") or {}
        reference_keys = list(layout.get("reference_keys") or [])
        total_modes = list(table.get("reference_total_modes") or [])
        if ws is None:
            continue
        period_keys = [item for item in reference_keys if item.startswith("period_")]
        total_column = reference_keys.index("total") + 1 if "total" in reference_keys else None
        for row_offset, field in enumerate(fields):
            if not field or field not in (layout.get("keys") or []):
                continue
            engine_column = (layout.get("keys") or []).index(field) + 1
            target_row = int(reference.get("first") or 4) + row_offset
            linked_cells: list[str] = []
            for period_offset, period_key in enumerate(period_keys):
                source_row = int(layout.get("first") or 4) + period_offset
                if source_row > int(layout.get("last") or 3):
                    break
                target_column = reference_keys.index(period_key) + 1
                source_ref = ws.cell(source_row, engine_column).coordinate
                target_cell = ws.cell(target_row, target_column, f"={source_ref}")
                target_cell.number_format = '#,##0.00'
                linked_cells.append(target_cell.coordinate)
                lineage.append({
                    "target": f"{sheet_name}!{target_cell.coordinate}",
                    "source": f"{sheet_name}!{source_ref}",
                    "method": "excel_formula",
                })
                formula_families[key].add("reference_projection")
            if total_column and linked_cells:
                mode = total_modes[row_offset] if row_offset < len(total_modes) else "sum"
                formula = f"={linked_cells[-1]}" if mode == "last" else "=SUM(" + ",".join(linked_cells) + ")"
                ws.cell(target_row, total_column, formula).number_format = '#,##0.00'

    # 按冻结 reference schema 生成跨表依赖复核块。真实工作簿有 15 张参考附表，
    # 引擎交付合并为 13 张，因此先把 supporting sheet 映射到承载它的引擎表。
    from lvke_mcp.domains.finance.reference_schema import load_reference_table_schema

    reference_contract = load_reference_table_schema()
    mapping = ((reference_contract.get("machine_contract") or {}).get("engine_reference_mapping") or {})
    reference_to_delivery = {
        str(item.get("reference_sheet") or ""): _DELIVERY_SHEETS.get(key, "")
        for key, item in mapping.items()
    }
    reference_to_delivery.update({
        "附表6-1": "附表6",
        "附表6-2": "附表6",
        "附表6-3": "附表6-1",
        "附表6-5": "附表6-2",
        "附表6-6": "附表6-3",
    })
    contract_expected_edges: set[tuple[str, str]] = set()

    def dependency_source_cell(source_key: str) -> Optional[str]:
        source_layout = layouts.get(source_key) or {}
        source_ws = source_layout.get("ws")
        if source_ws is None:
            return None
        for row in range(int(source_layout.get("first") or 4), int(source_layout.get("last") or 3) + 1):
            for column in range(1, len(source_layout.get("keys") or []) + 1):
                value = source_ws.cell(row, column).value
                if isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith("=")):
                    return source_ws.cell(row, column).coordinate
        return None

    delivery_to_key = {sheet: key for key, sheet in _DELIVERY_SHEETS.items()}
    for target_key, target_sheet in _DELIVERY_SHEETS.items():
        dependencies = list((mapping.get(target_key) or {}).get("formula_dependency_sources") or [])
        normalized_sources = []
        for reference_sheet in dependencies:
            source_sheet = reference_to_delivery.get(str(reference_sheet) or "")
            if source_sheet and source_sheet != target_sheet and source_sheet not in normalized_sources:
                normalized_sources.append(source_sheet)
        target_ws = layouts[target_key]["ws"]
        for source_sheet in normalized_sources:
            contract_expected_edges.add((target_sheet, source_sheet))
            source_key = delivery_to_key.get(source_sheet)
            source_cell = dependency_source_cell(source_key or "")
            if not source_key or not source_cell:
                continue
            target_row = target_ws.max_row + 1
            target_ws.cell(target_row, 1, f"跨表依赖复核：{source_sheet}")
            dep_cell = target_ws.cell(target_row, 2, f"='{source_sheet}'!{source_cell}")
            dep_cell.number_format = '#,##0.00'
            lineage.append({
                "target": f"{target_sheet}!{dep_cell.coordinate}",
                "source": f"'{source_sheet}'!{source_cell}",
                "method": "excel_formula",
                "family": "contract_dependency",
            })
            formula_families[target_key].add("contract_dependency")

    # Deep delivery checks are intentionally stricter than engine consistency.
    # A workbook is not formal-ready merely because 13 sheets contain numbers.
    def has_label(sheet: str, labels: set[str]) -> bool:
        ws = wb[sheet]
        return any(str(cell.value).strip() in labels for row in ws.iter_rows() for cell in row if cell.value is not None)

    def _is_contract_dependency_formula(value: Any, *, sheet: str = "", coordinate: str = "") -> bool:
        text = str(value or "")
        if not text.startswith("="):
            return False
        if "跨表依赖复核" in text:
            return True
        target = f"{sheet}!{coordinate}" if sheet and coordinate else ""
        if target and any(
            str(row.get("family") or "") == "contract_dependency"
            and str(row.get("target") or "") == target
            for row in lineage
        ):
            return True
        return False

    def _formula_cells(sheet: str) -> list[tuple[str, Any]]:
        ws = wb[sheet]
        return [
            (cell.coordinate, cell.value)
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]

    def _non_dependency_formula_count(sheet: str) -> int:
        return sum(
            1
            for coordinate, value in _formula_cells(sheet)
            if not _is_contract_dependency_formula(value, sheet=sheet, coordinate=coordinate)
        )

    investment_rows = pack.get("investment", {}).get("rows") or []
    investment_cols = [str(c.get("key") or "") for c in pack.get("investment", {}).get("columns") or []]
    qty_idx = investment_cols.index("quantity") if "quantity" in investment_cols else -1
    indicator_idx = investment_cols.index("indicator") if "indicator" in investment_cols else -1
    detail_rows = [r for r in investment_rows if str(r[0] if r else "").startswith("1.")]
    quantity_indicator_ok = bool(detail_rows) and all(
        len(r) > max(qty_idx, indicator_idx) and r[qty_idx] not in (None, "") and r[indicator_idx] not in (None, "")
        for r in detail_rows
    )
    cost_items = (fin.get("input_revision") or fin.get("finance_inputs") or {}).get("cost_items") or {}
    wc = (fin.get("annual") or {}).get("working_capital") or {}
    wc_stated = wc.get("stated_total")
    wc_delta = wc.get("delta_vs_stated")
    wc_tolerance = max(1.0, abs(float(wc_stated or 0.0)) * 0.01)
    inv_wc = float((fin.get("investment") or {}).get("working_capital") or 0.0)
    # ``wc_actionable`` records why the gate closed, so the blocker can name the
    # missing input instead of only the check name.  The pass/fail rules below are
    # unchanged.
    wc_actionable: str | None = None
    if inv_wc <= 0.01:
        wc_reconciled = True  # no WC to reconcile
    elif wc.get("method") == "turnover_days":
        # Honesty rules:
        # - force-scale is never formal-ok
        # - no stated total → turnover structure is the authority
        # - stated total present → require small delta (no silent scale)
        scaled = bool(wc.get("scaled_to_stated_total"))
        if scaled:
            wc_reconciled = False
            wc_actionable = (
                "附表3流动资金被 scaled_to_stated_total 强制缩放到附表1声明值，"
                "正式交付不接受强制缩放。调用 finance_run_model 时修正 working_capital "
                "的周转天数输入（应收账款/应付账款/存货/现金），让周转法自身算出的总额"
                "与附表1一致，而不是事后缩放。"
            )
        elif wc_stated is None:
            wc_reconciled = True
        else:
            wc_reconciled = (
                wc_delta is None
                or abs(float(wc_delta)) <= max(wc_tolerance, 1.0)
            )
            if not wc_reconciled:
                wc_actionable = (
                    f"附表3周转法流动资金 {wc.get('total')} 与附表1声明值 {wc_stated} "
                    f"相差 {wc_delta}，超过容差 {round(max(wc_tolerance, 1.0), 2)}。"
                    "调用 finance_run_model 时二者选一改到一致：修正 working_capital "
                    "的周转天数输入，或修正投资估算里的流动资金声明值。"
                )
    else:
        wc_reconciled = False
        wc_actionable = (
            f"附表3流动资金计算方法为 {wc.get('method') or 'unset'}，"
            "正式交付只接受 turnover_days（周转天数法），因为只有它能逐项复算。"
            "调用 finance_run_model 时把 working_capital.method 设为 turnover_days "
            "并提供各项周转天数。"
        )
    product_tree_ok = bool(pack.get("income-statement", {}).get("product_tree"))
    delivery_formula_count = {
        sheet: _non_dependency_formula_count(sheet) for sheet in _DELIVERY_SHEETS.values()
    }
    dependency_only_formula_count = {
        sheet: sum(
            1
            for coordinate, value in _formula_cells(sheet)
            if _is_contract_dependency_formula(value, sheet=sheet, coordinate=coordinate)
        )
        for sheet in _DELIVERY_SHEETS.values()
    }
    idc_rows = (fin.get("annual") or {}).get("interest_during_construction") or []
    idc_formula_total = round(sum(
        float(r.get("interest") or 0.0)
        for r in idc_rows if isinstance(r, dict)
    ), 2)
    stated_interest = round(float((fin.get("investment") or {}).get("interest") or 0.0), 2)

    # Independent arithmetic checks use the structured values before Excel formulas
    # replace cells.  Formula presence alone is not proof that the supplied facts
    # and deterministic result agree.
    investment_recalc_ok = bool(detail_rows)
    investment_recalc_deltas: list[float] = []
    total_idx = investment_cols.index("total") if "total" in investment_cols else -1
    if quantity_indicator_ok and total_idx >= 0:
        for row in detail_rows:
            expected = float(row[qty_idx]) * float(row[indicator_idx]) / 10000.0
            actual = float(row[total_idx] or 0.0)
            delta = round(actual - expected, 6)
            investment_recalc_deltas.append(delta)
            if abs(delta) > 0.01:
                investment_recalc_ok = False
    else:
        investment_recalc_ok = False
    actual_edges = set()
    for row in lineage:
        if row.get("method") != "excel_formula":
            continue
        source = str(row.get("source") or "")
        for sheet in _DELIVERY_SHEETS.values():
            if f"'{sheet}'!" in source:
                target = str(row.get("target") or "").split("!", 1)[0]
                actual_edges.add((target, sheet))
    expected_edges = contract_expected_edges
    # 附表 → 对应 pack 输入键。缺公式时用它把 blocker 指向该补的那份输入。
    supporting_schedule_inputs = {
        "附表2": "interest-during-construction",
        "附表3": "working-capital",
        "附表6-1": "wage",
        "附表6-3": "amortization",
        "附表8": "debt-service",
    }
    supporting_schedule_gaps = [
        (sheet, key)
        for sheet, key in supporting_schedule_inputs.items()
        if delivery_formula_count.get(sheet, 0) == 0
        and (pack.get(key) or {}).get("grade") != "not_applicable"
    ]
    semantic_checks = {
        "all_tables_reference_grade": {
            "ok": all(
                (pack.get(key) or {}).get("grade") in {"reference", "not_applicable"}
                for key in _DELIVERY_SHEETS
            ),
            "detail": "正式交付要求适用附表均达 reference 级；not_applicable 不计失败",
            "non_reference": [
                _DELIVERY_SHEETS[key]
                for key in _DELIVERY_SHEETS
                if (pack.get(key) or {}).get("grade") not in {"reference", "not_applicable"}
            ],
        },
        "investment_quantity_indicator": {
            "ok": quantity_indicator_ok,
            "detail": "附表1工程量与估算指标必须有输入值，不能只填金额",
            "actionable": (
                "附表1明细行（编号 1.x）的 quantity（工程量）和 indicator（估算指标）"
                "列必须有值，不能只填 total（金额）。调用 finance_run_model 时在 "
                "quantity_items 输入中补全每项的工程量与单位指标。"
            ) if not quantity_indicator_ok else None,
        },
        "funding_year_plan": {
            "ok": "分年" in [str(x) for x in (pack.get("funding", {}).get("column_labels") or [])]
                or any(str(x).startswith("建设期第") for x in (pack.get("funding", {}).get("column_labels") or [])),
            "detail": "附表4必须展示建设期分年使用计划",
        },
        "construction_interest_reconciled": {
            "ok": bool(idc_rows) and abs(idc_formula_total - stated_interest) <= 0.01,
            "detail": (
                "附表2不可变分年计息计划必须与附表1一致"
                f"（formula={idc_formula_total}, stated={stated_interest}）"
            ),
        },
        "income_product_tree": {"ok": product_tree_ok, "detail": "附表5必须有分产品量价与爬坡块"},
        "working_capital_reconciled": {
            "ok": wc_reconciled,
            "detail": (
                "附表3周转法流动资金必须与附表1投资流动资金对齐"
                f"（stated={wc_stated}, turnover={wc.get('total')}, delta={wc_delta}）"
            ),
            "actionable": wc_actionable,
        },
        "income_formula_driven": {
            "ok": product_tree_ok and delivery_formula_count.get("附表5", 0) > 0,
            "detail": "附表5收入/税费不能全部固化为引擎值，至少应有产品量价或跨表公式",
        },
        "cost_item_tree": {
            "ok": bool(cost_items) and has_label("附表6", set(map(str, cost_items.keys()))),
            "detail": "附表6必须展示成本明细树",
        },
        "depreciation_rollforward": {
            "ok": has_label("附表6-2", {"当期折旧费"}) and has_label("附表6-2", {"净值"}),
            "detail": "附表6-2必须有累计折旧与期末净值",
        },
        "supporting_schedules_formula_driven": {
            "ok": not supporting_schedule_gaps,
            "detail": "适用的附表2/3/6-1/6-3/8必须具备表内复算公式；not_applicable 跳过",
            "actionable": (
                "以下附表缺少表内非依赖公式且未标记 not_applicable："
                + "、".join(
                    f"{sheet}（对应输入 {key}）"
                    for sheet, key in supporting_schedule_gaps
                )
                + "。调用 finance_run_model 时补全对应输入的明细结构（如建设期借款分年"
                "计息、工资明细、还本付息计划），或把确实不适用的附表标记为 not_applicable。"
            ) if supporting_schedule_gaps else None,
            "missing_schedules": [sheet for sheet, _ in supporting_schedule_gaps],
        },
        "cashflow_row_tree": {
            "ok": has_label("附表9", {"现金流入", "现金流出"}) and has_label("附表10", {"现金流入", "现金流出"}),
            "detail": "附表9/10必须分别展示现金流入、现金流出组成树",
        },
        "cross_sheet_dependencies": {
            "ok": expected_edges.issubset(actual_edges),
            "detail": "关键附表之间必须有可追踪跨 sheet 公式依赖",
            "missing": sorted(expected_edges - actual_edges),
        },
    }
    pack_meta = pack.get("_meta") or {}
    applicable_keys = [
        key for key in _DELIVERY_SHEETS
        if (pack.get(key) or {}).get("applicable", True)
    ]
    # Formula coverage is declared by the frozen schema.  Projection/dependency
    # links alone never satisfy a business formula family.
    formula_coverage_by_table: dict[str, float] = {}
    required_formula_families: dict[str, list[str]] = {}
    actual_formula_families: dict[str, list[str]] = {}
    for key in applicable_keys:
        sheet = _DELIVERY_SHEETS[key]
        contract = (reference_contract.get("tables") or {}).get(key) or {}
        required = list(contract.get("required_formula_families") or _REQUIRED_FORMULA_FAMILIES.get(key, ()))
        families = {
            name for name in (formula_families.get(key) or set())
            if name not in {"contract_dependency", "reference_projection"}
        }
        required_formula_families[key] = required
        actual_formula_families[key] = sorted(families)
        passed = sum(1 for family in required if family in families)
        formula_coverage_by_table[key] = round(passed / len(required), 4) if required else 1.0
    formula_coverage = (
        round(sum(formula_coverage_by_table.values()) / len(formula_coverage_by_table), 4)
        if formula_coverage_by_table else 1.0
    )

    # Cashflow composition independent checks (catch 附表9/10 mapping bugs).
    cashflow_pack = pack.get("cashflow") or {}
    capital_pack = pack.get("capital-cashflow") or {}
    cashflow_rows = cashflow_pack.get("rows") or []
    capital_rows = capital_pack.get("rows") or []
    cashflow_labels = {
        str(row[1] if len(row) > 1 else ""): row
        for row in cashflow_rows if isinstance(row, (list, tuple)) and len(row) > 1
    }
    capital_labels = {
        str(row[1] if len(row) > 1 else ""): row
        for row in capital_rows if isinstance(row, (list, tuple)) and len(row) > 1
    }

    def _row_values(row: Any) -> list[float]:
        if not isinstance(row, (list, tuple)) or len(row) < 3:
            return []
        out: list[float] = []
        # Reference matrices are [no, item, total, period_1, ...].  The total
        # is not a year and must not shift period identities.
        for value in row[3:]:
            try:
                out.append(float(value or 0.0))
            except (TypeError, ValueError):
                out.append(0.0)
        return out

    fixed_vals = _row_values(cashflow_labels.get("回收固定资产余值"))
    wc_vals = _row_values(cashflow_labels.get("回收新增流动资金"))
    pre_tax_vals = _row_values(cashflow_labels.get("所得税前净现金流量"))
    post_tax_vals = _row_values(cashflow_labels.get("所得税后净现金流量"))
    # Fail if fixed and WC recover are identical non-zero series (classic double-map bug).
    recover_distinct = True
    if fixed_vals and wc_vals and any(abs(v) > 1e-9 for v in fixed_vals):
        if fixed_vals == wc_vals:
            recover_distinct = False
    pre_post_distinct = True
    if pre_tax_vals and post_tax_vals and any(abs(a - b) > 1e-9 for a, b in zip(pre_tax_vals, post_tax_vals)):
        pre_post_distinct = True
    elif pre_tax_vals and post_tax_vals and pre_tax_vals == post_tax_vals and any(
        abs(v) > 1e-9 for v in pre_tax_vals
    ):
        # Identical non-zero series is only acceptable when income tax is zero every year.
        tax_vals = _row_values(cashflow_labels.get("所得税"))
        pre_post_distinct = bool(tax_vals) and all(abs(v) < 1e-9 for v in tax_vals)

    cap_fixed = _row_values(capital_labels.get("回收固定资产余值"))
    cap_wc = _row_values(capital_labels.get("回收流动资金"))
    cap_tax = _row_values(capital_labels.get("所得税"))
    cap_inflow = _row_values(capital_labels.get("现金流入"))
    cap_outflow = _row_values(capital_labels.get("现金流出"))
    cap_net = _row_values(capital_labels.get("资本金净现金流量"))
    cap_revenue = _row_values(capital_labels.get("营业收入"))
    cap_cost = _row_values(capital_labels.get("经营成本"))
    cap_surtax = _row_values(capital_labels.get("税金及附加"))
    cap_capital = _row_values(capital_labels.get("项目资本金"))
    cap_principal = _row_values(capital_labels.get("借款本金偿还"))
    cap_interest = _row_values(capital_labels.get("借款利息支付"))
    effective_inputs = (
        fin.get("input_revision")
        if isinstance(fin.get("input_revision"), dict)
        else fin.get("raw")
        if isinstance(fin.get("raw"), dict)
        else {}
    )
    explicit_zero_terminal_policy = (
        "terminal_fixed_asset_recover_wan" in effective_inputs
        and "terminal_working_capital_recover_wan" in effective_inputs
        and abs(float(effective_inputs.get("terminal_fixed_asset_recover_wan") or 0.0)) <= 1e-9
        and abs(float(effective_inputs.get("terminal_working_capital_recover_wan") or 0.0)) <= 1e-9
    )
    capital_recover_present = bool(cap_fixed or cap_wc)
    if capital_recover_present:
        capital_recover_ok = (
            any(abs(v) > 1e-9 for v in (cap_fixed + cap_wc))
            or explicit_zero_terminal_policy
        )
    else:
        capital_recover_ok = (
            explicit_zero_terminal_policy
            or (pack.get("capital-cashflow") or {}).get("grade") == "not_applicable"
        )
    # 所得税 row must not equal 税金及附加+所得税 mix: compare against income_statement tax only.
    # Capital cashflow uses financing-before adjusted income tax from the
    # project cashflow rows, not the accounting tax row in the profit statement.
    project_rows_for_tax = (fin.get("annual") or {}).get("project_cashflow") or []
    income_tax_series = [
        float(row.get("income_tax") or 0.0)
        for row in project_rows_for_tax
        if isinstance(row, dict) and row.get("phase") == "运营期"
    ]
    capital_tax_ok = True
    if cap_tax and income_tax_series:
        # Align from operating years if lengths differ.
        shorter = min(len(cap_tax), len(income_tax_series))
        if shorter:
            # Compare tail (operating) when capital has build years prefix.
            cap_tail = cap_tax[-shorter:]
            tax_tail = income_tax_series[-shorter:]
            capital_tax_ok = all(abs(a - b) <= 0.05 for a, b in zip(cap_tail, tax_tail))
    capital_identity_ok = bool(cap_inflow and cap_outflow and cap_net)
    if capital_identity_ok:
        n = min(
            len(cap_inflow), len(cap_outflow), len(cap_net), len(cap_revenue),
            len(cap_fixed), len(cap_wc), len(cap_capital), len(cap_principal),
            len(cap_interest), len(cap_cost), len(cap_surtax), len(cap_tax),
        )
        capital_identity_ok = n > 0 and all(
            abs(cap_inflow[i] - (cap_revenue[i] + cap_fixed[i] + cap_wc[i])) <= 0.05
            and abs(cap_outflow[i] - (
                cap_capital[i] + cap_cost[i] + cap_surtax[i] + cap_tax[i]
                + cap_principal[i] + cap_interest[i]
            )) <= 0.05
            and abs(cap_net[i] - (cap_inflow[i] - cap_outflow[i])) <= 0.05
            for i in range(n)
        )

    independent_recalc_checks = {
        "investment_quantity_indicator": {
            "ok": investment_recalc_ok,
            "max_abs_delta_wan": max((abs(x) for x in investment_recalc_deltas), default=None),
        },
        "construction_interest": {
            "ok": bool(semantic_checks["construction_interest_reconciled"]["ok"])
            or (pack.get("interest-during-construction") or {}).get("grade") == "not_applicable",
            "delta_wan": round(idc_formula_total - stated_interest, 6),
        },
        "working_capital": {
            "ok": bool(semantic_checks["working_capital_reconciled"]["ok"])
            or (pack.get("working-capital") or {}).get("grade") == "not_applicable",
            "delta_wan": wc_delta,
            "reconcile_conflict": bool(wc.get("reconcile_conflict")),
        },
        "cashflow_recover_split": {
            "ok": (
                (pack.get("cashflow") or {}).get("grade") == "not_applicable"
                or recover_distinct
            ),
            "detail": "附表9 回收固定资产余值与回收流动资金不得同映射为完整 recover",
        },
        "cashflow_pre_post_tax": {
            "ok": (
                (pack.get("cashflow") or {}).get("grade") == "not_applicable"
                or pre_post_distinct
            ),
            "detail": "附表9 税前/税后净现金流不得无税差异时仍完全相同",
        },
        "capital_cashflow_terminal_recover": {
            "ok": capital_recover_ok or (pack.get("capital-cashflow") or {}).get("grade") == "not_applicable",
            "detail": (
                "附表10 接受来源明确的显式零回收；未显式给定时，有终值不得固定为 0"
            ),
        },
        "capital_cashflow_income_tax_only": {
            "ok": capital_tax_ok or (pack.get("capital-cashflow") or {}).get("grade") == "not_applicable",
            "detail": "附表10 所得税行不得混入税金及附加",
        },
        "capital_cashflow_composition": {
            "ok": (
                (pack.get("capital-cashflow") or {}).get("grade") == "not_applicable"
                or bool((pack.get("capital-cashflow") or {}).get("op_inflow_not_used_as_cash_inflow"))
                and bool((pack.get("capital-cashflow") or {}).get("atomic_capital_cashflow_complete"))
                and capital_identity_ok
            ),
            "detail": "附表10 不得用净额型 op_inflow 充当现金流入；须原子组成重算",
        },
        "funding_uses_sources_balance": {
            "ok": (
                (pack.get("funding") or {}).get("grade") == "not_applicable"
                or (
                    (pack.get("funding") or {}).get("funding_balance_ok") is True
                    and (pack.get("funding") or {}).get("funding_plan_source")
                    != "proportional_spread_fallback"
                )
            ),
            "detail": "附表4 资金用途与来源须分年闭合，且不得为比例摊分回退",
        },
    }
    independent_recalc_coverage = round(
        sum(1 for item in independent_recalc_checks.values() if item.get("ok"))
        / len(independent_recalc_checks),
        4,
    )
    formula_signatures: dict[str, list[dict[str, Any]]] = {}
    dependency_dag: list[dict[str, str]] = []
    sheet_ref_pattern = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_\u4e00-\u9fff-]+))!\$?([A-Z]+)\$?(\d+)")
    for sheet in _DELIVERY_SHEETS.values():
        signatures: list[dict[str, Any]] = []
        for coordinate, value in _formula_cells(sheet):
            normalized = re.sub(r"\s+", "", str(value or "")).upper()
            normalized = normalized.replace("$", "")
            signatures.append({
                "cell": coordinate,
                "normalized_ast": normalized,
                "ast_sha256": hashlib.sha256(normalized.encode("utf-8")).hexdigest(),
            })
            for match in sheet_ref_pattern.finditer(str(value or "")):
                source_sheet = match.group(1) or match.group(2) or ""
                if source_sheet and source_sheet != sheet:
                    dependency_dag.append({"target": sheet, "source": source_sheet, "cell": coordinate})
        formula_signatures[sheet] = signatures
    # Grade ceiling must be consumed at export formal gate.
    fact_pack: dict[str, Any] = {}
    for source in (fin.get("input_revision"), fin.get("finance_inputs"), fin.get("raw"), fin):
        if isinstance(source, dict) and isinstance(source.get("finance_fact_pack"), dict):
            fact_pack = source.get("finance_fact_pack") or {}
            break
    ceiling = str(fact_pack.get("delivery_grade_ceiling") or pack_meta.get("delivery_grade_ceiling") or "summary")
    depth_ok = bool((fact_pack.get("depth_assessment") or {}).get("ok")) if isinstance(fact_pack.get("depth_assessment"), dict) else False
    grade_gate_ok = ceiling == "formal_candidate" and depth_ok
    quality_dimensions = {
        "structure": {
            "ok": bool(pack_meta.get("reference_structure_ready")),
            "coverage": 1.0 if pack_meta.get("reference_structure_ready") else 0.0,
        },
        "source": {
            "ok": float(pack_meta.get("source_coverage") or 0.0) >= 0.999,
            "coverage": float(pack_meta.get("source_coverage") or 0.0),
            "issues": pack_meta.get("source_coverage_issues") or [],
        },
        "formula": {
            "ok": formula_coverage >= 0.999 and all(value >= 0.999 for value in formula_coverage_by_table.values()),
            "coverage": formula_coverage,
            "by_table": formula_coverage_by_table,
            "required_formula_families": required_formula_families,
            "actual_formula_families": actual_formula_families,
            "dependency_only_formula_count": dependency_only_formula_count,
        },
        "independent_recalc": {
            "ok": independent_recalc_coverage >= 0.999,
            "coverage": independent_recalc_coverage,
            "checks": independent_recalc_checks,
        },
        "grade_ceiling": {
            "ok": grade_gate_ok,
            "delivery_grade_ceiling": ceiling,
            "depth_ok": depth_ok,
        },
    }
    semantic_ok = all(bool(v.get("ok")) for v in semantic_checks.values())
    formal_ready = semantic_ok and all(
        bool(item.get("ok")) for item in quality_dimensions.values()
    )
    return pack, lineage, {
        "validation_complete": formal_ready,
        "semantic_checks": semantic_checks,
        "delivery_formula_count": delivery_formula_count,
        "formula_signatures": formula_signatures,
        "formula_dependency_dag": dependency_dag,
        "actual_dependency_edges": sorted(actual_edges),
        "quality_dimensions": quality_dimensions,
        "formula_coverage_by_table": formula_coverage_by_table,
        "independent_recalc_checks": independent_recalc_checks,
        "delivery_grade_ceiling": ceiling,
    }


def assess_finance_delivery_quality(fin: dict[str, Any]) -> dict[str, Any]:
    """Evaluate the exact workbook contract without writing an XLSX artifact."""
    if not fin or not fin.get("available"):
        raise FinanceExportError("finance result unavailable")
    openpyxl, Font, PatternFill, Alignment, Border, Side, _ = _require_openpyxl()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    delivery_pack, lineage, delivery_quality = _write_delivery_tables(
        workbook, fin, Font, PatternFill, Alignment, Border, Side,
    )
    return {
        "delivery_pack": delivery_pack,
        "cell_lineage_count": len(lineage),
        "delivery_quality": delivery_quality,
    }


def export_finance_workbook(
    fin: dict[str, Any],
    path: Union[str, Path],
    *,
    model_version: str = "finance_model.v1",
    run_id: str = "",
    include_control_sheets: bool = False,
) -> dict[str, Any]:
    """Write finance review workbook to ``path``. Returns summary dict."""
    if not fin or not fin.get("available"):
        raise FinanceExportError("finance result unavailable")

    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _require_openpyxl()
    wb = openpyxl.Workbook()

    # Inputs
    ws_in = wb.active
    ws_in.title = "Inputs"
    key_rows = _write_inputs(ws_in, fin, Font, PatternFill, Alignment, Border, Side)

    # The 13 delivery sheets are the formal artifact. Legacy review sheets below remain for compatibility.
    delivery_pack, lineage, delivery_quality = _write_delivery_tables(
        wb, fin, Font, PatternFill, Alignment, Border, Side,
    )

    # Indicators
    ws_ind = wb.create_sheet("Indicators")
    _write_indicators(ws_ind, fin, key_rows, Font)

    annual = fin.get("annual") or {}
    header_font = Font(bold=True)

    ws_inc = wb.create_sheet("Income")
    _write_year_table(
        ws_inc,
        annual.get("income_statement") or [],
        ["revenue", "operating_cost", "tax_surtax", "income_tax", "net_profit", "ebit"],
        header_font,
    )
    # Add net profit formula for each data row: E = B - C - D approx (revenue - op_cost - tax) if cols match
    # Columns: A year, B revenue, C operating_cost, D tax_surtax, E income_tax, F net_profit, G ebit
    # Provide helper column H =B-C-D-E as check formula
    if ws_inc.max_row >= 2 and ws_inc.cell(1, 1).value == "year":
        ws_inc.cell(1, 8, "net_profit_check")
        ws_inc.cell(1, 8).font = header_font
        for r in range(2, ws_inc.max_row + 1):
            ws_inc.cell(r, 8, f"=B{r}-C{r}-D{r}-E{r}")

    ws_tc = wb.create_sheet("TotalCost")
    _write_year_table(
        ws_tc,
        annual.get("total_cost") or [],
        ["operating_cost", "depreciation", "amortization", "interest", "total"],
        header_font,
    )
    if ws_tc.max_row >= 2 and ws_tc.cell(1, 1).value == "year":
        ws_tc.cell(1, 7, "total_check")
        ws_tc.cell(1, 7).font = header_font
        for r in range(2, ws_tc.max_row + 1):
            ws_tc.cell(r, 7, f"=B{r}+C{r}+D{r}+E{r}")

    ws_pcf = wb.create_sheet("ProjectCF")
    _write_year_table(
        ws_pcf,
        annual.get("project_cashflow") or [],
        ["net_cashflow", "cumulative", "revenue", "op_cash_cost", "construction", "wc_change", "recover"],
        header_font,
    )

    ws_ccf = wb.create_sheet("CapitalCF")
    cap_rows = annual.get("capital_cashflow") or []
    # capital rows may be dict list with varying keys
    cap_fields = []
    if cap_rows and isinstance(cap_rows[0], dict):
        cap_fields = [k for k in cap_rows[0].keys() if k not in ("year", "period", "phase")][:8]
    _write_year_table(ws_ccf, cap_rows if isinstance(cap_rows, list) else [], cap_fields, header_font)

    ws_fp = wb.create_sheet("FinancialPlan")
    fp = annual.get("financial_plan") or []
    _write_year_table(
        ws_fp,
        fp,
        ["finance_in", "operating_net", "invest_out", "debt_service", "net_cashflow", "cumulative", "gap"],
        header_font,
    )

    ws_debt = wb.create_sheet("Debt")
    _write_year_table(
        ws_debt,
        annual.get("debt_service") or [],
        ["begin", "principal", "interest", "end", "dscr", "icr"],
        header_font,
    )

    ws_chk = wb.create_sheet("Checks")
    _write_checks(ws_chk, fin, Font)

    ws_lin = wb.create_sheet("CellLineage")
    ws_lin.append(["target_cell", "source", "method"])
    for row in lineage:
        ws_lin.append([row.get("target"), row.get("source"), row.get("method")])

    ws_diff = wb.create_sheet("TemplateDiff")
    ws_diff.append(["table", "grade", "effective", "missing_fields", "reference_schema"])
    for key, sheet_name in _DELIVERY_SHEETS.items():
        table = delivery_pack.get(key) or {}
        ws_diff.append([
            sheet_name,
            table.get("grade"),
            bool(table.get("effective")),
            "；".join(table.get("missing_fields") or []),
            table.get("reference_schema"),
        ])

    ws_formula = wb.create_sheet("FormulaAudit")
    ws_formula.append(["table", "required_formula_families", "actual_formula_families", "coverage"])
    formula_dimension = (delivery_quality.get("quality_dimensions") or {}).get("formula") or {}
    required_by_table = formula_dimension.get("required_formula_families") or {}
    actual_by_table = formula_dimension.get("actual_formula_families") or {}
    coverage_by_table = formula_dimension.get("by_table") or {}
    for key in _DELIVERY_SHEETS:
        ws_formula.append([
            key,
            ",".join(required_by_table.get(key) or []),
            ",".join(actual_by_table.get(key) or []),
            coverage_by_table.get(key),
        ])
    ws_formula.append([])
    ws_formula.append(["sheet", "cell", "normalized_ast", "ast_sha256"])
    for sheet, signatures in (delivery_quality.get("formula_signatures") or {}).items():
        for signature in signatures:
            ws_formula.append([sheet, signature.get("cell"), "'" + str(signature.get("normalized_ast") or ""), signature.get("ast_sha256")])

    ws_meta = wb.create_sheet("Meta")
    meta_font = Font(bold=True)
    ws_meta.append(["key", "value"])
    for cell in ws_meta[1]:
        cell.font = meta_font
    schema = fin.get("schema") or {}
    manifest = fin.get("model_manifest") or {}
    meta_rows = [
        ("run_id", run_id or ""),
        ("model_version", model_version),
        ("manifest_hash", fin.get("manifest_hash") or ""),
        ("manifest_version", manifest.get("manifest_version") or ""),
        ("spec_schema_version", manifest.get("spec_schema_version") or ""),
        ("policy_version", manifest.get("policy_version") or ""),
        ("industry_profile_version", manifest.get("industry_profile_version") or ""),
        ("gate_version", manifest.get("gate_version") or ""),
        ("valuation_date", fin.get("valuation_date") or ""),
        ("finance_schema_version", fin.get("finance_schema_version") or schema.get("finance_schema_version") or ""),
        ("invest_type", fin.get("invest_type") or ""),
        ("industry", fin.get("industry") or ""),
        ("exported_at", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("export_profile", "professional_13_table_v1"),
        ("delivery_sheet_count", len(_DELIVERY_SHEETS)),
        ("validation_complete", bool(delivery_quality.get("validation_complete"))),
    ]
    for k, v in meta_rows:
        ws_meta.append([k, v])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:  # noqa: BLE001
        pass

    # 【P0-2修复】手动触发计算并缓存所有公式单元格的值
    # 问题: 附表9/10有大量空单元格(公式未缓存),data_only=True读取时显示空白
    # 方案: 在保存前遍历所有公式单元格,尝试触发计算(openpyxl限制:无法真正计算,但可以保留公式)
    # 注意: openpyxl无法执行公式计算,真正的缓存需要Excel/LibreOffice打开后才能生成
    # 此修复只是确保公式被正确写入,实际缓存由_recalculate_with_soffice完成
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 公式单元格
                    formula_count += 1
                    # openpyxl会在保存时自动写入公式,这里只是统计

    if not include_control_sheets:
        delivery_names = set(_DELIVERY_SHEETS.values())
        for sheet_name in list(wb.sheetnames):
            if sheet_name not in delivery_names:
                del wb[sheet_name]

    # P1-015 修复：先写临时文件，校验完整后原子发布。原先 wb.save 直接写最终
    # 路径，1946-1959 行的 validation_complete 和 recalculation 校验失败时文件
    # 已经落盘，属于非原子写入（与 transport.py:517 的结构性问题同源）。
    # 临时文件必须保留 .xlsx 后缀，否则 _recalculate_with_soffice 的 LibreOffice
    # --convert-to xlsx 会因格式检测失败而报错。
    temporary = out.parent / f".{out.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    lock_path = str(out) + ".lock"
    try:
        wb.save(str(temporary))

        formal_requested = bool(delivery_quality.get("validation_complete"))
        recalculation = (
            _recalculate_with_soffice(temporary)
            if formal_requested
            else {"ok": False, "available": False, "skipped": True, "issues": ["reference 导出不要求实际重算"]}
        )
        if formal_requested and not recalculation.get("ok"):
            delivery_quality = dict(delivery_quality)
            delivery_quality["validation_complete"] = False
            delivery_quality["recalculation"] = recalculation
            delivery_quality.setdefault("quality_dimensions", {})["libreoffice_recalculation"] = {
                "ok": False,
                "detail": recalculation.get("issues") or [],
            }
        else:
            delivery_quality = dict(delivery_quality)
            delivery_quality["recalculation"] = recalculation

        # 临时文件已写完且校验通过（或 reference 级不要求重算），原子发布到最终路径
        with FileLock(lock_path, timeout=30):
            os.replace(temporary, out)
            temporary = None  # 成功后标记已发布，finally 不删
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()

    formula_cells = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    formula_cells += 1

    return {
        "ok": True,
        "path": str(out.resolve()),
        "sheets": wb.sheetnames,
        "formula_cells": formula_cells,
        "input_keys": list(key_rows.keys()),
        "delivery_sheets": list(_DELIVERY_SHEETS.values()),
        "delivery_sheet_count": len(_DELIVERY_SHEETS),
        "control_sheets_included": include_control_sheets,
        "cell_lineage_count": len(lineage),
        "validation_complete": bool(delivery_quality.get("validation_complete")),
        "table_quality": delivery_pack.get("_meta") or {},
        "delivery_quality": delivery_quality,
        "recalculation": recalculation,
    }
