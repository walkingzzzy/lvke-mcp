"""轻量级 xls/xlsx 读取器。

策略:
1. ``.xlsx`` 首选 ``openpyxl``(若已安装), ``.xls`` 使用 ``xlrd``。
2. 兜底使用 zipfile + xml.etree 解析 xlsx(zip 内的 ``xl/worksheets/sheet*.xml``
   + ``xl/sharedStrings.xml``),只能读纯值,不计算公式。

调用入口:

    >>> backend = pick_backend()
    >>> backend.list_sheets(path)
    >>> backend.read_sheet(path, sheet_name, max_rows, max_cols)
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


@dataclass
class ReadResult:
    sheet: str
    rows: list[list[Any]]
    row_count: int
    col_count: int
    backend: str


# ── openpyxl 后端 ──────────────────────────────────────────────────


class _OpenpyxlBackend:
    name = "openpyxl"

    def __init__(self) -> None:
        import openpyxl  # type: ignore

        self.openpyxl = openpyxl

    def list_sheets(self, path: Path) -> list[str]:
        wb = self.openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def read_sheet(
        self,
        path: Path,
        sheet: str | None,
        max_rows: int,
        max_cols: int,
    ) -> ReadResult:
        wb = self.openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet is None:
                ws = wb[wb.sheetnames[0]]
            else:
                if sheet not in wb.sheetnames:
                    raise KeyError(sheet)
                ws = wb[sheet]
            rows: list[list[Any]] = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= max_rows:
                    break
                trimmed = list(row[:max_cols])
                rows.append([_to_jsonable(v) for v in trimmed])
            return ReadResult(
                sheet=ws.title,
                rows=rows,
                row_count=len(rows),
                col_count=max(len(r) for r in rows) if rows else 0,
                backend=self.name,
            )
        finally:
            wb.close()


class _XlrdBackend:
    """Read legacy BIFF8 ``.xls`` workbooks used by the Hengli statements."""

    name = "xlrd"

    def __init__(self) -> None:
        import xlrd  # type: ignore

        self.xlrd = xlrd

    def _book(self, path: Path):
        return self.xlrd.open_workbook(str(path), on_demand=True)

    def list_sheets(self, path: Path) -> list[str]:
        book = self._book(path)
        try:
            return list(book.sheet_names())
        finally:
            book.release_resources()

    def read_sheet(
        self,
        path: Path,
        sheet: str | None,
        max_rows: int,
        max_cols: int,
    ) -> ReadResult:
        book = self._book(path)
        try:
            ws = book.sheet_by_name(sheet) if sheet else book.sheet_by_index(0)
            rows = [
                [_to_jsonable(value) for value in ws.row_values(index)[:max_cols]]
                for index in range(min(ws.nrows, max_rows))
            ]
            return ReadResult(
                sheet=ws.name,
                rows=rows,
                row_count=len(rows),
                col_count=max((len(row) for row in rows), default=0),
                backend=self.name,
            )
        finally:
            book.release_resources()


class _HybridBackend:
    name = "hybrid-openpyxl-xlrd"

    def __init__(self) -> None:
        try:
            self.xlsx = _OpenpyxlBackend()
        except Exception:
            self.xlsx = _StdlibBackend()
        self.xls = _XlrdBackend()

    def _backend(self, path: Path):
        return self.xls if path.suffix.lower() == ".xls" else self.xlsx

    def list_sheets(self, path: Path) -> list[str]:
        return self._backend(path).list_sheets(path)

    def read_sheet(self, path: Path, sheet: str | None, max_rows: int, max_cols: int) -> ReadResult:
        return self._backend(path).read_sheet(path, sheet, max_rows, max_cols)

# ── 内置 zip+xml fallback ──────────────────────────────────────────


def _col_letter_to_idx(letter: str) -> int:
    """Excel 列字母转 0-based 数字索引。例:A→0,Z→25,AA→26。"""

    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return idx - 1


def _cell_ref(ref: str) -> tuple[int, int]:
    """如 'B12' → (row=11, col=1)"""

    letters = ""
    digits = ""
    for ch in ref:
        if ch.isalpha():
            letters += ch
        else:
            digits += ch
    return int(digits) - 1, _col_letter_to_idx(letters)


class _StdlibBackend:
    name = "stdlib-zip-xml"

    def list_sheets(self, path: Path) -> list[str]:
        with zipfile.ZipFile(path) as z:
            with z.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
            sheets = []
            for node in tree.getroot().iter():
                if node.tag.endswith("}sheet"):
                    sheets.append(node.get("name", ""))
            return sheets

    def _read_shared_strings(self, z: zipfile.ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in z.namelist():
            return []
        with z.open("xl/sharedStrings.xml") as f:
            tree = ET.parse(f)
        out: list[str] = []
        for si in tree.getroot().findall("main:si", NS):
            # 串接所有 t 节点
            parts: list[str] = []
            for t in si.iter():
                if t.tag.endswith("}t") and t.text:
                    parts.append(t.text)
            out.append("".join(parts))
        return out

    def read_sheet(
        self,
        path: Path,
        sheet: str | None,
        max_rows: int,
        max_cols: int,
    ) -> ReadResult:
        with zipfile.ZipFile(path) as z:
            shared_strings = self._read_shared_strings(z)
            with z.open("xl/workbook.xml") as f:
                wb_root = ET.parse(f).getroot()
            wb_sheets: list[tuple[str, str]] = []
            for node in wb_root.iter():
                if node.tag.endswith("}sheet"):
                    name = node.get("name", "")
                    rid = node.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
                    wb_sheets.append((name, rid))
            # 关系映射
            rels: dict[str, str] = {}
            with z.open("xl/_rels/workbook.xml.rels") as f:
                rels_root = ET.parse(f).getroot()
            for r in rels_root.iter():
                if r.tag.endswith("}Relationship"):
                    rels[r.get("Id", "")] = r.get("Target", "")
            # 选 sheet
            target_name = sheet or (wb_sheets[0][0] if wb_sheets else "")
            target_rel: str | None = None
            for name, rid in wb_sheets:
                if name == target_name:
                    target_rel = rels.get(rid)
                    break
            if target_rel is None:
                raise KeyError(target_name)
            sheet_xml = "xl/" + target_rel.lstrip("/")
            if sheet_xml not in z.namelist():
                # 部分压缩包用 ../ 相对路径
                sheet_xml = target_rel
            with z.open(sheet_xml) as f:
                tree = ET.parse(f)
            # 解析单元格
            cells: dict[tuple[int, int], Any] = {}
            for c in tree.getroot().iter():
                if not c.tag.endswith("}c"):
                    continue
                ref = c.get("r")
                if not ref:
                    continue
                row_i, col_i = _cell_ref(ref)
                if row_i >= max_rows or col_i >= max_cols:
                    continue
                ctype = c.get("t", "n")
                v = c.find("main:v", NS)
                val: Any = None
                if v is not None and v.text is not None:
                    if ctype == "s":
                        try:
                            val = shared_strings[int(v.text)]
                        except (ValueError, IndexError):
                            val = v.text
                    elif ctype == "b":
                        val = bool(int(v.text))
                    else:
                        # 数字 / 日期
                        try:
                            val = int(v.text) if "." not in v.text else float(v.text)
                        except ValueError:
                            val = v.text
                else:
                    inline = c.find("main:is/main:t", NS)
                    if inline is not None:
                        val = inline.text
                cells[(row_i, col_i)] = val
            # 投影到矩阵
            if not cells:
                return ReadResult(
                    sheet=target_name,
                    rows=[],
                    row_count=0,
                    col_count=0,
                    backend=self.name,
                )
            max_r = min(max_rows, max(r for r, _ in cells.keys()) + 1)
            max_c = min(max_cols, max(c for _, c in cells.keys()) + 1)
            rows: list[list[Any]] = [
                [cells.get((r, c)) for c in range(max_c)] for r in range(max_r)
            ]
            return ReadResult(
                sheet=target_name,
                rows=rows,
                row_count=max_r,
                col_count=max_c,
                backend=self.name,
            )


def _to_jsonable(v: Any) -> Any:
    # openpyxl 可能返回 datetime / Decimal 等;转成普通类型
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def pick_backend():
    try:
        return _HybridBackend()
    except Exception:
        return _StdlibBackend()
