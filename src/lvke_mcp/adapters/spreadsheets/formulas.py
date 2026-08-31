"""公式解析后端（P2：Excel 对照能力服务化）。

`reader.py` 用 ``data_only=True`` 只读缓存值、读不到公式；本模块用
``openpyxl`` 双开工作簿（``data_only=False`` 读公式 + ``data_only=True`` 读缓存值），
把根目录离线脚本（``excel_deep_analyzer.py`` / ``formula_deep_analyzer.py`` /
``evidence_chain_audit.py``）的核心解析算法沉淀为可复用后端：

- ``read_formulas``:    抽取某 sheet 的公式单元格（公式文本 + 缓存值 + 引用）。
- ``cross_sheet_refs``: 汇总跨表引用（from → to 及公式）。
- ``dependency_tree``:  对某单元格递归追公式依赖树（跨表，带环保护）。

依赖 ``openpyxl``；不可用时抛 ``FormulaBackendUnavailable``，由 server 层转成
明确的降级错误（不静默返回空，避免"读不到公式"被误判为"没有公式"）。
"""

from __future__ import annotations

import re
from collections import defaultdict
from typing import Any

# 单元格/区间: A1 或 A1:B2（允许 $ 绝对符）
_CELL = r"\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?"
# 跨表引用（启发式，无 sheet 名清单时用）: Sheet!A1 或 'Sheet Name'!A1。
# 名称段允许中文/字母/数字/连字符/下划线（覆盖「附表6-3」这类含连字符的表名），
# 但不含运算符/括号/引号；优先用 extract_references(sheet_names=...) 的精确匹配。
_CROSS_SHEET = re.compile(r"'([^']+)'!(" + _CELL + r")|([\w一-鿿-]+?)!(" + _CELL + r")")
# 同表引用: A1, A1:B2
_SAME_SHEET = re.compile(r"(?<![A-Za-z0-9_!一-鿿])(" + _CELL + r")")


class FormulaBackendUnavailable(RuntimeError):
    """openpyxl 不可用时抛出（公式解析强依赖它，不能静默降级为空）。"""


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as exc:  # noqa: BLE001
        raise FormulaBackendUnavailable(
            "公式解析需要 openpyxl（data_only=False 读公式），当前环境不可用"
        ) from exc


def extract_references(formula: str, sheet_names: list[str] | None = None) -> list[str]:
    """从公式文本提取单元格引用（跨表在前、同表在后，去重去 $ 绝对符）。

    ``sheet_names`` 提供时优先做**精确表名匹配**——这是可靠路径：像「附表6-3」
    这类含连字符的表名，启发式正则会在 ``-`` 处误切，只有拿到真实表名清单才能
    正确切出 ``附表6-3!D5``。无清单时退启发式（可能对含连字符表名不准）。
    """
    refs: list[str] = []
    seen: set[str] = set()
    consumed_spans: list[tuple[int, int]] = []

    def _add(ref: str) -> None:
        if ref not in seen:
            seen.add(ref)
            refs.append(ref)

    if sheet_names:
        # 精确匹配：按表名长度降序（先匹配「附表6-3」再匹配「附表6」），
        # 支持带/不带单引号包裹的表名。
        cell_re = re.compile(_CELL)
        for name in sorted(sheet_names, key=len, reverse=True):
            for quoted in (f"'{name}'!", f"{name}!"):
                start = 0
                while True:
                    idx = formula.find(quoted, start)
                    if idx < 0:
                        break
                    after = idx + len(quoted)
                    cm = cell_re.match(formula, after)
                    if cm:
                        _add(f"{name}!{cm.group(0).replace('$', '')}")
                        consumed_spans.append((idx, cm.end()))
                    start = after
    else:
        for m in _CROSS_SHEET.finditer(formula):
            sheet = (m.group(1) or m.group(3) or "").strip()
            cell = (m.group(2) or m.group(4) or "").replace("$", "")
            if sheet and cell:
                _add(f"{sheet}!{cell}")
                consumed_spans.append(m.span())

    for m in _SAME_SHEET.finditer(formula):
        # 跳过已被跨表引用消费的区段（避免把 Sheet!A1 的 A1 再当同表引用）
        if any(s <= m.start() < e for s, e in consumed_spans):
            continue
        _add(m.group(1).replace("$", ""))
    return refs


class FormulaBackend:
    """openpyxl 双开工作簿：公式 + 缓存值。按需缓存已加载的工作簿。"""

    name = "openpyxl-formula"

    def __init__(self, path: str) -> None:
        self._openpyxl = _require_openpyxl()
        self.path = path
        self._wb_f = None  # data_only=False（公式）
        self._wb_v = None  # data_only=True（缓存值）

    def _formula_wb(self):
        if self._wb_f is None:
            self._wb_f = self._openpyxl.load_workbook(self.path, data_only=False)
        return self._wb_f

    def _value_wb(self):
        if self._wb_v is None:
            self._wb_v = self._openpyxl.load_workbook(self.path, data_only=True)
        return self._wb_v

    def close(self) -> None:
        for wb in (self._wb_f, self._wb_v):
            try:
                if wb is not None:
                    wb.close()
            except Exception:  # noqa: BLE001
                pass
        self._wb_f = self._wb_v = None

    def sheet_names(self) -> list[str]:
        return list(self._formula_wb().sheetnames)

    def read_formulas(self, sheet: str, *, max_rows: int = 500, max_cols: int = 60) -> dict[str, Any]:
        """抽取某 sheet 的公式单元格：公式文本 + 缓存值 + 引用。"""
        from openpyxl.utils import get_column_letter  # type: ignore

        wb_f = self._formula_wb()
        wb_v = self._value_wb()
        if sheet not in wb_f.sheetnames:
            raise KeyError(sheet)
        ws = wb_f[sheet]
        wsv = wb_v[sheet] if sheet in wb_v.sheetnames else None
        names = list(wb_f.sheetnames)
        cells: list[dict[str, Any]] = []
        rmax = min(max_rows, ws.max_row or 0)
        cmax = min(max_cols, ws.max_column or 0)
        for row in range(1, rmax + 1):
            for col in range(1, cmax + 1):
                cell = ws.cell(row, col)
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                ref = f"{get_column_letter(col)}{row}"
                formula = v[1:]
                cached = wsv.cell(row, col).value if wsv is not None else None
                cells.append({
                    "cell": ref,
                    "formula": formula,
                    "cached_value": _jsonable(cached),
                    "references": extract_references(formula, names),
                })
        return {
            "sheet": sheet,
            "formula_count": len(cells),
            "cells": cells,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }

    def cross_sheet_refs(self) -> dict[str, Any]:
        """汇总全工作簿的跨表引用（from sheet → to sheet 的 from/to 明细）。"""
        from openpyxl.utils import get_column_letter  # type: ignore

        wb_f = self._formula_wb()
        refs: dict[str, list[dict[str, str]]] = defaultdict(list)
        matrix: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        names = list(wb_f.sheetnames)
        for sheet in wb_f.sheetnames:
            ws = wb_f[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    src = f"{get_column_letter(cell.column)}{cell.row}"
                    for r in extract_references(v[1:], names):
                        if "!" not in r:
                            continue
                        tgt_sheet = r.split("!", 1)[0]
                        if tgt_sheet == sheet:
                            continue
                        refs[sheet].append({"from": src, "to": r, "formula": v[1:]})
                        matrix[sheet][tgt_sheet] += 1
        return {
            "cross_sheet_refs": {k: v for k, v in refs.items()},
            "matrix": {k: dict(v) for k, v in matrix.items()},
            "total": sum(len(v) for v in refs.values()),
        }

    def dependency_tree(
        self,
        sheet: str,
        cell: str,
        *,
        max_depth: int = 6,
        max_range_cells: int = 64,
        max_nodes: int = 2000,
    ) -> dict[str, Any]:
        """对 sheet!cell 递归追公式依赖树（跨表，带环保护，区间展开且截断可见）。

        缺陷（A2）：区间引用此前被 ``c_cell.split(":", 1)[0]`` 压成左上角单元格，
        ``成本明细!B6 = SUM(B2:B5)`` 只展开出 B2，B3/B4/B5 静默消失且输出侧没有
        任何标记 —— 调用方拿到的是一棵看着完整、实际漏了 3/4 依赖的树。更糟的是
        ``SUM(汇总!B2:B3)`` 复用了深度/环保护那个 ``truncated:true``，让"区间被砍"
        与"递归到底"不可区分。

        现在区间会真的展开（``max_range_cells`` 为每个区间的展开上限），并把区间
        本身登记进节点的 ``range_references``：单元格数、实际展开数、是否被截断、
        区间原文。超上限时子节点带 ``range_truncated``，与深度/环保护的
        ``truncated`` 分开命名，三种截断（``cycle_or_max_depth``/``already_visited``/
        ``max_range_cells``/``max_nodes``）各自可归因，不再混为一个 ``truncated``。

        ``max_nodes`` 是展开后新增的总量护栏：原实现靠"区间只取一格"隐式控规模，
        去掉那个压缩后必须有个显式上限，否则一条 ``SUM(A:Z)`` 就能把树撑爆。
        """

        wb_f = self._formula_wb()
        wb_v = self._value_wb()
        names = list(wb_f.sheetnames)

        def _formula_at(sh: str, cl: str) -> tuple[str | None, Any]:
            if sh not in wb_f.sheetnames:
                return None, None
            ws = wb_f[sh]
            m = re.match(r"\$?([A-Z]+)\$?(\d+)", cl)
            if not m:
                return None, None
            from openpyxl.utils import column_index_from_string  # type: ignore

            c = column_index_from_string(m.group(1))
            r = int(m.group(2))
            val = ws.cell(r, c).value
            cached = wb_v[sh].cell(r, c).value if sh in wb_v.sheetnames else None
            if isinstance(val, str) and val.startswith("="):
                return val[1:], _jsonable(cached)
            return None, _jsonable(val if val is not None else cached)

        visited: set[str] = set()
        budget = {"nodes": max(1, int(max_nodes))}

        def _walk(sh: str, cl: str, depth: int) -> dict[str, Any]:
            node_id = f"{sh}!{cl}"
            # 三种截断分别归因。原实现把"已访问过"和"超深度"合成一个
            # ``truncated:true``；区间展开后重复引用（B2:C3 与单点 B2 同现）会
            # 大量命中"已访问"，若仍报 cycle_or_max_depth 就是在误导。
            if node_id in visited:
                return {
                    "id": node_id,
                    "truncated": True,
                    "truncation_reason": "already_visited",
                }
            if depth > max_depth:
                return {
                    "id": node_id,
                    "truncated": True,
                    "truncation_reason": "max_depth",
                }
            if budget["nodes"] <= 0:
                return {
                    "id": node_id,
                    "truncated": True,
                    "truncation_reason": "max_nodes",
                }
            budget["nodes"] -= 1
            visited.add(node_id)
            formula, value = _formula_at(sh, cl)
            node: dict[str, Any] = {"id": node_id, "value": value}
            if formula is None:
                node["type"] = "input"
                return node
            node["type"] = "formula"
            node["formula"] = formula
            children: list[dict[str, Any]] = []
            range_refs: list[dict[str, Any]] = []
            for ref in extract_references(formula, names):
                if "!" in ref:
                    c_sheet, c_ref = ref.split("!", 1)
                else:
                    c_sheet, c_ref = sh, ref
                if ":" not in c_ref:
                    children.append(_walk(c_sheet, c_ref, depth + 1))
                    continue
                cells, total = _expand_range(c_ref, limit=max_range_cells)
                if not cells:
                    # 区间解析不了：绝不退回"取左上角"，那会伪装成完整依赖。
                    range_refs.append({
                        "reference": f"{c_sheet}!{c_ref}",
                        "range": c_ref,
                        "cell_count": None,
                        "expanded_count": 0,
                        "range_truncated": True,
                        "reason": "range_unparsable",
                    })
                    children.append({
                        "id": f"{c_sheet}!{c_ref}",
                        "type": "range",
                        "range_truncated": True,
                        "truncation_reason": "range_unparsable",
                    })
                    continue
                truncated = total > len(cells)
                range_refs.append({
                    "reference": f"{c_sheet}!{c_ref}",
                    "range": c_ref,
                    "cell_count": total,
                    "expanded_count": len(cells),
                    "range_truncated": truncated,
                    **({"reason": "max_range_cells"} if truncated else {}),
                })
                for c_cell in cells:
                    child = _walk(c_sheet, c_cell, depth + 1)
                    # 标注这个子节点来自哪个区间，让"哪些格属于 B2:B5"可追。
                    child["from_range"] = f"{c_sheet}!{c_ref}"
                    children.append(child)
                if truncated:
                    children.append({
                        "id": f"{c_sheet}!{c_ref}",
                        "type": "range",
                        "range_truncated": True,
                        "truncation_reason": "max_range_cells",
                        "cell_count": total,
                        "expanded_count": len(cells),
                    })
            if children:
                node["depends_on"] = children
            if range_refs:
                node["range_references"] = range_refs
                if any(item["range_truncated"] for item in range_refs):
                    node["range_truncated"] = True
            return node

        return _walk(sheet, cell.replace("$", ""), 0)


_RANGE_ENDPOINT = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")


def _expand_range(ref: str, *, limit: int) -> tuple[list[str], int]:
    """把 ``B2:B5`` 展开成 ``["B2","B3","B4","B5"]``。

    返回 ``(展开出的单元格, 区间总格数)``。总数独立返回，是为了让调用方能区分
    "全展开"与"到 limit 就停了"——只给列表的话，截断在输出上不可见，正是 A2 的
    失效模式。解析不了时返回 ``([], 0)``：宁可显式报不可解析，也不猜。
    """

    from openpyxl.utils import column_index_from_string, get_column_letter  # type: ignore

    parts = str(ref or "").split(":")
    if len(parts) != 2:
        return [], 0
    start, end = (_RANGE_ENDPOINT.match(part.strip()) for part in parts)
    if start is None or end is None:
        return [], 0
    try:
        col_start = column_index_from_string(start.group(1).upper())
        col_end = column_index_from_string(end.group(1).upper())
    except ValueError:
        return [], 0
    row_start, row_end = int(start.group(2)), int(end.group(2))
    # Excel 允许倒序写法（B5:B2），归一化后再展开。
    col_start, col_end = min(col_start, col_end), max(col_start, col_end)
    row_start, row_end = min(row_start, row_end), max(row_start, row_end)
    total = (col_end - col_start + 1) * (row_end - row_start + 1)
    cells: list[str] = []
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            if len(cells) >= max(1, int(limit)):
                return cells, total
            cells.append(f"{get_column_letter(col)}{row}")
    return cells, total


def _jsonable(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)

# 门面模块的公开面。显式声明而不是靠"碰巧 import 了"——API 快照门禁
# (tests/integration/test_refactor_guardrails.py) 要求这些 re-export 保持
# 可达,而 ruff F401 会把它们判成未使用。写成 __all__ 让两个门禁同时成立,
# 也让"哪些名字是刻意对外的"可读。
__all__ = [
    "Any",
    "FormulaBackend",
    "FormulaBackendUnavailable",
    "_CELL",
    "_CROSS_SHEET",
    "_SAME_SHEET",
    "_expand_range",
    "_jsonable",
    "_require_openpyxl",
    "defaultdict",
    "extract_references",
    "re",
]
