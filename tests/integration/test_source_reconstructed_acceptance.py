from __future__ import annotations

import hashlib
import json
import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from lvke_mcp.domains.finance.evidence_binding import bind_finance_spec_evidence
from lvke_mcp.domains.finance.hengli_reference import scenario_matrix
from lvke_mcp.adapters.spreadsheets.reader import pick_backend
from lvke_mcp.servers.lvke_data_analysis import service as analysis
from lvke_mcp.runtime.source_reconstruction import reconstruction_errors
from lvke_mcp.servers.lvke_feasibility_delivery import service as delivery
from lvke_mcp.testing.source_reconstructed_acceptance import (
    load_nine_chapter_bodies,
    run_reconstructed_acquisition_case,
    run_reconstructed_delivery_release,
    run_reconstructed_finance_case,
    run_reconstructed_planning_case,
    run_reconstructed_report_case,
    run_reconstructed_research_case,
    run_reconstructed_review_closure,
)


ROOT = Path(__file__).resolve().parents[2]
MANIFEST = ROOT / "tests/fixtures/source_reconstructed/manifest.json"
class SourceReconstructedAcceptanceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory(prefix="lvke-source-reconstructed-")
        self.previous = os.environ.get("LVKE_MCP_DATA_DIR")
        self.previous_golden = os.environ.get("LVKE_GOLDEN_DATA_ROOT")
        self.previous_seal_secret = os.environ.get("LVKE_FACT_PACK_SEAL_SECRET")
        os.environ["LVKE_MCP_DATA_DIR"] = self.tempdir.name
        os.environ["LVKE_GOLDEN_DATA_ROOT"] = str(ROOT / "docs")
        os.environ["LVKE_FACT_PACK_SEAL_SECRET"] = "source-reconstructed-acceptance-test"

    def tearDown(self) -> None:
        if self.previous is None:
            os.environ.pop("LVKE_MCP_DATA_DIR", None)
        else:
            os.environ["LVKE_MCP_DATA_DIR"] = self.previous
        if self.previous_golden is None:
            os.environ.pop("LVKE_GOLDEN_DATA_ROOT", None)
        else:
            os.environ["LVKE_GOLDEN_DATA_ROOT"] = self.previous_golden
        if self.previous_seal_secret is None:
            os.environ.pop("LVKE_FACT_PACK_SEAL_SECRET", None)
        else:
            os.environ["LVKE_FACT_PACK_SEAL_SECRET"] = self.previous_seal_secret
        self.tempdir.cleanup()

    def test_real_material_manifest_hashes_and_report_revisions(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
        self.assertEqual(manifest["evidence_policy"], "source_reconstructed")
        self.assertFalse(manifest["project_fact_certified"])
        for source in manifest["sources"]:
            path = ROOT / source["path"]
            self.assertTrue(path.is_file(), source["path"])
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            self.assertEqual(source["content_hash"], f"sha256:{digest}")
            self.assertTrue(source["locator"])
            self.assertTrue(source["method"])
        for case in ("cy_xiangyuan", "qianshan_forest_park"):
            revision = ROOT / manifest["cases"][case]["revised_report"]
            text = revision.read_text(encoding="utf-8")
            self.assertEqual(sum(1 for i in range(1, 10) if f"第{i}章" in text), 9)

    def test_finance_templates_have_sheets_and_formulas(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
        template_ids = {
            "srcx_template_real_estate",
            "srcx_template_product_1",
            "srcx_template_product_2",
            "srcx_template_lease",
            "srcx_template_cemetery",
        }
        rows = [row for row in manifest["sources"] if row["source_id"] in template_ids]
        self.assertEqual(len(rows), 5)
        for row in rows:
            workbook = load_workbook(ROOT / row["path"], data_only=False, read_only=True)
            self.assertTrue(workbook.sheetnames, row["path"])
            formula_count = 0
            for sheet in workbook.worksheets:
                for values in sheet.iter_rows(values_only=True):
                    formula_count += sum(
                        1 for value in values if isinstance(value, str) and value.startswith("=")
                    )
            workbook.close()
            self.assertGreater(formula_count, 0, row["path"])

    def test_real_template_runs_through_formal_finance_chain(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
        source = next(
            row for row in manifest["sources"]
            if row["source_id"] == "srcx_template_product_1"
        )
        result = run_reconstructed_finance_case(
            ROOT / source["path"],
            workspace_id="source-reconstructed-real-finance",
            valuation_date="2026-08-05",
            case_key="product-template",
        )
        self.assertTrue(result["finance_run_id"].startswith("run_"), result)
        self.assertTrue(result["finance_tables_package_id"].startswith("ftp_"), result)
        self.assertTrue(result["xlsx_resource_uri"].endswith("/xlsx"), result)
        self.assertTrue(result["formal_validation"]["success"], result)
        self.assertEqual(result["evidence_policy"], "source_reconstructed")
        self.assertFalse(result["project_fact_certified"])

    def test_reconstruction_contract_and_finance_binding(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
        record = manifest["sources"][0]
        self.assertEqual(reconstruction_errors(record), [])
        incomplete = dict(record)
        incomplete.pop("locator")
        self.assertIn("locator_required", reconstruction_errors(incomplete))
        blocked_pack = analysis.build_evidence_pack(
            "source-reconstructed-test",
            "missing-analysis-task",
            None,
            [],
            [],
            evidence_track="source_reconstructed",
            reconstruction_records=[incomplete],
        )
        self.assertFalse(blocked_pack["success"])
        self.assertEqual(blocked_pack["code"], "source_reconstruction_invalid")
        spec = {
            "evidence_policy": "source_reconstructed",
            "reconstruction_records": [record],
        }
        binding = bind_finance_spec_evidence("source-reconstructed-test", spec)
        self.assertTrue(binding["formal_ok"], binding)
        self.assertEqual(binding["evidence_policy"], "source_reconstructed")
        self.assertFalse(binding["project_fact_certified"])
        self.assertEqual(binding["reconstruction_ids"], [record["source_id"]])

    def test_two_client_reports_run_real_nine_chapter_review_and_release_chains(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))

        def source(source_id: str) -> dict:
            return next(row for row in manifest["sources"] if row["source_id"] == source_id)

        configurations = (
            (
                "cy_xiangyuan",
                "崇阳香苑小区一期贷款项目",
                "real_estate",
                "srcx_template_product_1",
            ),
            (
                "qianshan_forest_park",
                "潜山国家森林公园文旅项目",
                "cultural_tourism",
                "srcx_template_product_1",
            ),
        )
        releases = []
        for case_id, project_name, industry_code, runnable_template_id in configurations:
            case = manifest["cases"][case_id]
            additional_ids = [
                case["report_source_id"],
                *[
                    item for item in case["template_source_ids"]
                    if item != runnable_template_id
                ],
            ]
            finance = run_reconstructed_finance_case(
                ROOT / source(runnable_template_id)["path"],
                workspace_id=f"source-reconstructed-{case_id}",
                valuation_date="2026-08-05",
                case_key=case_id,
                additional_reconstruction_records=[source(item) for item in additional_ids],
            )
            planning = run_reconstructed_planning_case(
                finance,
                project_name=project_name,
                industry_code=industry_code,
            )
            research = run_reconstructed_research_case(
                finance,
                planning,
                topic=f"{project_name}来源重建研究",
                limitations=case["unresolved_inputs"],
            )
            report = run_reconstructed_report_case(
                finance,
                planning,
                research,
                chapter_contents=load_nine_chapter_bodies(ROOT / case["revised_report"]),
                report_key=case_id,
                unresolved_inputs=case["unresolved_inputs"],
            )
            review = run_reconstructed_review_closure(
                finance,
                report,
                review_key=case_id,
            )
            released = run_reconstructed_delivery_release(
                finance,
                planning,
                research,
                review,
                release_key=case_id,
                unresolved_inputs=case["unresolved_inputs"],
            )
            self.assertTrue(released["release_id"].startswith("fdrp_"), released)
            self.assertEqual(released["release_scope"], "process_acceptance")
            self.assertFalse(released["project_fact_certified"])
            self.assertTrue(released["lineage_hash"].startswith("sha256:"))
            self.assertEqual(review["finding_status"], "resolved")
            self.assertEqual(len(report["outline"]), 9)
            releases.append(released)
        self.assertEqual(len({row["delivery_run_id"] for row in releases}), 2)
        self.assertEqual(len({row["report_revision_id"] for row in releases}), 2)

    def _complete_run(self, workspace: str, *, release_scope: str) -> tuple[str, dict]:
        started = delivery.start({
            "workspace_id": workspace,
            "delivery_mode": "formal_release",
            "evidence_policy": "source_reconstructed",
            "release_scope": release_scope,
            "project_fact_certified": False,
            "reconstructed_source_ids": ["srcx_cy_report"],
            "unresolved_inputs": ["original_boe"],
            "release_limitations": ["report-derived values are reconstructed"],
            "idempotency_key": f"start-{workspace}",
        })
        self.assertTrue(started["success"], started)
        run_id = started["delivery_run_id"]
        rejected = delivery.stage({
            "workspace_id": workspace,
            "delivery_run_id": run_id,
            "stage": "project",
            "status": "completed",
            "output_refs": ["project-does-not-exist"],
            "basis_hash": "sha256:" + hashlib.sha256(b"project").hexdigest(),
            "idempotency_key": f"stage-{workspace}-project",
        })
        self.assertFalse(rejected["success"], rejected)
        self.assertEqual(rejected["code"], "stage_reference_invalid")
        return run_id, delivery.validate({
            "workspace_id": workspace,
            "delivery_run_id": run_id,
            "scope": "formal",
        })

    def test_fake_process_acceptance_is_blocked_and_project_delivery_stays_blocked(self) -> None:
        process_run, validation = self._complete_run("cy-process", release_scope="process_acceptance")
        self.assertFalse(validation["success"], validation)
        self.assertIn("reconstruction_records_missing", validation["blockers"])
        released = delivery.release({
            "workspace_id": "cy-process",
            "delivery_run_id": process_run,
            "release_scope": "process_acceptance",
            "release_note": "source reconstructed process acceptance",
            "idempotency_key": "release-cy-process",
        })
        self.assertFalse(released["success"], released)
        self.assertEqual(released["code"], "formal_validation_required")

        delivery_run, blocked_validation = self._complete_run("cy-delivery", release_scope="project_delivery")
        self.assertFalse(blocked_validation["success"])
        self.assertIn("project_fact_evidence_missing", blocked_validation["blockers"])
        blocked = delivery.release({
            "workspace_id": "cy-delivery",
            "delivery_run_id": delivery_run,
            "release_scope": "project_delivery",
            "idempotency_key": "release-cy-delivery",
        })
        self.assertFalse(blocked["success"])
        self.assertEqual(blocked["code"], "project_fact_evidence_missing")

        controlled = delivery.start({
            "workspace_id": "controlled-formal",
            "delivery_mode": "formal_release",
            "evidence_policy": "controlled_assumption",
            "release_scope": "process_acceptance",
            "idempotency_key": "controlled-start",
        })
        self.assertTrue(controlled["success"], controlled)
        controlled_validation = delivery.validate({
            "workspace_id": "controlled-formal",
            "delivery_run_id": controlled["delivery_run_id"],
            "scope": "formal",
        })
        self.assertFalse(controlled_validation["success"])
        self.assertIn("controlled_assumption_formal_forbidden", controlled_validation["blockers"])

    def test_hengli_six_scenarios_and_historical_statements(self) -> None:
        manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
        hengli = manifest["cases"]["hengli_hotel"]
        self.assertEqual(hengli["scenario_purchase_prices_wan"], [2000, 2200, 2400, 2600, 2800, 3000])
        self.assertEqual(len(hengli["historical_source_ids"]), 6)
        for source_id in hengli["historical_source_ids"]:
            row = next(item for item in manifest["sources"] if item["source_id"] == source_id)
            path = ROOT / row["path"]
            self.assertTrue(path.is_file())
            backend = pick_backend()
            sheets = backend.list_sheets(path)
            self.assertTrue(sheets, row["path"])
            read = backend.read_sheet(path, sheets[0], max_rows=20, max_cols=20)
            self.assertEqual(read.sheet, sheets[0])
            self.assertGreaterEqual(read.row_count, 1)
        reference = scenario_matrix()
        self.assertTrue(reference["valid"], reference)
        self.assertEqual(
            [item["purchase_price_wan"] for item in reference["scenarios"]],
            hengli["scenario_purchase_prices_wan"],
        )
        self.assertEqual(hengli["business_decision_status"], "not_selected")
        reconstruction_records = [
            next(item for item in manifest["sources"] if item["source_id"] == source_id)
            for source_id in hengli["historical_source_ids"]
        ]
        releases = []
        for scenario in reference["scenarios"]:
            price = int(scenario["purchase_price_wan"])
            finance = run_reconstructed_acquisition_case(
                workspace_id=f"source-reconstructed-hengli-{price}",
                scenario=scenario,
                reconstruction_records=reconstruction_records,
                unresolved_inputs=hengli["unresolved_inputs"],
            )
            planning = run_reconstructed_planning_case(
                finance,
                project_name=f"恒立酒店资产收购{price}万元情景",
                industry_code="hotel",
                project_type="asset_acquisition",
            )
            research = run_reconstructed_research_case(
                finance,
                planning,
                topic=f"恒立酒店{price}万元情景研究",
                limitations=hengli["unresolved_inputs"],
            )
            citation = reconstruction_records[0]["locator"]
            chapter_contents = [
                f"恒立酒店来源重建过程验收第{chapter}章。"
                f"收购价{price}万元，评估值4027.53万元，总投资"
                f"{scenario['total_investment_wan']}万元（来源: {citation}）。"
                "经营模式、交易类型和最终收购价保持未选择。"
                for chapter in range(1, 10)
            ]
            report = run_reconstructed_report_case(
                finance,
                planning,
                research,
                chapter_contents=chapter_contents,
                report_key=f"hengli-{price}",
                unresolved_inputs=hengli["unresolved_inputs"],
            )
            review = run_reconstructed_review_closure(
                finance,
                report,
                review_key=f"hengli-{price}",
            )
            released = run_reconstructed_delivery_release(
                finance,
                planning,
                research,
                review,
                release_key=f"hengli-{price}",
                unresolved_inputs=hengli["unresolved_inputs"],
                business_decision_status="not_selected",
            )
            self.assertAlmostEqual(finance["total_investment_wan"], scenario["total_investment_wan"], places=2)
            self.assertEqual(len(finance["csv_resource_uris"]), 13)
            self.assertEqual(released["business_decision_status"], "not_selected")
            self.assertTrue(released["release_id"].startswith("fdrp_"))
            releases.append(released)
        for field in (
            "finance_run_id",
            "finance_tables_package_id",
            "report_revision_id",
            "review_run_id",
            "lineage_hash",
        ):
            self.assertEqual(len({row[field] for row in releases}), 6, field)


if __name__ == "__main__":
    unittest.main()
