from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from lvke_mcp.domains.finance import tables_service
from lvke_mcp.domains.finance.industry_scenario_factory import build_industry_scenarios
from lvke_mcp.domains.finance.model_application import run_model


class XlsxTechnicalPreviewScopeTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory(prefix="lvke-xlsx-scope-")
        self.previous_data_dir = os.environ.get("LVKE_MCP_DATA_DIR")
        self.previous_golden_root = os.environ.get("LVKE_GOLDEN_DATA_ROOT")
        os.environ["LVKE_MCP_DATA_DIR"] = self.tempdir.name
        os.environ["LVKE_GOLDEN_DATA_ROOT"] = str(
            Path(__file__).resolve().parents[2] / "docs"
        )
        self.workspace = "xlsx-scope-test"
        scenario = next(
            item
            for item in build_industry_scenarios("transport_logistics")
            if item["archetype_id"] == "urban_rail" and item["variant_id"] == "base"
        )
        result = run_model({
            "workspace_id": self.workspace,
            "spec": scenario["spec"],
            "input_revision": scenario["finance"],
            "mode": "estimate_preview",
            "idempotency_key": "xlsx-scope-run",
        })
        self.assertTrue(result["success"], result)
        self.run_id = result["run_id"]
        self.package_id = tables_service.render(self.workspace, self.run_id)[
            "finance_tables_package_id"
        ]

    def tearDown(self) -> None:
        for name, previous in (
            ("LVKE_MCP_DATA_DIR", self.previous_data_dir),
            ("LVKE_GOLDEN_DATA_ROOT", self.previous_golden_root),
        ):
            if previous is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = previous
        self.tempdir.cleanup()

    def _export(self, scope: str) -> dict:
        return tables_service.export_xlsx(
            self.workspace,
            self.run_id,
            "",
            self.package_id,
            scope,
        )

    def test_invalid_scope_is_rejected(self) -> None:
        self.assertEqual(self._export("other")["code"], "validation_scope_invalid")

    def test_technical_xlsx_has_exactly_thirteen_marked_sheets(self) -> None:
        exported = self._export("technical")
        self.assertEqual(exported["validation_scope"], "technical")
        self.assertEqual(exported["release_grade"], "technical_preview")
        self.assertFalse(exported["validation_complete"])
        self.assertTrue(exported["not_for_formal_use"])
        self.assertEqual(exported["blockers"], [])
        self.assertTrue(exported["xlsx_resource"].endswith("/xlsx-technical"))
        self.assertFalse(exported["xlsx_validation"]["validation_complete"])
        resolved = tables_service.resolve_resource(
            exported["xlsx_resource"], self.workspace,
        )
        self.assertIsNotNone(resolved)
        self.assertIsInstance((resolved or (None, ""))[0], bytes)
        workbook = load_workbook(exported["deliverable_path"], read_only=True, data_only=False)
        try:
            self.assertEqual(len(workbook.sheetnames), 13)
            for sheet in workbook.worksheets:
                with self.subTest(sheet=sheet.title):
                    notice = str(sheet["A2"].value or "")
                    self.assertIn("估算预览", notice)
                    self.assertIn("过程验收", notice)
                    self.assertIn("不得作为正式投资决策依据", notice)
        finally:
            workbook.close()

    def test_technical_xlsx_keeps_the_requested_package_binding(self) -> None:
        exported = self._export("technical")
        self.assertEqual(exported["source_package_id"], self.package_id)
        self.assertEqual(exported["source_run_id"], self.run_id)
        self.assertTrue(exported["manifest_hash"])

    def test_formal_scope_produces_formal_candidate_when_qualification_is_incomplete(self) -> None:
        before = sorted(Path(self.tempdir.name).rglob("*.xlsx"))
        exported = self._export("formal")
        self.assertEqual(exported["validation_scope"], "formal")
        self.assertEqual(exported["release_grade"], "formal_candidate")
        self.assertFalse(exported["technical_preview"])
        self.assertTrue(exported["not_for_formal_use"])
        self.assertFalse(exported["validation_complete"])
        self.assertEqual(exported["delivery_mode"], "draft")
        self.assertEqual(exported["blockers"], [])
        self.assertIn("xlsx_formal_quality_incomplete", exported["quality_issues"])
        self.assertIn("xlsx_resource", exported)
        self.assertIn("deliverable_path", exported)
        self.assertTrue(Path(exported["deliverable_path"]).is_file())
        self.assertEqual(
            len(sorted(Path(self.tempdir.name).rglob("*.xlsx"))),
            len(before) + 1,
        )

    def test_formal_candidate_file_has_the_candidate_banner(self) -> None:
        exported = self._export("formal")
        self.assertIn("【正式候选·含限制】", exported["in_file_marking"])
        workbook = load_workbook(exported["deliverable_path"], read_only=True, data_only=False)
        try:
            for sheet in workbook.worksheets:
                with self.subTest(sheet=sheet.title):
                    notice = str(sheet["A2"].value or "")
                    self.assertIn("正式候选", notice)
                    self.assertIn("不得作为对外正式交付物", notice)
        finally:
            workbook.close()

    def test_technical_and_formal_candidate_coexist(self) -> None:
        technical = self._export("technical")
        technical_path = Path(technical["deliverable_path"])
        technical_bytes = technical_path.read_bytes()
        formal = self._export("formal")
        formal_path = Path(formal["deliverable_path"])
        # 两个文件路径不同
        self.assertNotEqual(str(technical_path), str(formal_path))
        # technical 文件未受影响
        self.assertEqual(technical_path.read_bytes(), technical_bytes)
        # formal 文件存在且独立
        self.assertTrue(formal_path.is_file())
        xlsx_files = sorted(Path(self.tempdir.name).rglob("*.xlsx"))
        self.assertIn(technical_path, xlsx_files)
        self.assertIn(formal_path, xlsx_files)


if __name__ == "__main__":
    unittest.main()
