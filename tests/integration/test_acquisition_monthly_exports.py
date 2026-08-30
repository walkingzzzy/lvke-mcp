from __future__ import annotations

import hashlib
import json
import os
import tempfile
import unittest
from pathlib import Path

from jsonschema import Draft202012Validator
from openpyxl import load_workbook

from lvke_mcp.domains.asset_acquisition import backend, resources, tables
from lvke_mcp.servers.lvke_asset_acquisition import server as acquisition_server


def _monthly_preview_spec() -> dict:
    return {
        "version": "finance_spec.v3",
        "finance_kind": "asset_acquisition",
        "invest_type": "asset_acquisition",
        "asset_type": "hotel_lease",
        "selected_scenario_id": "base",
        "confirmation_status": "confirmed",
        "delivery_mode": "estimate_preview",
        "controlled_assumptions": [{
            "field": "monthly_drivers",
            "value": "integration_fixture",
            "unit": "mixed",
            "basis": "deterministic_test_fixture",
            "impact": "monthly_model",
            "sensitivity": "covered_by_driver_tests",
            "validation_condition": "replace_with_formal_project_evidence",
        }],
        "transaction": {
            "calculation_granularity": "monthly",
            "model_start_date": "2026-01-01",
            "opening_date": "2026-01-01",
            "closing_date": "2026-01-01",
            "operating_mode": "mixed_owner_operator",
            "acquisition_type": "asset",
            "purchase_price": 1000.0,
            "valuation_value": 1000.0,
            "transaction_taxes": {},
            "financing_ratio": 0.4,
            "interest_rate": 0.05,
            "tenor": 1,
            "exit_year": 1,
            "repayment": "equal_principal",
            "asset_scope": [{
                "scope_id": "hotel-building",
                "type": "building",
                "name": "Hotel building",
                "included": True,
                "status": "confirmed",
                "accounting_treatment": "depreciable",
                "value_wan": 1000.0,
                "depreciable_basis_wan": 1000.0,
                "depreciation_years": 20,
                "residual_rate": 0.0,
                "evidence_ids": [],
            }],
        },
        "hotel_operation": {
            "rooms": 10,
            "adr": {"monthly_values": [200.0] * 12},
            "occupancy": {"monthly_values": [0.5] * 12},
            "ancillary_revenue": {
                "annual_values": 120.0,
                "seasonal_factors": [1.0] * 12,
            },
            "payroll": {"monthly_values": [2.0] * 12},
            "utilities": 12.0,
            "consumables": 24.0,
            "maintenance": {"monthly_values": [1.0] * 12},
            "operating_calendar": {
                "basis": "operating_days",
                "monthly_days": [20.0] * 12,
            },
        },
        "lease_portfolio": {"projection_years": 1, "units": []},
        "cost": {
            "annual_owner_operating_cost_wan": {
                "monthly_values": [1.0] * 12,
            },
        },
        "tax": {"income_tax_rate": 0.25, "vat_rate": 0.06, "surtax_rate": 0.12},
        "revenue": {"model": "flat"},
    }


class AcquisitionMonthlyExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory(prefix="lvke-acquisition-monthly-export-")
        self.previous_data = os.environ.get("LVKE_MCP_DATA_DIR")
        self.previous_deliverable = os.environ.get("LVKE_DELIVERABLE_DIR")
        os.environ["LVKE_MCP_DATA_DIR"] = self.tempdir.name
        os.environ["LVKE_DELIVERABLE_DIR"] = str(Path(self.tempdir.name) / "deliverables")
        self.workspace = "acquisition-monthly-export"
        run = backend.create_run(
            self.workspace,
            _monthly_preview_spec(),
            idempotency_key="monthly-export-run",
        )
        self.assertTrue(run.get("ok"), run)
        self.run = run
        rendered = tables.render(self.workspace, str(run["run_id"]))
        self.assertEqual(rendered.get("integrity", {}).get("status"), "passed", rendered)
        self.package_id = str(rendered["acquisition_tables_package_id"])

    def tearDown(self) -> None:
        for name, previous in (
            ("LVKE_MCP_DATA_DIR", self.previous_data),
            ("LVKE_DELIVERABLE_DIR", self.previous_deliverable),
        ):
            if previous is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = previous
        self.tempdir.cleanup()

    def test_public_schema_exposes_all_monthly_driver_forms(self) -> None:
        schema = acquisition_server.SERVER._tools["acquisition_validate_spec"].input_schema
        errors = list(Draft202012Validator(schema).iter_errors({"spec": _monthly_preview_spec()}))
        self.assertEqual(errors, [], [error.message for error in errors])

    def test_csv_export_writes_monthly_tables_and_verifiable_manifest(self) -> None:
        exported = tables.export_csv(self.workspace, self.package_id)
        directory = Path(exported["deliverable_path"])
        csv_files = sorted(directory.glob("*.csv"))
        self.assertEqual(len(csv_files), 17)
        self.assertTrue((directory / "monthly_income_statement.csv").is_file())
        self.assertTrue((directory / "monthly_balance_sheet.csv").is_file())
        manifest_path = directory / "manifest.json"
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest, exported["monthly_export_manifest"])
        self.assertEqual(manifest["package_id"], self.package_id)
        self.assertEqual(manifest["run_id"], self.run["run_id"])
        self.assertEqual(manifest["operating_calendar"]["manifest"]["month_count"], 12)
        self.assertEqual(len(manifest["annual_reconciliation"]), 14)
        self.assertTrue(all(row["status"] == "passed" for row in manifest["annual_reconciliation"]))
        self.assertEqual(manifest["monthly_driver_manifest"]["maintenance"]["source"], "explicit_monthly")
        for key, digest in manifest["files"].items():
            actual = "sha256:" + hashlib.sha256((directory / f"{key}.csv").read_bytes()).hexdigest()
            self.assertEqual(actual, digest)
        manifest_hash = "sha256:" + hashlib.sha256(manifest_path.read_bytes()).hexdigest()
        self.assertEqual(manifest_hash, exported["monthly_export_manifest_hash"])

    def test_xlsx_and_all_export_resources_are_readable(self) -> None:
        csv_export = tables.export_csv(self.workspace, self.package_id)
        xlsx_export = tables.export_xlsx(self.workspace, self.package_id)
        workbook_path = Path(xlsx_export["deliverable_path"])
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        try:
            self.assertEqual(len(workbook.sheetnames), 17)
            self.assertTrue(any("月度利润表" in name for name in workbook.sheetnames))
            self.assertTrue(any("月度资产负债表" in name for name in workbook.sheetnames))
        finally:
            workbook.close()

        for uri in [
            xlsx_export["xlsx_resource_uri"],
            xlsx_export["monthly_export_manifest_uri"],
            csv_export["monthly_export_manifest_uri"],
            next(uri for uri in csv_export["csv_resource_uris"] if uri.endswith("/monthly_income_statement")),
            next(uri for uri in csv_export["csv_resource_uris"] if uri.endswith("/monthly_balance_sheet")),
        ]:
            with self.subTest(uri=uri):
                read = resources.read_resource(self.workspace, uri)
                self.assertTrue(read["success"], read)
                self.assertTrue(read["content"])

        listed = resources.list_resources(self.workspace, limit=100)
        listed_uris = {row["uri"] for row in listed["resources"]}
        self.assertIn(xlsx_export["monthly_export_manifest_uri"], listed_uris)
        self.assertIn(csv_export["monthly_export_manifest_uri"], listed_uris)
        self.assertTrue(any(uri.endswith("/csv/monthly_income_statement") for uri in listed_uris))
        denied = resources.read_resource("other-workspace", xlsx_export["xlsx_resource_uri"])
        self.assertEqual(denied["code"], "RESOURCE_WORKSPACE_MISMATCH")


if __name__ == "__main__":
    unittest.main()
